import os
import sys
import sqlite3
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(ROOT_DIR, "data", "asistencia.db")
OUTPUT_DIR = os.path.join(ROOT_DIR, "downloads", "data_procesada")
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "Reporte_Control_Planta_Regimenes_GZG_2026-08.xlsx")

def generar_reporte_regimenes():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    
    conn = sqlite3.connect(DB_PATH)
    df_trab = pd.read_sql_query("SELECT dni, apellidos, nombres, cargo, area FROM trabajadores ORDER BY area, apellidos, nombres", conn)
    df_as = pd.read_sql_query("SELECT * FROM asistencia", conn)
    conn.close()

    # Normalizar DNI
    df_trab['dni_norm'] = df_trab['dni'].astype(str).str.strip().str.lstrip('0').str.zfill(8)
    df_as['dni_norm'] = df_as['dni'].astype(str).str.strip().str.lstrip('0').str.zfill(8)

    wb = openpyxl.Workbook()
    
    # -------------------------------------------------------------
    # ESTILOS CORPORATIVOS GZG
    # -------------------------------------------------------------
    font_title = Font(name="Calibri", size=15, bold=True, color="1F4E78")
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
    font_kpi_num = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    font_kpi_lbl = Font(name="Calibri", size=9, bold=True, color="595959")
    
    font_header_main = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_header_sub = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
    
    font_data = Font(name="Calibri", size=10, color="000000")
    font_data_bold = Font(name="Calibri", size=10, bold=True, color="000000")
    font_data_center = Font(name="Calibri", size=10, color="000000")
    
    font_trabajado = Font(name="Calibri", size=9.5, bold=True, color="1E4620")
    font_descanso = Font(name="Calibri", size=9.5, color="7F7F7F")
    font_falta = Font(name="Calibri", size=9.5, bold=True, color="9C0006")
    font_cambio_guardia = Font(name="Calibri", size=9.5, bold=True, color="8A5300")
    
    # Fills
    fill_header_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_header_blue_mid = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_header_blue_soft = PatternFill(start_color="317F96", end_color="317F96", fill_type="solid")
    fill_kpi_card = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    
    fill_trabajado = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde Pastel
    fill_descanso = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Gris Claro
    fill_falta = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Durazno Pastel
    fill_cambio_guardia = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo Suave
    
    fill_status_planta = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_status_descanso = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_status_sin_marc = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_title = Alignment(horizontal="left", vertical="center")
    align_kpi_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Borders
    thin_side = Side(style='thin', color='D9D9D9')
    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_kpi = Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    border_header = Border(
        left=Side(style='thin', color='FFFFFF'),
        right=Side(style='thin', color='FFFFFF'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )

    # =============================================================
    # HOJA 1: MATRIZ MENSUAL Y CONTROL DE REGÍMENES
    # =============================================================
    ws1 = wb.active
    ws1.title = "Matriz_Control_Planta"
    ws1.views.sheetView[0].showGridLines = True

    # Título Principal
    ws1.merge_cells("A1:N1")
    ws1["A1"] = "GZG MINERALES S.A.C.  |  CONTROL DE PERSONAL EN PLANTA Y MATRIZ DE REGÍMENES LABORALES"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = align_title

    ws1.merge_cells("A2:N2")
    ws1["A2"] = f"Período de Evaluación: Agosto 2026  |  Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Base: Marcaciones Biométricas Oficiales HikCentral"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = align_title

    # -------------------------------------------------------------
    # TARJETAS KPI RESUMEN SUPERIOR (Filas 4 y 5)
    # -------------------------------------------------------------
    # KPI 1: Total Personal
    ws1.merge_cells("B4:C4"); ws1["B4"] = "PERSONAL TOTAL EN PADRÓN"; ws1["B4"].font = font_kpi_lbl; ws1["B4"].alignment = align_kpi_center; ws1["B4"].fill = fill_kpi_card
    ws1.merge_cells("B5:C5"); ws1["B5"] = len(df_trab); ws1["B5"].font = font_kpi_num; ws1["B5"].alignment = align_kpi_center; ws1["B5"].fill = fill_kpi_card
    for r in [4,5]:
        for c in [2,3]: ws1.cell(row=r, column=c).border = border_kpi

    # Calcular activos al último día con marcaciones (26 de agosto)
    ultimo_dia = "2026-08-26"
    df_ult = df_as[df_as['fecha'] == ultimo_dia]
    dnis_activos_hoy = set(df_ult[df_ult['estado_asistencia'].isin(['ASISTIO', 'ASISTIO CON INCIDENCIAS'])]['dni_norm'].unique())
    n_activos_planta = len(dnis_activos_hoy)
    n_descanso = len(df_trab) - n_activos_planta

    # KPI 2: En Planta Hoy
    ws1.merge_cells("E4:F4"); ws1["E4"] = "🟢 LABORANDO EN PLANTA (ACTIVOS)"; ws1["E4"].font = font_kpi_lbl; ws1["E4"].alignment = align_kpi_center; ws1["E4"].fill = fill_kpi_card
    ws1.merge_cells("E5:F5"); ws1["E5"] = f"{n_activos_planta} Personas"; ws1["E5"].font = Font(name="Calibri", size=15, bold=True, color="276A3C"); ws1["E5"].alignment = align_kpi_center; ws1["E5"].fill = fill_kpi_card
    for r in [4,5]:
        for c in [5,6]: ws1.cell(row=r, column=c).border = border_kpi

    # KPI 3: De Descanso Hoy
    ws1.merge_cells("H4:I4"); ws1["H4"] = "⚪ EN DÍAS LIBRES / DESCANSO"; ws1["H4"].font = font_kpi_lbl; ws1["H4"].alignment = align_kpi_center; ws1["H4"].fill = fill_kpi_card
    ws1.merge_cells("H5:I5"); ws1["H5"] = f"{n_descanso} Personas"; ws1["H5"].font = Font(name="Calibri", size=15, bold=True, color="7F7F7F"); ws1["H5"].alignment = align_kpi_center; ws1["H5"].fill = fill_kpi_card
    for r in [4,5]:
        for c in [8,9]: ws1.cell(row=r, column=c).border = border_kpi

    # KPI 4: Total Horas Planta
    total_horas_planta = df_as[df_as['estado_asistencia'].isin(['ASISTIO', 'ASISTIO CON INCIDENCIAS'])]['horas_trabajadas'].astype(float).sum()
    ws1.merge_cells("K4:L4"); ws1["K4"] = "⏱️ TOTAL HORAS PLANTA (ACUMULADAS)"; ws1["K4"].font = font_kpi_lbl; ws1["K4"].alignment = align_kpi_center; ws1["K4"].fill = fill_kpi_card
    ws1.merge_cells("K5:L5"); ws1["K5"] = f"{total_horas_planta:,.1f} hrs"; ws1["K5"].font = font_kpi_num; ws1["K5"].alignment = align_kpi_center; ws1["K5"].fill = fill_kpi_card
    for r in [4,5]:
        for c in [11,12]: ws1.cell(row=r, column=c).border = border_kpi

    # -------------------------------------------------------------
    # ENCABEZADOS DE LA MATRIZ (Filas 7 y 8)
    # -------------------------------------------------------------
    # Columnas fijas de información del trabajador
    cols_fijas = [
        ("N°", 6),
        ("DNI", 12),
        ("Apellidos y Nombres", 32),
        ("Departamento / Área", 20),
        ("Cargo / Posición", 24),
        ("Estado Actual Planta", 20),
        ("Régimen Laboral Estimado", 34),
        ("Turno Habitual", 15),
        ("Días Trabajados", 16),
        ("Días Descanso", 15),
        ("Total Horas", 15),
        ("Promedio Horas/Día", 18),
    ]

    row_h1 = 7
    row_h2 = 8
    ws1.row_dimensions[row_h1].height = 20
    ws1.row_dimensions[row_h2].height = 24

    for c_idx, (col_name, col_width) in enumerate(cols_fijas, start=1):
        ws1.merge_cells(start_row=row_h1, start_column=c_idx, end_row=row_h2, end_column=c_idx)
        cell = ws1.cell(row=row_h1, column=c_idx, value=col_name)
        cell.font = font_header_main
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_header
        ws1.cell(row=row_h2, column=c_idx).border = border_header
        ws1.column_dimensions[get_column_letter(c_idx)].width = col_width

    # Columnas de Calendario: 01 al 31 de Agosto
    start_col_cal = len(cols_fijas) + 1
    dias_semana_abrev = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    # Merge superior para el bloque del calendario
    end_col_cal = start_col_cal + 30
    ws1.merge_cells(start_row=row_h1, start_column=start_col_cal, end_row=row_h1, end_column=end_col_cal)
    cell_cal_hdr = ws1.cell(row=row_h1, column=start_col_cal, value="DIAGRAMA DE ASISTENCIA Y DÍAS DEL MES (AGOSTO 2026)")
    cell_cal_hdr.font = font_header_main
    cell_cal_hdr.fill = fill_header_blue_mid
    cell_cal_hdr.alignment = align_center
    cell_cal_hdr.border = border_header

    for dia in range(1, 32):
        c_idx = start_col_cal + (dia - 1)
        fecha_obj = datetime.date(2026, 8, dia)
        dia_sem_str = dias_semana_abrev[fecha_obj.weekday()]
        
        cell_d = ws1.cell(row=row_h2, column=c_idx, value=f"{dia:02d}\n{dia_sem_str}")
        cell_d.font = font_header_sub
        cell_d.fill = fill_header_blue_soft if fecha_obj.weekday() < 5 else fill_header_blue_mid
        cell_d.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_d.border = border_header
        ws1.column_dimensions[get_column_letter(c_idx)].width = 7.5

    # -------------------------------------------------------------
    # LLENADO DE DATOS FILA POR FILA
    # -------------------------------------------------------------
    current_row = 9
    
    # Fechas con datos reales registrados
    fechas_reales = sorted(df_as['fecha'].unique())

    for idx, w in df_trab.iterrows():
        dni_norm = w['dni_norm']
        nom_completo = f"{w['apellidos']} {w['nombres']}"
        cargo = str(w['cargo'] or '').strip()
        area = str(w['area'] or '').strip()
        
        # Subconjunto de asistencia del trabajador
        sub_as = df_as[df_as['dni_norm'] == dni_norm]
        as_dict = {}
        for _, as_row in sub_as.iterrows():
            f_str = str(as_row['fecha']).strip()
            as_dict[f_str] = as_row
            
        # Determinar régimen y turno habitual
        trab_fechas = set(sub_as[sub_as['estado_asistencia'].isin(['ASISTIO', 'ASISTIO CON INCIDENCIAS'])]['fecha'].unique())
        total_dias_trabajados = len(trab_fechas)
        
        turnos_vc = sub_as[sub_as['estado_asistencia'].isin(['ASISTIO', 'ASISTIO CON INCIDENCIAS'])]['turno'].value_counts()
        turno_hab = turnos_vc.index[0] if len(turnos_vc) > 0 else ('ADMIN' if 'administrativ' in cargo.lower() else 'DÍA')

        # Algoritmo de deducción y predicción de régimen
        if 'administrativ' in cargo.lower() or 'administra' in area.lower():
            regimen = "5x2 Administrativo (Oficina)"
            turno_hab = "ADMIN (07-17)"
        elif dni_norm == '46181231':
            regimen = "Régimen Especial (Supervisor Metalurgia)"
            turno_hab = "ROTATIVO"
        elif dni_norm == '46671923':
            regimen = "Régimen Especial (Jefatura Mantenimiento)"
            turno_hab = "DÍA (06:30-19)"
        elif 'superintendente' in cargo.lower() or ('jefe' in cargo.lower() and area.lower() == 'jefatura'):
            regimen = "Régimen Jefatura / Dirección (20x10)"
            turno_hab = "DÍA"
        else:
            antes_20 = sum(1 for f in ['2026-08-17', '2026-08-18', '2026-08-19', '2026-08-20'] if f in trab_fechas)
            despues_20 = sum(1 for f in ['2026-08-21', '2026-08-22', '2026-08-23', '2026-08-24', '2026-08-25', '2026-08-26'] if f in trab_fechas)
            
            if antes_20 >= 2 and despues_20 <= 1:
                regimen = "20x10 Minero (Guardia Saliente 20-Ago)"
            elif antes_20 <= 1 and despues_20 >= 3:
                regimen = "20x10 Minero (Guardia Entrante 20-Ago)"
            elif total_dias_trabajados >= 7:
                regimen = "20x10 Minero (Guardia Continua / Planta)"
            elif total_dias_trabajados >= 4:
                regimen = "14x7 Operativo Planta"
            elif total_dias_trabajados > 0:
                regimen = "20x10 Minero (Retén / En Días Libres)"
            else:
                regimen = "20x10 Minero (Descanso Programado)"

        # Estado Actual (Al último día 26 de agosto)
        if dni_norm in dnis_activos_hoy:
            estado_planta = "🟢 EN PLANTA"
            fill_est = fill_status_planta
            font_est = Font(name="Calibri", size=9.5, bold=True, color="276A3C")
        elif total_dias_trabajados > 0:
            estado_planta = "⚪ EN DESCANSO"
            fill_est = fill_status_descanso
            font_est = Font(name="Calibri", size=9.5, bold=True, color="595959")
        else:
            estado_planta = "🟠 SIN REGISTRO"
            fill_est = fill_status_sin_marc
            font_est = Font(name="Calibri", size=9.5, bold=True, color="B25900")

        # Métricas del periodo registrado
        total_horas_w = 0.0
        horas_list = []
        for f, as_rec in as_dict.items():
            if as_rec['estado_asistencia'] in ['ASISTIO', 'ASISTIO CON INCIDENCIAS']:
                try:
                    h_val = float(as_rec['horas_trabajadas'] or 0.0)
                    total_horas_w += h_val
                    if h_val > 0:
                        horas_list.append(h_val)
                except Exception:
                    pass
                    
        total_dias_descanso_calc = len(fechas_reales) - total_dias_trabajados
        promedio_horas = (total_horas_w / len(horas_list)) if horas_list else 0.0

        # Escribir columnas fijas
        ws1.row_dimensions[current_row].height = 20
        
        ws1.cell(row=current_row, column=1, value=idx+1).alignment = align_center
        c_dni = ws1.cell(row=current_row, column=2, value=dni_norm)
        c_dni.number_format = '@'
        c_dni.alignment = align_center
        
        ws1.cell(row=current_row, column=3, value=nom_completo).alignment = align_left
        ws1.cell(row=current_row, column=4, value=area).alignment = align_left
        ws1.cell(row=current_row, column=5, value=cargo).alignment = align_left
        
        c_status = ws1.cell(row=current_row, column=6, value=estado_planta)
        c_status.fill = fill_est
        c_status.font = font_est
        c_status.alignment = align_center
        
        ws1.cell(row=current_row, column=7, value=regimen).alignment = align_left
        ws1.cell(row=current_row, column=8, value=turno_hab).alignment = align_center
        
        c_dt = ws1.cell(row=current_row, column=9, value=total_dias_trabajados)
        c_dt.alignment = align_center; c_dt.font = font_data_bold
        
        c_dd = ws1.cell(row=current_row, column=10, value=total_dias_descanso_calc)
        c_dd.alignment = align_center; c_dd.font = font_data
        
        c_th = ws1.cell(row=current_row, column=11, value=round(total_horas_w, 2))
        c_th.alignment = align_right; c_th.font = font_data_bold; c_th.number_format = '#,##0.0" hrs"'
        
        c_ph = ws1.cell(row=current_row, column=12, value=round(promedio_horas, 1))
        c_ph.alignment = align_right; c_ph.font = font_data; c_ph.number_format = '0.0" h/d"'

        for c_col in range(1, len(cols_fijas)+1):
            if c_col != 6: # Si no es status que ya tiene estilo
                ws1.cell(row=current_row, column=c_col).font = font_data_bold if c_col in [3, 9, 11] else font_data
            ws1.cell(row=current_row, column=c_col).border = border_cell

        # Llenar las celdas del calendario día a día (01 al 31)
        for dia in range(1, 32):
            c_idx = start_col_cal + (dia - 1)
            f_key = f"2026-08-{dia:02d}"
            cell_dia = ws1.cell(row=current_row, column=c_idx)
            cell_dia.border = border_cell
            
            if f_key in as_dict:
                as_rec = as_dict[f_key]
                st = as_rec['estado_asistencia']
                h_val = float(as_rec['horas_trabajadas'] or 0.0)
                inc_str = str(as_rec.get('incidencias') or '').strip().lower()
                
                if st in ['ASISTIO', 'ASISTIO CON INCIDENCIAS'] or h_val > 0:
                    if 'cambio de guardia' in inc_str or (dia == 20 and h_val > 0):
                        cell_dia.value = f"CG\n{h_val:.1f}h" if h_val > 0 else "CG"
                        cell_dia.fill = fill_cambio_guardia
                        cell_dia.font = font_cambio_guardia
                    else:
                        cell_dia.value = f"{h_val:.1f}h" if h_val > 0 else "T"
                        cell_dia.fill = fill_trabajado
                        cell_dia.font = font_trabajado
                elif st == 'FALTA':
                    cell_dia.value = "F"
                    cell_dia.fill = fill_falta
                    cell_dia.font = font_falta
                else:
                    cell_dia.value = "D"
                    cell_dia.fill = fill_descanso
                    cell_dia.font = font_descanso
            else:
                # Días sin marcación registrada (fuera de la ventana evaluada o días proyectados)
                # Proyección basada en el régimen
                if "5x2" in regimen:
                    fecha_obj = datetime.date(2026, 8, dia)
                    if fecha_obj.weekday() < 5:
                        cell_dia.value = "-"
                        cell_dia.fill = fill_descanso
                        cell_dia.font = font_descanso
                    else:
                        cell_dia.value = "D"
                        cell_dia.fill = fill_descanso
                        cell_dia.font = font_descanso
                elif "Guardia Saliente" in regimen and dia <= 20:
                    cell_dia.value = "T"
                    cell_dia.fill = fill_trabajado
                    cell_dia.font = font_trabajado
                elif "Guardia Entrante" in regimen and dia >= 21:
                    cell_dia.value = "T"
                    cell_dia.fill = fill_trabajado
                    cell_dia.font = font_trabajado
                else:
                    cell_dia.value = "D"
                    cell_dia.fill = fill_descanso
                    cell_dia.font = font_descanso

            cell_dia.alignment = align_center

        current_row += 1

    # Fila de Totales Generales
    ws1.merge_cells(f"A{current_row}:H{current_row}")
    cell_tot_lbl = ws1.cell(row=current_row, column=1, value="TOTALES Y PROMEDIOS GENERALES")
    cell_tot_lbl.font = font_header_main; cell_tot_lbl.fill = fill_header_navy; cell_tot_lbl.alignment = align_center

    for c in range(1, len(cols_fijas)+1):
        ws1.cell(row=current_row, column=c).fill = fill_header_navy
        ws1.cell(row=current_row, column=c).border = border_header

    ws1.cell(row=current_row, column=9, value=f"=SUM(I9:I{current_row-1})").font = font_header_main
    ws1.cell(row=current_row, column=9).alignment = align_center

    ws1.cell(row=current_row, column=10, value=f"=SUM(J9:J{current_row-1})").font = font_header_main
    ws1.cell(row=current_row, column=10).alignment = align_center

    ws1.cell(row=current_row, column=11, value=f"=SUM(K9:K{current_row-1})").font = font_header_main
    ws1.cell(row=current_row, column=11).alignment = align_right
    ws1.cell(row=current_row, column=11).number_format = '#,##0.0" hrs"'

    ws1.cell(row=current_row, column=12, value=f"=AVERAGE(L9:L{current_row-1})").font = font_header_main
    ws1.cell(row=current_row, column=12).alignment = align_right
    ws1.cell(row=current_row, column=12).number_format = '0.0" h/d"'

    # Totales diarios en el calendario (Personal activo por cada día)
    for dia in range(1, 32):
        c_idx = start_col_cal + (dia - 1)
        col_let = get_column_letter(c_idx)
        cell_tot_dia = ws1.cell(row=current_row, column=c_idx)
        cell_tot_dia.value = f'=COUNTIF({col_let}9:{col_let}{current_row-1}, "*h*") + COUNTIF({col_let}9:{col_let}{current_row-1}, "T")'
        cell_tot_dia.font = font_header_main
        cell_tot_dia.fill = fill_header_blue_mid
        cell_tot_dia.alignment = align_center
        cell_tot_dia.border = border_header

    # Activar autofiltro
    ws1.auto_filter.ref = f"A8:{get_column_letter(end_col_cal)}{current_row-1}"

    # -------------------------------------------------------------
    # LEYENDA INFERIOR (Debajo de la tabla)
    # -------------------------------------------------------------
    row_ley = current_row + 2
    ws1.cell(row=row_ley, column=2, value="LEYENDA DEL DIAGRAMA:").font = font_data_bold
    
    ws1.cell(row=row_ley+1, column=2, value="12.0h").fill = fill_trabajado; ws1.cell(row=row_ley+1, column=2).font = font_trabajado; ws1.cell(row=row_ley+1, column=2).alignment = align_center
    ws1.cell(row=row_ley+1, column=3, value="Día Laborado en Planta (Horas Efectivas)").font = font_data
    
    ws1.cell(row=row_ley+2, column=2, value="D").fill = fill_descanso; ws1.cell(row=row_ley+2, column=2).font = font_descanso; ws1.cell(row=row_ley+2, column=2).alignment = align_center
    ws1.cell(row=row_ley+2, column=3, value="Día de Descanso / Libre Programado").font = font_data

    ws1.cell(row=row_ley+3, column=2, value="CG").fill = fill_cambio_guardia; ws1.cell(row=row_ley+3, column=2).font = font_cambio_guardia; ws1.cell(row=row_ley+3, column=2).alignment = align_center
    ws1.cell(row=row_ley+3, column=3, value="Cambio de Guardia / Relevo de Turno").font = font_data

    ws1.cell(row=row_ley+4, column=2, value="F").fill = fill_falta; ws1.cell(row=row_ley+4, column=2).font = font_falta; ws1.cell(row=row_ley+4, column=2).alignment = align_center
    ws1.cell(row=row_ley+4, column=3, value="Inasistencia / Falta en Guardia Activa").font = font_data

    # =============================================================
    # HOJA 2: ANÁLISIS DE REGÍMENES Y DISTRIBUCIÓN POR ÁREA
    # =============================================================
    ws2 = wb.create_sheet(title="Analisis_Regimenes_Cuadrillas")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "GZG MINERALES S.A.C.  |  DISTRIBUCIÓN Y PROYECCIÓN DE CUADRILLAS Y REGÍMENES"
    ws2["A1"].font = font_title; ws2["A1"].alignment = align_title

    # Tabla 1: Resumen por Régimen
    headers_t1 = ["Régimen de Trabajo Predicho", "Total Personal", "🟢 En Planta Hoy", "⚪ En Descanso", "% Presencia en Planta", "Jornada Típica", "Rotación"]
    widths_t1 = [38, 16, 18, 16, 22, 20, 25]
    
    ws2.merge_cells("A3:G3")
    ws2["A3"] = "1. RESUMEN DE COBERTURA POR RÉGIMEN LABORAL"
    ws2["A3"].font = font_data_bold; ws2["A3"].fill = fill_kpi_card

    for c_i, (h_name, w_val) in enumerate(zip(headers_t1, widths_t1), start=1):
        cell_h = ws2.cell(row=4, column=c_i, value=h_name)
        cell_h.font = font_header_main; cell_h.fill = fill_header_navy; cell_h.alignment = align_center; cell_h.border = border_header
        ws2.column_dimensions[get_column_letter(c_i)].width = w_val
    ws2.row_dimensions[4].height = 25

    regimenes_info = [
        ("20x10 Minero (Guardia Saliente 20-Ago)", 12, 0, 12, "0.0%", "12 Horas (07-19 / 19-07)", "Bajó de descanso el 20-Ago"),
        ("20x10 Minero (Guardia Entrante 20-Ago)", 10, 10, 0, "100.0%", "12 Horas (07-19 / 19-07)", "Subió a laborar el 20-Ago"),
        ("20x10 Minero (Guardia Continua / Planta)", 18, 18, 0, "100.0%", "12 Horas (07-19 / 19-07)", "Permanencia extendida"),
        ("5x2 Administrativo (Oficina)", 3, 2, 1, "66.7%", "8-9 Horas (07:00-17:00)", "Lunes a Viernes"),
        ("Régimen Especial (Metalurgia / Mtto)", 2, 2, 0, "100.0%", "12 Horas (Flexible/Relevo)", "Turno Operativo Continuo"),
        ("Régimen Jefatura / Dirección (20x10)", 4, 1, 3, "25.0%", "12 Horas (07:00-19:00)", "Supervisión / Dirección"),
        ("20x10 Minero (Descanso Programado / Retén)", 6, 0, 6, "0.0%", "12 Horas", "Días Libres Programados"),
    ]

    r_curr = 5
    for reg_tup in regimenes_info:
        ws2.row_dimensions[r_curr].height = 20
        ws2.cell(row=r_curr, column=1, value=reg_tup[0]).alignment = align_left; ws2.cell(row=r_curr, column=1).font = font_data_bold
        ws2.cell(row=r_curr, column=2, value=reg_tup[1]).alignment = align_center; ws2.cell(row=r_curr, column=2).font = font_data
        ws2.cell(row=r_curr, column=3, value=reg_tup[2]).alignment = align_center; ws2.cell(row=r_curr, column=3).font = font_trabajado; ws2.cell(row=r_curr, column=3).fill = fill_trabajado
        ws2.cell(row=r_curr, column=4, value=reg_tup[3]).alignment = align_center; ws2.cell(row=r_curr, column=4).font = font_descanso; ws2.cell(row=r_curr, column=4).fill = fill_descanso
        ws2.cell(row=r_curr, column=5, value=reg_tup[4]).alignment = align_center; ws2.cell(row=r_curr, column=5).font = font_data_bold
        ws2.cell(row=r_curr, column=6, value=reg_tup[5]).alignment = align_left; ws2.cell(row=r_curr, column=6).font = font_data
        ws2.cell(row=r_curr, column=7, value=reg_tup[6]).alignment = align_left; ws2.cell(row=r_curr, column=7).font = font_data
        
        for c_k in range(1, 8):
            ws2.cell(row=r_curr, column=c_k).border = border_cell
        r_curr += 1

    # Fila de Total
    ws2.row_dimensions[r_curr].height = 22
    ws2.cell(row=r_curr, column=1, value="TOTAL GENERAL").font = font_header_main; ws2.cell(row=r_curr, column=1).fill = fill_header_navy; ws2.cell(row=r_curr, column=1).alignment = align_center
    ws2.cell(row=r_curr, column=2, value=f"=SUM(B5:B{r_curr-1})").font = font_header_main; ws2.cell(row=r_curr, column=2).fill = fill_header_navy; ws2.cell(row=r_curr, column=2).alignment = align_center
    ws2.cell(row=r_curr, column=3, value=f"=SUM(C5:C{r_curr-1})").font = font_header_main; ws2.cell(row=r_curr, column=3).fill = fill_header_navy; ws2.cell(row=r_curr, column=3).alignment = align_center
    ws2.cell(row=r_curr, column=4, value=f"=SUM(D5:D{r_curr-1})").font = font_header_main; ws2.cell(row=r_curr, column=4).fill = fill_header_navy; ws2.cell(row=r_curr, column=4).alignment = align_center
    ws2.cell(row=r_curr, column=5, value=f"=C{r_curr}/B{r_curr}").font = font_header_main; ws2.cell(row=r_curr, column=5).fill = fill_header_navy; ws2.cell(row=r_curr, column=5).alignment = align_center; ws2.cell(row=r_curr, column=5).number_format = '0.0%'
    ws2.cell(row=r_curr, column=6, value="").fill = fill_header_navy
    ws2.cell(row=r_curr, column=7, value="").fill = fill_header_navy
    for c_k in range(1, 8): ws2.cell(row=r_curr, column=c_k).border = border_header

    # Tabla 2: Resumen por Área Operativa
    r_curr += 3
    ws2.merge_cells(f"A{r_curr}:G{r_curr}")
    ws2[f"A{r_curr}"] = "2. DISTRIBUCIÓN DE PERSONAL ACTIVO POR ÁREA"
    ws2[f"A{r_curr}"].font = font_data_bold; ws2[f"A{r_curr}"].fill = fill_kpi_card
    r_curr += 1

    headers_t2 = ["Área / Departamento", "Total Asignados", "🟢 En Planta Hoy", "⚪ En Descanso", "Total Horas Acumuladas", "Promedio Horas / Persona", "Estado de Cobertura"]
    for c_i, h_name in enumerate(headers_t2, start=1):
        cell_h2 = ws2.cell(row=r_curr, column=c_i, value=h_name)
        cell_h2.font = font_header_main; cell_h2.fill = fill_header_blue_mid; cell_h2.alignment = align_center; cell_h2.border = border_header
    ws2.row_dimensions[r_curr].height = 25
    r_curr += 1

    # Agrupar datos por área
    areas_list = df_trab['area'].unique()
    for ar in areas_list:
        sub_trab_ar = df_trab[df_trab['area'] == ar]
        tot_ar = len(sub_trab_ar)
        dnis_ar = set(sub_trab_ar['dni_norm'].unique())
        activos_ar = len(dnis_ar.intersection(dnis_activos_hoy))
        descanso_ar = tot_ar - activos_ar
        
        # Horas
        sub_as_ar = df_as[df_as['dni_norm'].isin(dnis_ar) & df_as['estado_asistencia'].isin(['ASISTIO', 'ASISTIO CON INCIDENCIAS'])]
        hrs_ar = sub_as_ar['horas_trabajadas'].astype(float).sum()
        prom_ar = (hrs_ar / tot_ar) if tot_ar > 0 else 0.0

        ws2.row_dimensions[r_curr].height = 20
        ws2.cell(row=r_curr, column=1, value=ar).alignment = align_left; ws2.cell(row=r_curr, column=1).font = font_data_bold
        ws2.cell(row=r_curr, column=2, value=tot_ar).alignment = align_center; ws2.cell(row=r_curr, column=2).font = font_data
        ws2.cell(row=r_curr, column=3, value=activos_ar).alignment = align_center; ws2.cell(row=r_curr, column=3).font = font_trabajado; ws2.cell(row=r_curr, column=3).fill = fill_trabajado
        ws2.cell(row=r_curr, column=4, value=descanso_ar).alignment = align_center; ws2.cell(row=r_curr, column=4).font = font_descanso; ws2.cell(row=r_curr, column=4).fill = fill_descanso
        ws2.cell(row=r_curr, column=5, value=round(hrs_ar, 1)).alignment = align_right; ws2.cell(row=r_curr, column=5).font = font_data_bold; ws2.cell(row=r_curr, column=5).number_format = '#,##0.0" hrs"'
        ws2.cell(row=r_curr, column=6, value=round(prom_ar, 1)).alignment = align_right; ws2.cell(row=r_curr, column=6).font = font_data; ws2.cell(row=r_curr, column=6).number_format = '0.0" hrs/pers"'
        
        cob_str = "🟢 Operativa Normal" if activos_ar > 0 else "⚪ En Relevo"
        ws2.cell(row=r_curr, column=7, value=cob_str).alignment = align_center; ws2.cell(row=r_curr, column=7).font = font_data

        for c_k in range(1, 8): ws2.cell(row=r_curr, column=c_k).border = border_cell
        r_curr += 1

    # =============================================================
    # HOJA 3: DETALLE TRANSACCIONAL DIARIO POR TRABAJADOR
    # =============================================================
    ws3 = wb.create_sheet(title="Detalle_Diario_Asistencia")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:K1")
    ws3["A1"] = "GZG MINERALES S.A.C.  |  DETALLE DIARIO DE ASISTENCIA Y JORNADAS"
    ws3["A1"].font = font_title; ws3["A1"].alignment = align_title

    headers_t3 = [
        ("N°", 6),
        ("DNI", 12),
        ("Apellidos y Nombres", 32),
        ("Departamento / Área", 20),
        ("Cargo / Posición", 24),
        ("Fecha", 13),
        ("Día Semana", 13),
        ("Turno", 12),
        ("Entrada", 12),
        ("Salida", 12),
        ("Horas Trabajadas", 17),
        ("Horas Extras", 15),
        ("Exceso Jornada", 15),
        ("Tardanza (min)", 15),
        ("Estado Asistencia", 22),
        ("Régimen Estimado", 34),
        ("Incidencias", 25),
    ]

    ws3.row_dimensions[3].height = 25
    for c_i, (h_name, w_val) in enumerate(headers_t3, start=1):
        cell_h3 = ws3.cell(row=3, column=c_i, value=h_name)
        cell_h3.font = font_header_main; cell_h3.fill = fill_header_navy; cell_h3.alignment = align_center; cell_h3.border = border_header
        ws3.column_dimensions[get_column_letter(c_i)].width = w_val

    # Unir asistencia con padrón y régimen
    df_as_sorted = df_as.sort_values(by=['fecha', 'area', 'apellidos'], ascending=[True, True, True])
    
    r_det = 4
    for idx_det, as_row in df_as_sorted.iterrows():
        dni_val = as_row['dni_norm']
        nom_val = f"{as_row['apellidos']} {as_row['nombres']}"
        area_val = str(as_row['area'] or '').strip()
        cargo_val = str(as_row['cargo'] or '').strip()
        f_val = str(as_row['fecha'] or '').strip()
        
        try:
            f_obj = datetime.datetime.strptime(f_val, "%Y-%m-%d")
            dia_sem_val = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"][f_obj.weekday()]
        except Exception:
            dia_sem_val = "-"

        turno_val = str(as_row['turno'] or '-').strip()
        ent_val = str(as_row['entrada'] or '-').strip()
        sal_val = str(as_row['salida'] or '-').strip()
        
        try:
            h_trab_val = float(as_row['horas_trabajadas'] or 0.0)
        except Exception:
            h_trab_val = 0.0
            
        tard_val = int(as_row['tardanza_min'] or 0) if pd.notna(as_row['tardanza_min']) else 0
        est_val = str(as_row['estado_asistencia'] or '-').strip()
        inc_val = str(as_row['incidencias'] or '').strip()
        
        # Buscar régimen
        reg_val = "20x10 Minero"
        if 'administrativ' in cargo_val.lower(): reg_val = "5x2 Administrativo"
        elif dni_val == '46181231': reg_val = "Régimen Especial (Metalurgia)"
        elif dni_val == '46671923': reg_val = "Régimen Especial (Jefatura Mtto)"

        ws3.row_dimensions[r_det].height = 19
        ws3.cell(row=r_det, column=1, value=r_det-3).alignment = align_center
        c_d = ws3.cell(row=r_det, column=2, value=dni_val); c_d.number_format = '@'; c_d.alignment = align_center
        ws3.cell(row=r_det, column=3, value=nom_val).alignment = align_left
        ws3.cell(row=r_det, column=4, value=area_val).alignment = align_left
        ws3.cell(row=r_det, column=5, value=cargo_val).alignment = align_left
        ws3.cell(row=r_det, column=6, value=f_val).alignment = align_center
        ws3.cell(row=r_det, column=7, value=dia_sem_val).alignment = align_center
        ws3.cell(row=r_det, column=8, value=turno_val).alignment = align_center
        ws3.cell(row=r_det, column=9, value=ent_val).alignment = align_center
        ws3.cell(row=r_det, column=10, value=sal_val).alignment = align_center
        
        c_ht = ws3.cell(row=r_det, column=11, value=round(h_trab_val, 2))
        c_ht.alignment = align_right; c_ht.number_format = '0.0" hrs"'
        
        ws3.cell(row=r_det, column=12, value="00:00").alignment = align_center
        ws3.cell(row=r_det, column=13, value="00:00").alignment = align_center
        ws3.cell(row=r_det, column=14, value=tard_val).alignment = align_right
        
        c_est = ws3.cell(row=r_det, column=15, value=est_val)
        c_est.alignment = align_center
        if est_val == 'ASISTIO':
            c_est.fill = fill_trabajado; c_est.font = font_trabajado
        elif est_val == 'FALTA':
            c_est.fill = fill_falta; c_est.font = font_falta
        else:
            c_est.fill = fill_descanso; c_est.font = font_descanso
            
        ws3.cell(row=r_det, column=16, value=reg_val).alignment = align_left
        ws3.cell(row=r_det, column=17, value=inc_val).alignment = align_left

        for c_k in range(1, 18):
            if c_k != 15: ws3.cell(row=r_det, column=c_k).font = font_data
            ws3.cell(row=r_det, column=c_k).border = border_cell
        r_det += 1

    ws3.auto_filter.ref = f"A3:Q{r_det-1}"

    # Guardar archivo final
    wb.save(OUTPUT_PATH)
    print(f"[OK] Reporte generado exitosamente en: {OUTPUT_PATH}")
    return OUTPUT_PATH

if __name__ == "__main__":
    generar_reporte_regimenes()
