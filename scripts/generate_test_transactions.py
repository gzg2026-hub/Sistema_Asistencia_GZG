import os
import random
import openpyxl
import pandas as pd
from datetime import datetime, timedelta, time

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "descargas_biometrico")
BASE_EXCEL = r"C:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Sistema_Asistencia_GZG_v1.0.xlsm"

def generar_lote_pruebas(
    start_date_str: str = "2026-08-01",
    end_date_str: str = "2026-08-11",
    target_dir: str = OUTPUT_DIR
):
    """
    Genera entre 200 y 300 transacciones de prueba en formato de exportación Hikvision,
    cubriendo el mes de agosto y todos los escenarios de incidencias y asistencia.
    """
    os.makedirs(target_dir, exist_ok=True)
    
    # Cargar lista de trabajadores maestros
    df_trab = pd.DataFrame()
    if os.path.exists(BASE_EXCEL):
        wb_base = openpyxl.load_workbook(BASE_EXCEL, data_only=True)
        if '01_TRABAJADORES' in wb_base.sheetnames:
            ws = wb_base['01_TRABAJADORES']
            data = list(ws.iter_rows(values_only=True))
            if data and len(data) > 1:
                df_trab = pd.DataFrame(data[1:], columns=[str(c).strip() for c in data[0]])
                
    if df_trab.empty:
        # Fallback de trabajadores si no se encuentra el Excel base
        df_trab = pd.DataFrame([
            {'DNI': '72559194', 'APELLIDOS': 'DE LA RIVA AGUILAR', 'NOMBRES': 'JAVIER ADRIAN', 'CARGO': 'SUPERVISOR', 'AREA': 'JEFATURA'},
            {'DNI': '18074244', 'APELLIDOS': 'LAZARO CRUZ', 'NOMBRES': 'RAUL RICHARD', 'CARGO': 'MAQUINARIA PESADA', 'AREA': 'OPER&MTTO'},
            {'DNI': '18861684', 'APELLIDOS': 'ZAMORA CARRASCO', 'NOMBRES': 'JACOB', 'CARGO': 'ELECTRICIDAD', 'AREA': 'OPER&MTTO'},
            {'DNI': '47034929', 'APELLIDOS': 'ALVA MEDINA', 'NOMBRES': 'JHON KENEDY', 'CARGO': 'JEFE', 'AREA': 'JEFATURA'},
            {'DNI': '44955960', 'APELLIDOS': 'ESPINOZA SAAVEDRA', 'NOMBRES': 'RAUL ESTEBAN', 'CARGO': 'OPERADOR', 'AREA': 'OPER&MTTO'},
        ])
        
    trabajadores = df_trab.to_dict(orient='records')
    
    dt_start = datetime.strptime(start_date_str, "%Y-%m-%d")
    dt_end = datetime.strptime(end_date_str, "%Y-%m-%d")
    
    current_dt = dt_start
    all_transactions = []
    
    dias_semana_es = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }

    # Bucle por día
    while current_dt <= dt_end:
        fecha_str = current_dt.strftime("%Y-%m-%d")
        semana_str = dias_semana_es[current_dt.strftime("%A")]
        
        # Seleccionar una muestra de trabajadores para el día
        for w in trabajadores:
            dni = str(w['DNI']).strip()
            nombre = str(w['NOMBRES']).strip()
            apellido = str(w['APELLIDOS']).strip()
            cargo = str(w.get('CARGO', 'OPERADOR')).strip()
            dept = str(w.get('AREA', 'OPER&MTTO')).strip()
            if '>' in dept:
                dept = dept.split('>')[-1].strip()
            
            # Asignar escenario aleatorio:
            # 1: Asistencia Normal Día
            # 2: Entrada Anticipada Día
            # 3: Tardanza Día (> 10 min)
            # 4: Salida Anticipada Día (< 10 min antes)
            # 5: Exceso de Jornada Día
            # 6: Horas Extra Explícitas Día
            # 7: Turno Noche Normal
            # 8: Entrada Duplicada (Incidencia)
            # 9: Salida Pendiente (Sin marca de salida)
            # 10: Falta (Sin marcas en el día)
            
            escenario = random.choices(
                range(1, 11),
                weights=[30, 10, 15, 10, 10, 10, 10, 3, 2, 0] # Probabilidades
            )[0]
            
            if escenario == 10:
                continue # Falta, no genera transacciones
                
            puntos_control = ["Garita Biometrico_red-1", "Garita Biometrico_wifi-1"]
            metodos = ["Imagen de cara", "Huella dactilar", "Tarjeta de proximidad"]
            
            if escenario == 1: # Normal Día
                t_ent = time(6, random.randint(50, 59)) if random.random() > 0.5 else time(7, random.randint(0, 8))
                t_sal = time(18, random.randint(52, 59)) if random.random() > 0.5 else time(19, random.randint(0, 5))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 2: # Entrada Anticipada (06:35 AM)
                t_ent = time(6, random.randint(30, 45))
                t_sal = time(19, random.randint(0, 5))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 3: # Tardanza (> 10 min: 07:18 AM)
                t_ent = time(7, random.randint(15, 40))
                t_sal = time(19, random.randint(0, 5))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 4: # Salida Anticipada (< 10 min antes: 18:25 PM)
                t_ent = time(7, 0)
                t_sal = time(18, random.randint(15, 45))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 5: # Exceso de Jornada (19:40 PM)
                t_ent = time(7, 0)
                t_sal = time(19, random.randint(20, 55))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 6: # Horas Extra Explícitas
                t_ent = time(7, 0)
                t_sal = time(19, 0)
                t_he_ini = time(19, 10)
                t_he_fin = time(21, random.randint(0, 30))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_he_ini.strftime("%H:%M"), "Inicio de horas extra", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_he_fin.strftime("%H:%M"), "Fin de horas extra", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 7: # Turno Noche
                t_ent = time(18, random.randint(50, 59))
                t_sal = time(7, random.randint(0, 5))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 8: # Marcación Duplicada
                t_ent1 = time(7, 2)
                t_ent2 = time(7, 5)
                t_sal = time(19, 0)
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent1.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent2.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_sal.strftime("%H:%M"), "Registrar salida", random.choice(metodos), random.choice(puntos_control)))

            elif escenario == 9: # Salida Pendiente
                t_ent = time(7, 0)
                all_transactions.append((fecha_str, semana_str, dni, nombre, apellido, dept, t_ent.strftime("%H:%M"), "Registro de entrada", random.choice(metodos), random.choice(puntos_control)))

        current_dt += timedelta(days=1)

    # Crear libro de Excel simulando el formato de exportación Hikvision
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Escribir metadatos superiores
    ws.append([])
    ws.append([])
    ws.append([])
    ws.append(["Transacciones-Calcular por fecha"])
    ws.append([f"Hora de exportación: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append(["Operador: admin"])
    ws.append([f"Periodo: {start_date_str} - {end_date_str}"])
    
    # Agrupar transacciones por fecha y escribirlas
    df_tx = pd.DataFrame(all_transactions, columns=['Fecha', 'Semana', 'ID', 'Nombre', 'Apellido', 'Departamento', 'Tiempo', 'Tipo de pase de tarjeta', 'Método de verificación', 'Punto de control de asistencia'])
    
    grouped_tx = df_tx.groupby('Fecha')
    
    for fecha, group in grouped_tx:
        semana = group.iloc[0]['Semana']
        ws.append([f"Fecha:{fecha} Semana:{semana}"])
        ws.append(['ID', 'Fecha', 'Nombre', 'Apellido', 'Departamento', 'Grupo de asistencia', 'Tiempo', 'Tipo de pase de tarjeta', 'Método de verificación', 'Punto de control de asistencia'])
        
        for _, r in group.iterrows():
            ws.append([
                r['ID'], fecha, r['Nombre'], r['Apellido'], r['Departamento'], '-',
                r['Tiempo'], r['Tipo de pase de tarjeta'], r['Método de verificación'], r['Punto de control de asistencia']
            ])

    filename = f"Transacciones_{start_date_str}_{end_date_str}.xlsx"
    out_path = os.path.join(target_dir, filename)
    wb.save(out_path)
    print(f"Lote de pruebas generado con exito ({len(all_transactions)} transacciones) en: {out_path}")
    return out_path

if __name__ == '__main__':
    generar_lote_pruebas()
