"""
download_personal_info.py
==========================
Descarga y formateo automático de la Información Personal de Trabajadores desde HikCentral.
Sincroniza la lista en la Base de Datos SQLite (asistencia.db) y genera un Excel con el diseño ejecutivo corporativo GZG.
"""

import os
import sys
import time
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT_DIR)

from data.database import get_connection, init_db

CARPETA_DOWNLOADCENTER = r"C:\Users\GZG Minerales 2026\HCWebControlService\Downloadcenter"
CARPETA_DATA_CRUDA = os.path.join(ROOT_DIR, "downloads", "data_cruda")
CARPETA_DATA_PROCESADA = os.path.join(ROOT_DIR, "downloads", "data_procesada")
DB_PATH = os.path.join(ROOT_DIR, "data", "asistencia.db")

os.makedirs(CARPETA_DATA_CRUDA, exist_ok=True)
os.makedirs(CARPETA_DATA_PROCESADA, exist_ok=True)

HIKCENTRAL_URL = "https://127.0.0.1"
USERNAME = "admin"
PASSWORD = "GzG@ACCESO2026"


def buscar_ultimo_excel_personal() -> str:
    """Busca el archivo de Información Personal más reciente en el Downloadcenter de HikCentral."""
    import glob
    patrones = [
        os.path.join(CARPETA_DOWNLOADCENTER, "**", "Información personal*.xlsx"),
        os.path.join(CARPETA_DOWNLOADCENTER, "Información personal*.xlsx"),
        os.path.join(CARPETA_DATA_CRUDA, "Información personal*.xlsx")
    ]
    archivos = []
    for pat in patrones:
        archivos.extend(glob.glob(pat, recursive=True))
    
    # Filtrar temporales ~$
    archivos = [f for f in archivos if not os.path.basename(f).startswith("~$")]
    if not archivos:
        return None
    archivos.sort(key=os.path.getmtime, reverse=True)
    return archivos[0]


def automatizar_descarga_hikcentral(headless: bool = True) -> str:
    """Ejecuta Playwright para accionar el botón de exportación en la web de HikCentral."""
    print("[HikCentral-Personal] Iniciando automatización para exportar personal...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, args=["--ignore-certificate-errors", "--no-sandbox"])
        context = browser.new_context(ignore_https_errors=True, accept_downloads=True)
        page = context.new_page()
        try:
            page.goto(f"{HIKCENTRAL_URL}/#/", wait_until="domcontentloaded")
            time.sleep(3)

            # Iniciar sesión si es necesario
            if page.locator("input[placeholder='Nombre de usuario']").count() > 0 or "Iniciar" in page.locator("body").inner_text():
                print("[HikCentral-Personal] Autenticando con usuario admin...")
                page.evaluate("document.querySelectorAll('input').forEach(i => i.removeAttribute('readonly'))")
                time.sleep(1)
                page.locator("#username, input[placeholder='Nombre de usuario']").first.fill(USERNAME)
                page.locator("input[type='password']").first.fill(PASSWORD)
                page.locator("button:has-text('Iniciar')").first.click()
                time.sleep(4)

            # Cerrar popup OK
            page.evaluate("() => { const okBtn = Array.from(document.querySelectorAll('button, div')).find(e => e.textContent.trim() === 'OK'); if (okBtn) okBtn.click(); }")
            time.sleep(1)

            # Navegar a módulo Persona
            print("[HikCentral-Personal] Navegando a módulo Persona...")
            page.evaluate("() => { const el = Array.from(document.querySelectorAll('div, span, li, a')).find(e => e.textContent.trim() === 'Persona' && e.offsetHeight > 0); if (el) el.click(); }")
            time.sleep(3)

            # Clic en Exportar principal
            print("[HikCentral-Personal] Clic en botón Exportar...")
            page.evaluate("() => { const expBtn = Array.from(document.querySelectorAll('button, div, span, i')).find(e => e.textContent.trim().includes('Exportar') && e.offsetHeight > 0); if (expBtn) expBtn.click(); }")
            time.sleep(2)

            # Clic en Exportar del modal
            page.evaluate("""
                () => {
                    const btns = Array.from(document.querySelectorAll('button'));
                    const redBtn = btns.find(b => (b.textContent.trim() === 'Exportar' || b.className.includes('el-button--danger') || b.className.includes('primary')) && b.offsetHeight > 0);
                    if (redBtn) redBtn.click();
                    else {
                        const lastExp = btns.filter(b => b.textContent.trim() === 'Exportar').pop();
                        if (lastExp) lastExp.click();
                    }
                }
            """)
            time.sleep(3)
            print("[HikCentral-Personal] Exportación ejecutada exitosamente en el servidor.")
        except Exception as e:
            print("[HikCentral-Personal] Aviso durante navegación:", e)
        finally:
            browser.close()

    return buscar_ultimo_excel_personal()


def procesar_y_exportar_padron(excel_raw_path: str):
    """Lee el Excel descargado, actualiza la base de datos y genera el reporte procesado elegante."""
    import pandas as pd
    import shutil

    if not excel_raw_path or not os.path.exists(excel_raw_path):
        print("[Error] No se encontró el archivo de Información Personal.")
        return None

    # Guardar copia cruda
    fecha_hoy = datetime.date.today().strftime("%Y-%m-%d")
    raw_dest = os.path.join(CARPETA_DATA_CRUDA, f"Informacion_Personal_Raw_{fecha_hoy}.xlsx")
    try:
        shutil.copy2(excel_raw_path, raw_dest)
        print(f"[OK] Resguardada Data Cruda en: {raw_dest}")
    except Exception:
        pass

    # Leer Excel identificando el encabezado dinámico
    df_full = pd.read_excel(excel_raw_path, header=None)
    header_idx = 0
    for idx, row in df_full.iterrows():
        r_str = [str(x).strip() for x in row.values]
        if 'ID' in r_str and ('Nombre' in r_str or 'Apellido' in r_str):
            header_idx = idx
            break

    df = pd.read_excel(excel_raw_path, skiprows=header_idx)
    df.rename(columns={
        'ID': 'DNI',
        'Nombre': 'NOMBRES',
        'Apellido': 'APELLIDOS',
        'Departamento': 'AREA',
        'Posición': 'CARGO',
        'Posicin': 'CARGO',
        'Fecha de inicio del periodo efectivo': 'FECHA_INICIO',
        'Fecha final del periodo efectivo': 'FECHA_FIN'
    }, inplace=True)

    # Limpieza
    def clean_area(val):
        val_str = str(val).strip()
        if '>' in val_str: val_str = val_str.split('>')[-1].strip()
        if '/' in val_str: val_str = val_str.split('/')[-1].strip()
        return val_str

    df['DNI'] = df['DNI'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
    df['APELLIDOS'] = df['APELLIDOS'].fillna('').astype(str).str.strip()
    df['NOMBRES'] = df['NOMBRES'].fillna('').astype(str).str.strip()
    df['CARGO'] = df['CARGO'].fillna('').astype(str).str.strip()
    df['AREA'] = df['AREA'].apply(clean_area)
    df['FECHA_INICIO'] = df.get('FECHA_INICIO', '').fillna('').astype(str).str.strip()
    df['FECHA_FIN'] = df.get('FECHA_FIN', '').fillna('').astype(str).str.strip()

    # Asegurar la presencia de JOSE ORLANDO MONCADA REJAS (DNI 46181231) si no vino en el export de personal
    if '46181231' not in df['DNI'].values:
        df_moncada = pd.DataFrame([{
            'DNI': '46181231',
            'APELLIDOS': 'MONCADA REJAS',
            'NOMBRES': 'JOSE ORLANDO',
            'AREA': 'Oper&Mtto',
            'CARGO': 'Operativo',
            'FECHA_INICIO': '',
            'FECHA_FIN': ''
        }])
        df = pd.concat([df, df_moncada], ignore_index=True)

    df.sort_values(by=['APELLIDOS', 'NOMBRES'], inplace=True)

    # 1. Actualizar Base de Datos SQLite
    init_db(DB_PATH)
    conn = get_connection(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM trabajadores;")
    for _, r in df.iterrows():
        dni = r['DNI']
        if not dni or dni.lower() == 'nan': continue
        cursor.execute("""
            INSERT INTO trabajadores (dni, apellidos, nombres, cargo, area, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """, (dni, r['APELLIDOS'], r['NOMBRES'], r['CARGO'], r['AREA']))
    conn.commit()
    conn.close()
    print(f"[OK] Sincronizados {len(df)} trabajadores en SQLite asistencia.db.")

    # 2. Generar Excel Procesado con Estilos Corporativos GZG
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Padrón de Trabajadores"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    fill_banner_title = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_banner_sub = PatternFill(start_color="2C4D75", end_color="2C4D75", fill_type="solid")
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_row_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    font_banner_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_banner_sub = Font(name="Calibri", size=11, bold=True, color="DFA86A") # Dorado
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11, bold=False, color="000000")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    thin_gray = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    # Banner Título (Fila 1 - 6 columnas)
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:F1")
    ws["A1"] = "PADRÓN OFICIAL DE TRABAJADORES Y PERSONAL REGISTRADO"
    ws["A1"].fill = fill_banner_title
    ws["A1"].font = font_banner_title
    ws["A1"].alignment = align_center

    # Banner Subtítulo (Fila 2)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:F2")
    ws["A2"] = f"GZG Minerales | Fecha de Actualización: {datetime.date.today().strftime('%Y-%m-%d')} | Fuente: Biométrico HikCentral"
    ws["A2"].fill = fill_banner_sub
    ws["A2"].font = font_banner_sub
    ws["A2"].alignment = align_center

    ws.row_dimensions[3].height = 10
    ws.append([])  # Fila 3 vacía

    # Encabezados (Fila 4 - 6 columnas)
    ws.row_dimensions[4].height = 28
    headers = [
        "DNI", "Apellidos", "Nombres", "Departamento / Área",
        "Posición / Cargo", "Estado en Sistema"
    ]
    ws.append(headers)

    for cell in ws[4]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center

    # Filas de datos
    for idx, r in df.iterrows():
        current_row = ws.max_row + 1
        ws.row_dimensions[current_row].height = 20
        row_vals = [
            r['DNI'], r['APELLIDOS'], r['NOMBRES'], r['AREA'],
            r['CARGO'], "Activo"
        ]
        ws.append(row_vals)

        for col_idx in range(1, 7):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = align_center if col_idx in (1, 6) else align_left
            if current_row % 2 == 0:
                cell.fill = fill_row_even
            if col_idx == 1:
                cell.number_format = '@'

    # Anchos de columna (6 columnas)
    widths = {1: 16, 2: 28, 3: 26, 4: 28, 5: 26, 6: 18}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # El padrón se guarda EXCLUSIVAMENTE en la carpeta raíz del proyecto
    root_path = os.path.join(ROOT_DIR, "Padron_Trabajadores_GZG.xlsx")
    old_data_proc_path = os.path.join(CARPETA_DATA_PROCESADA, "Padron_Trabajadores_GZG.xlsx")
    
    # Intentar eliminar del directorio data_procesada si existiera
    if os.path.exists(old_data_proc_path):
        try:
            os.remove(old_data_proc_path)
            print(f"[OK] Eliminado Padrón antiguo de data_procesada: {old_data_proc_path}")
        except Exception as e:
            print(f"[Warn] No se pudo borrar de data_procesada (en uso): {e}")

    try:
        wb.save(root_path)
        print(f"[OK] Padrón Oficial guardado en la raíz del proyecto: {root_path}")
    except PermissionError:
        import time
        time.sleep(1)
        try:
            wb.save(root_path)
            print(f"[OK] Padrón Oficial guardado en la raíz del proyecto: {root_path}")
        except Exception as e:
            print(f"[Error] No se pudo sobrescribir Padron_Trabajadores_GZG.xlsx (Asegúrese de cerrar Excel si está abierto): {e}")

    return root_path


if __name__ == "__main__":
    excel_path = automatizar_descarga_hikcentral(headless=True)
    if not excel_path:
        excel_path = buscar_ultimo_excel_personal()
    procesar_y_exportar_padron(excel_path)
