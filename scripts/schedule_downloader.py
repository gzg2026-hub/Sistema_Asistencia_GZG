"""
schedule_downloader.py
======================
Descarga automática diaria de transacciones Hikvision a las 8:00 AM.
Descarga el DIA ANTERIOR al actual (ejemplo: si hoy es 18/08, descarga el 17/08).

MODOS DE USO:
  - Automático (Tarea Programada Windows): ejecutar sin argumentos
        python schedule_downloader.py

  - Manual inmediato (día anterior):
        python schedule_downloader.py ahora

  - Manual con fecha específica:
        python schedule_downloader.py manual

  - Manual con fecha desde argumento:
        python schedule_downloader.py 2026-08-17
        python schedule_downloader.py 2026-08-15 2026-08-17   (rango)
"""

import os
import sys
import time
import datetime

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT_DIR)

# ── Carpetas de descargas y procesamiento ──────────────────────────────────────
CARPETA_DATA_CRUDA = os.path.join(ROOT_DIR, "downloads", "data_cruda")
CARPETA_DATA_PROCESADA = os.path.join(ROOT_DIR, "downloads", "data_procesada")
CARPETA_DOWNLOADCENTER = r"C:\Users\GZG Minerales 2026\HCWebControlService\Downloadcenter"

os.makedirs(CARPETA_DATA_CRUDA, exist_ok=True)
os.makedirs(CARPETA_DATA_PROCESADA, exist_ok=True)

LOG_FILE = os.path.join(ROOT_DIR, "logs", "descarga_diaria.log")
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)


def _log(msg: str):
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    linea = f"[{ts}] {msg}"
    try:
        print(linea)
    except Exception:
        print(linea.encode('ascii', errors='replace').decode('ascii'))
    with open(LOG_FILE, "a", encoding="utf-8") as f:
        f.write(linea + "\n")


def _sincronizar_downloadcenter():
    """Copia archivos de data cruda exportados manualmente en Downloadcenter hacia downloads/data_cruda."""
    if os.path.exists(CARPETA_DOWNLOADCENTER):
        try:
            import shutil
            for root_dir, dirs, files in os.walk(CARPETA_DOWNLOADCENTER):
                for f in files:
                    if f.endswith(".xlsx") and not f.startswith("~$"):
                        src = os.path.join(root_dir, f)
                        dst = os.path.join(CARPETA_DATA_CRUDA, f)
                        if not os.path.exists(dst):
                            shutil.copy2(src, dst)
                            _log(f"Sincronizado archivo desde Downloadcenter: {f}")
        except Exception as e:
            _log(f"Aviso al sincronizar Downloadcenter: {e}")


def _ejecutar_descarga(fecha_inicio: str, fecha_fin: str):
    """Descarga, procesa y guarda las transacciones del rango de fechas dado."""
    _log("=" * 60)
    _log(f"Descargando transacciones del {fecha_inicio} al {fecha_fin}...")

    # Sincronizar descargas manuales de Downloadcenter si existen
    _sincronizar_downloadcenter()

    try:
        from core.hikvision_downloader import descargar_transacciones_hikvision
        from data.data_loader import cargar_datos_excel, fusionar_y_deduplicar_data_cruda
        from core.attendance_engine import procesar_asistencia_df
        from data.database import (guardar_trabajadores, guardar_marcaciones_raw,
                                   guardar_asistencia_y_reportes)
        from data.exporter import guardar_excel_base, exportar_asistencia_excel
        from scripts.gdrive_uploader import subir_archivo_a_gdrive

        # 1. Descargar Data Cruda 1:1 desde Hikvision en downloads/data_cruda/
        excel_path_nuevo = descargar_transacciones_hikvision(
            carpeta_destino=CARPETA_DATA_CRUDA,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin
        )
        _log(f"Data Cruda descargada en: {excel_path_nuevo}")

        # 2. Cargar y Fusionar en el Archivo Maestro de Data Cruda sin Duplicados ni Faltantes
        df_trab_nuevo, df_marc_nuevo, df_he_nuevo = cargar_datos_excel(excel_path_nuevo)
        
        ruta_maestro_raw = os.path.join(CARPETA_DATA_CRUDA, "Transacciones_Acumuladas.xlsx")
        df_marc_master = fusionar_y_deduplicar_data_cruda(df_marc_nuevo, ruta_maestro_raw)

        # Limpieza estricta e inmediata de cualquier descarga o reporte en downloads/data_cruda
        # para conservar ÚNICAMENTE Transacciones_Acumuladas.xlsx
        try:
            for item_tmp in os.listdir(CARPETA_DATA_CRUDA):
                item_tmp_path = os.path.join(CARPETA_DATA_CRUDA, item_tmp)
                if os.path.isfile(item_tmp_path) and item_tmp.strip().lower() != "transacciones_acumuladas.xlsx" and not item_tmp.startswith("~$"):
                    try:
                        os.remove(item_tmp_path)
                    except Exception:
                        pass
            _log(f"Limpieza estricta de carpeta data_cruda completada. Conservando únicamente Maestro: {ruta_maestro_raw}")
        except Exception as e_clean:
            _log(f"Aviso al limpiar temporales en data_cruda: {e_clean}")

        # Fallback a base de datos de trabajadores si el Excel crudo no incluye pestaña de trabajadores
        df_trab = df_trab_nuevo
        if df_trab.empty:
            from data.database import obtener_trabajadores_master
            df_trab = obtener_trabajadores_master()

        _log(f"Data Cruda Maestro Acumulada: {len(df_marc_master)} marcaciones limpias sin duplicados, {len(df_trab)} trabajadores")

        if not df_marc_master.empty:
            # Completar Posición desde maestro de trabajadores si está vacía y quitar tildes
            if not df_trab.empty:
                c_col = 'CARGO' if 'CARGO' in df_trab.columns else ('Cargo' if 'Cargo' in df_trab.columns else None)
                d_col = 'DNI' if 'DNI' in df_trab.columns else None
                if c_col and d_col:
                    cargo_dict = dict(zip(df_trab[d_col].astype(str).str.strip().str.zfill(8), df_trab[c_col].astype(str).str.strip()))
                    pos_col = 'Posición' if 'Posición' in df_marc_master.columns else ('Posicion' if 'Posicion' in df_marc_master.columns else None)
                    id_col = 'ID' if 'ID' in df_marc_master.columns else 'DNI'
                    if pos_col and id_col:
                        for r_idx, r_val in df_marc_master.iterrows():
                            cur_p = str(r_val.get(pos_col, '')).strip()
                            d_id = str(r_val.get(id_col, '')).strip().zfill(8)
                            if (not cur_p or cur_p.lower() in ('nan', 'none', '', '-')) and d_id in cargo_dict:
                                df_marc_master.loc[r_idx, pos_col] = cargo_dict[d_id]

            for col_name in ['Nombre', 'Apellido', 'Nombres', 'Apellidos']:
                if col_name in df_marc_master.columns:
                    df_marc_master[col_name] = df_marc_master[col_name].astype(str).apply(quitar_tildes)

            # Guardar el Archivo Maestro de Data Cruda en Excel
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                wb_m = openpyxl.Workbook()
                ws_m = wb_m.active
                ws_m.title = "Transacciones"
                ws_m.views.sheetView[0].showGridLines = True
                ws_m.freeze_panes = "A2"

                fill_h = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                font_d = Font(name="Calibri", size=11, bold=False, color="000000")
                align_h = Alignment(horizontal="center", vertical="center", wrap_text=True)
                align_c = Alignment(horizontal="center", vertical="center")
                align_l = Alignment(horizontal="left", vertical="center")
                thin_gray = Side(border_style="thin", color="D3D3D3")
                thin_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

                cols_m = df_marc_master.columns.tolist()
                ws_m.append(cols_m)
                ws_m.row_dimensions[1].height = 28

                for cell in ws_m[1]:
                    cell.fill = fill_h
                    cell.font = font_h
                    cell.alignment = align_h
                    cell.border = thin_border

                for r_idx, r_m in enumerate(df_marc_master.itertuples(index=False), start=2):
                    ws_m.append(list(r_m))
                    ws_m.row_dimensions[r_idx].height = 20
                    for c_idx in range(1, len(cols_m) + 1):
                        cell = ws_m.cell(row=r_idx, column=c_idx)
                        cell.font = font_d
                        cell.border = thin_border
                        if c_idx in (1, 8, 10):
                            cell.alignment = align_c
                            cell.number_format = '@'
                        else:
                            cell.alignment = align_l

                for c_idx in range(1, len(cols_m) + 1):
                    col_letter = get_column_letter(c_idx)
                    max_len = 0
                    for r_idx in range(1, min(ws_m.max_row + 1, 100)):
                        v = ws_m.cell(row=r_idx, column=c_idx).value
                        if v is not None:
                            max_len = max(max_len, len(str(v)))
                    ws_m.column_dimensions[col_letter].width = max(max_len + 4, 12)

                wb_m.save(ruta_maestro_raw)
                _log(f"Archivo Maestro Data Cruda guardado con formato corporativo en: {ruta_maestro_raw}")
            except Exception as e_m:
                _log(f"Aviso guardando maestro data cruda: {e_m}")

            # Guardar y Subir Data Cruda Maestro (Transacciones_Acumuladas.xlsx) a Google Drive (directo a carpeta AGOSTO)
            subir_archivo_a_gdrive(ruta_maestro_raw)
            guardar_marcaciones_raw(df_marc_master, archivo_origen=ruta_maestro_raw)
            if not df_trab.empty:
                guardar_trabajadores(df_trab)
                df_asis, df_he_out, df_inc, kpis = procesar_asistencia_df(df_trab, df_marc_master)
                guardar_asistencia_y_reportes(df_asis, df_he_out, df_inc)
                
                # 3. Guardar Reporte Consolidado Período Completo en la raíz y en data_procesada (localmente)
                guardar_excel_base(df_trab, df_marc_master, df_asis, df_he_out, df_inc)
                _log(f"Procesamiento consolidado completado. Guardado en carpeta raíz y downloads/data_procesada/.")

                # 4. Generar y Subir a Google Drive ÚNICAMENTE los Reportes Procesados Diarios de Días Cerrados (directo a carpeta AGOSTO)
                carp_diario = os.path.join(CARPETA_DATA_PROCESADA, "diario")
                os.makedirs(carp_diario, exist_ok=True)

                hoy_str = datetime.date.today().strftime("%Y-%m-%d")
                if 'FECHA' in df_asis.columns:
                    fechas_unicas = sorted([str(f) for f in df_asis['FECHA'].dropna().unique()])
                    for f_dia in fechas_unicas:
                        # Estrictamente SOLO días cerrados/completados (anteriores a hoy)
                        if f_dia < hoy_str:
                            df_asis_dia = df_asis[df_asis['FECHA'].astype(str) == f_dia]
                            if not df_asis_dia.empty:
                                file_name_dia = f"Reporte_Asistencia_GZG_{f_dia}.xlsx"
                                file_path_dia = os.path.join(carp_diario, file_name_dia)

                                # Generar bytes de Excel diario completado
                                excel_bytes = exportar_asistencia_excel(df_trab, df_marc_master, df_asis_dia, df_he_out, df_inc)
                                with open(file_path_dia, "wb") as f_out:
                                    f_out.write(excel_bytes)
                                _log(f"Reporte diario completado generado para {f_dia} -> {file_path_dia}")

                                # ÚNICO ARCHIVO AUTORIZADO A SUBIR A GOOGLE DRIVE: Reporte Diario Procesado de Día Cerrado (directo a AGOSTO)
                                subir_archivo_a_gdrive(file_path_dia)
                        else:
                            _log(f"Día actual {f_dia} en curso: NO se genera reporte diario incompleto (se mantiene acumulado en Data Cruda).")
                
                # Actualizar únicamente de forma local en la PC el archivo raíz principal Sistema_Asistencia_GZG_v1.0.xlsx
                ruta_root_v1 = os.path.join(ROOT_DIR, "Sistema_Asistencia_GZG_v1.0.xlsx")
                excel_bytes_root = exportar_asistencia_excel(df_trab, df_marc_master, df_asis, df_he_out, df_inc)
                try:
                    with open(ruta_root_v1, "wb") as f_out:
                        f_out.write(excel_bytes_root)
                    _log(f"Archivo raíz principal actualizado en PC local: {ruta_root_v1}")
                except Exception as e_v1:
                    _log(f"Aviso actualizando archivo raíz principal local: {e_v1}")
            else:
                _log("AVISO: No se encontraron trabajadores en la base de datos para procesar asistencia.")
        else:
            _log("AVISO: No se encontraron marcaciones en el archivo.")

    except Exception as e:
        import traceback
        _log(f"ERROR: {e}")
        _log(traceback.format_exc())

    # 5. Sincronización Automática con GitHub para Streamlit Cloud
    try:
        _log("Sincronizando cambios diarios con GitHub / Streamlit Cloud...")
        os.system('git add data/asistencia.db downloads/ && git commit -m "auto: Daily attendance update 9AM" && git push origin main')
        _log("[OK] Sincronizado exitosamente con GitHub.")
    except Exception as e_git:
        _log(f"Aviso en sincronización git: {e_git}")

    _log("Descarga finalizada.")
    _log("=" * 60)


def _hoy() -> str:
    return datetime.date.today().strftime("%Y-%m-%d")


def _ayer() -> str:
    return (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")


def _menu_manual():
    """Menú interactivo para elegir la fecha a descargar manualmente."""
    print("\n" + "=" * 58)
    print("  DESCARGA MANUAL DE TRANSACCIONES HIKVISION - GZG")
    print("=" * 58)
    print(f"  Fecha actual (HOY) : {_hoy()} ({datetime.date.today().strftime('%d/%m/%Y')})")
    print(f"  Día anterior (AYER): {_ayer()}")
    print("=" * 58)
    print("\n  Opciones:")
    print("  [1] Descargar el día de HOY (Día actual)")
    print("  [2] Descargar el día de AYER")
    print("  [3] Descargar una fecha específica")
    print("  [4] Descargar un rango de fechas")
    print("  [5] Salir")
    print()

    opcion = input("  Seleccione opción (1-5): ").strip()

    if opcion == "1":
        fecha = _hoy()
        print(f"\n  Descargando HOY: {fecha}")
        _ejecutar_descarga(fecha, fecha)

    elif opcion == "2":
        fecha = _ayer()
        print(f"\n  Descargando AYER: {fecha}")
        _ejecutar_descarga(fecha, fecha)

    elif opcion == "3":
        fecha_str = input("  Ingrese la fecha (DD/MM/YYYY, YYYY-MM-DD o YYYY/MM/DD): ").strip()
        fecha = _parsear_fecha(fecha_str)
        if fecha:
            print(f"\n  Descargando: {fecha}")
            _ejecutar_descarga(fecha, fecha)
        else:
            print("  Fecha inválida. Use formato DD/MM/YYYY o YYYY-MM-DD.")

    elif opcion == "4":
        ini_str = input("  Fecha inicio (DD/MM/YYYY, YYYY-MM-DD o YYYY/MM/DD): ").strip()
        fin_str = input("  Fecha fin    (DD/MM/YYYY, YYYY-MM-DD o YYYY/MM/DD): ").strip()
        ini = _parsear_fecha(ini_str)
        fin = _parsear_fecha(fin_str)
        if ini and fin:
            if ini > fin:
                ini, fin = fin, ini
            print(f"\n  Descargando rango: {ini} → {fin}")
            _ejecutar_descarga(ini, fin)
        else:
            print("  Fechas inválidas.")

    elif opcion == "5":
        print("  Saliendo.")
    else:
        print("  Opción inválida.")

    input("\n  Presione Enter para cerrar...")


def _parsear_fecha(s: str) -> str | None:
    """Acepta DD/MM/YYYY, YYYY-MM-DD, YYYY/MM/DD, DD-MM-YYYY y devuelve YYYY-MM-DD."""
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _iniciar_programador():
    """Servicio que espera hasta las 09:00 AM y ejecuta la descarga del día anterior."""
    try:
        import schedule
    except ImportError:
        os.system(f'"{sys.executable}" -m pip install schedule')
        import schedule

    _log("Servicio de descarga diaria iniciado.")
    _log(f"  Horario    : todos los días a las 09:00 AM")
    _log(f"  Descarga   : día ANTERIOR al de ejecución (ayer)")
    _log(f"  Data Cruda : {CARPETA_DATA_CRUDA}")
    _log(f"  Procesada  : {CARPETA_DATA_PROCESADA}")
    _log(f"  Log        : {LOG_FILE}")

    def _tarea_9am():
        fecha = _ayer()
        _log(f"Tarea programada: descargando día anterior = {fecha}")
        _ejecutar_descarga(fecha, fecha)

    schedule.every().day.at("09:00").do(_tarea_9am)

    while True:
        schedule.run_pending()
        time.sleep(30)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    args = sys.argv[1:]

    if not args or args[0].lower() in ("ahora", "now", "auto", "automatico"):
        # Modo por defecto / Tarea Programada de Windows (a las 9:00 AM):
        # Descarga el rango de AYER a HOY para incluir salidas matutinas de turno noche
        ini = _ayer()
        fin = _hoy()
        _log(f"Ejecución AUTOMÁTICA (9:00 AM): descargando rango {ini} -> {fin} para emparejamiento completo de Turno Noche...")
        _ejecutar_descarga(ini, fin)

    elif args[0].lower() in ("daemon", "service", "servicio"):
        # Modo servicio continuo 24/7 (bucle)
        _iniciar_programador()

    elif args[0].lower() == "manual":
        # Menú interactivo para elegir fecha
        _menu_manual()

    elif len(args) == 1 and _parsear_fecha(args[0]):
        # Fecha específica como argumento: python schedule_downloader.py 2026-08-17
        fecha = _parsear_fecha(args[0])
        _log(f"Ejecución MANUAL con fecha: {fecha}")
        _ejecutar_descarga(fecha, fecha)

    elif len(args) == 2 and _parsear_fecha(args[0]) and _parsear_fecha(args[1]):
        # Rango: python schedule_downloader.py 2026-08-15 2026-08-17
        ini = _parsear_fecha(args[0])
        fin = _parsear_fecha(args[1])
        _log(f"Ejecución MANUAL con rango: {ini} -> {fin}")
        _ejecutar_descarga(ini, fin)

    else:
        print(__doc__)
