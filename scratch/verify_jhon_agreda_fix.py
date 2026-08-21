import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21_115415.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print("=== VERIFICACIÓN JHON AGREDA EN NUEVO REPORTE CON UMBRAL 30 MIN ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).strip() if row[1] is not None else ""
    if "AGREDA" in apellidos or "ÁGREDA" in apellidos:
        vals = [cell for cell in row]
        print(f"Fecha: {vals[5]} | Ent: {vals[8]} {vals[9]} | Sal: {vals[10]} {vals[11]} | Trab: {vals[16]} | Exc: {vals[18]} | HE Tot: {vals[19]} | Obs: {vals[21]}")

print("\n=== MUESTRA DE OBSERVACIONES ORDENADAS (H.E. / EXCESO PRIMERO, TARDANZA AL FINAL) ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    obs = str(row[21]).strip() if row[21] is not None else ""
    if obs and obs != "None":
        print(f"{row[1]} {row[2]} ({row[5]}): {obs}")
