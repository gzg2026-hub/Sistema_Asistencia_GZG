import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE HUAYAMA ADRIANO JOSMELL WALDIR EN {report_path} ===")

for r_idx in range(5, ws.max_row + 1):
    dni = str(ws.cell(row=r_idx, column=1).value or "").strip()
    if dni == "46671923":
        apellidos = ws.cell(row=r_idx, column=2).value
        nombres = ws.cell(row=r_idx, column=3).value
        fecha = ws.cell(row=r_idx, column=6).value
        h_ent = ws.cell(row=r_idx, column=10).value
        h_sal = ws.cell(row=r_idx, column=12).value
        h_turno = ws.cell(row=r_idx, column=17).value
        exc_turno = ws.cell(row=r_idx, column=18).value
        tot_adic = ws.cell(row=r_idx, column=20).value
        tipo_reg = ws.cell(row=r_idx, column=22).value
        obs = ws.cell(row=r_idx, column=23).value
        print(f"Fila {r_idx:3d} | {apellidos} {nombres} ({fecha}): Ent={h_ent} | Sal={h_sal} | Horas de Turno={h_turno} | Exceso de Turno={exc_turno} | Total Adic={tot_adic} | Tipo='{tipo_reg}' | Obs='{obs}'")
