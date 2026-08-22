import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE REGLA DE MANTENIMIENTO (CORTE 06:25 AM) EN {report_path} ===")

print("\n--- CASO HUAYAMA ADRIANO JOSMELL WALDIR (DNI 46671923) ---")
for r_idx in range(5, ws.max_row + 1):
    dni = str(ws.cell(row=r_idx, column=1).value or "").strip()
    if dni == "46671923":
        ap = ws.cell(row=r_idx, column=2).value
        nom = ws.cell(row=r_idx, column=3).value
        f = ws.cell(row=r_idx, column=6).value
        ent = ws.cell(row=r_idx, column=10).value
        sal = ws.cell(row=r_idx, column=12).value
        ht = ws.cell(row=r_idx, column=17).value
        exc = ws.cell(row=r_idx, column=18).value
        tot = ws.cell(row=r_idx, column=20).value
        tipo = ws.cell(row=r_idx, column=22).value
        obs = ws.cell(row=r_idx, column=23).value
        print(f"Fila {r_idx:3d} | {ap} {nom} ({f}): Ent={ent} | Sal={sal} | HT={ht} | Exceso={exc} | Total={tot} | Tipo='{tipo}' | Obs='{obs}'")

print("\n--- OTROS TRABAJADORES DE MANTENIMIENTO CON ENTRADA ANTES DE LAS 07:00 AM ---")
for r_idx in range(5, ws.max_row + 1):
    pos = str(ws.cell(row=r_idx, column=5).value or "").lower()
    dept = str(ws.cell(row=r_idx, column=4).value or "").lower()
    if "mantenimiento" in pos or "mantenimiento" in dept or "mtto" in pos or "mtto" in dept:
        ent = str(ws.cell(row=r_idx, column=10).value or "").strip()
        if ent and ent not in ("None", "-", ""):
            h_ent = int(ent.split(":")[0])
            m_ent = int(ent.split(":")[1])
            if 6 <= h_ent < 7:
                ap = ws.cell(row=r_idx, column=2).value
                nom = ws.cell(row=r_idx, column=3).value
                f = ws.cell(row=r_idx, column=6).value
                sal = ws.cell(row=r_idx, column=12).value
                ht = ws.cell(row=r_idx, column=17).value
                exc = ws.cell(row=r_idx, column=18).value
                tipo = ws.cell(row=r_idx, column=22).value
                obs = ws.cell(row=r_idx, column=23).value
                corte_info = "< 06:25 AM (Real)" if (h_ent == 6 and m_ent < 25) else ">= 06:25 AM (Snap 07:00)"
                print(f"Fila {r_idx:3d} | {ap} {nom} ({f}): Ent={ent} [{corte_info}] | Sal={sal} | HT={ht} | Exceso={exc} | Tipo='{tipo}' | Obs='{obs}'")
