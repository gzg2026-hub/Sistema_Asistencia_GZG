import sqlite3, os, sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
padron_path = os.path.join(ROOT_DIR, "Padron_Trabajadores_GZG.xlsx")

# 1. Obtener los 55 trabajadores con N1 y N2 desde la DB recuperada
conn_old = sqlite3.connect(os.path.join(ROOT_DIR, "scratch", "recovered_asistencia.db"))
df_trab = pd.read_sql_query("""
    SELECT dni, apellidos, nombres, area, cargo, 
           'Activo' as estado,
           COALESCE(aprobador_n1, '') as aprobador_n1,
           COALESCE(aprobador_n2, '') as aprobador_n2
    FROM trabajadores
    ORDER BY apellidos, nombres
""", conn_old)

df_users = pd.read_sql_query("""
    SELECT username as 'Usuario', 
           nombre_completo as 'Nombre Completo',
           rol as 'Rol',
           area_asignada as 'Área',
           cargo as 'Cargo',
           'Activo' as 'Estado'
    FROM usuarios
    ORDER BY id
""", conn_old)
conn_old.close()

print(f"Trabajadores recuperados: {len(df_trab)}")
print(f"Usuarios recuperados: {len(df_users)}")

# 2. Crear el libro de Excel con formato corporativo GZG
wb = openpyxl.Workbook()

# ==================== HOJA 1: TRABAJADORES ====================
ws1 = wb.active
ws1.title = "Trabajadores"
ws1.views.sheetView[0].showGridLines = True

# Colores y fuentes
fill_banner = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
font_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

fill_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

fill_header_aprob = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
font_header_aprob = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

font_data = Font(name="Calibri", size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

ws1.merge_cells("A1:H1")
ws1.row_dimensions[1].height = 28
ws1["A1"] = "PADRÓN OFICIAL DE TRABAJADORES Y PERSONAL REGISTRADO - GZG MINERALES"
ws1["A1"].fill = fill_banner
ws1["A1"].font = font_banner
ws1["A1"].alignment = align_center

ws1.merge_cells("A2:H2")
ws1.row_dimensions[2].height = 18
ws1["A2"] = "GZG Minerales | Estado Actualizado de Personal en Biométrico y Sistema de Aprobaciones"
ws1["A2"].fill = fill_sub
ws1["A2"].font = font_sub
ws1["A2"].alignment = align_center

headers1 = [
    "DNI", "Apellidos", "Nombres", "Departamento / Área", 
    "Posición / Cargo", "Estado en Sistema", 
    "Nivel de Aprobacion 1", "Nivel de Aprobacion 2"
]
ws1.row_dimensions[3].height = 26
ws1.append(headers1)

for col_idx, cell in enumerate(ws1[3], start=1):
    if col_idx in (7, 8):
        cell.fill = fill_header_aprob
    else:
        cell.fill = fill_header
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border

for _, r in df_trab.iterrows():
    dni = str(r['dni']).strip().zfill(8)
    ape = str(r['apellidos']).strip()
    nom = str(r['nombres']).strip()
    area = str(r['area']).strip()
    cargo = str(r['cargo']).strip()
    estado = str(r['estado']).strip()
    n1 = str(r['aprobador_n1']).strip() if r['aprobador_n1'] else ""
    n2 = str(r['aprobador_n2']).strip() if r['aprobador_n2'] else ""

    row_data = [dni, ape, nom, area, cargo, estado, n1, n2]
    ws1.append(row_data)
    curr_row = ws1.max_row
    ws1.row_dimensions[curr_row].height = 20

    for c_idx in range(1, 9):
        cell = ws1.cell(row=curr_row, column=c_idx)
        cell.font = font_data
        cell.border = thin_border
        if c_idx in (1, 6, 7, 8):
            cell.alignment = align_center
        else:
            cell.alignment = align_left
        if c_idx == 1:
            cell.number_format = '@'

widths1 = {1: 14, 2: 26, 3: 24, 4: 22, 5: 24, 6: 18, 7: 24, 8: 24}
for col_idx, w in widths1.items():
    ws1.column_dimensions[get_column_letter(col_idx)].width = w

ws1.auto_filter.ref = f"A3:H{ws1.max_row}"


# ==================== HOJA 2: USUARIOS Y ROLES ====================
ws2 = wb.create_sheet(title="Usuarios del Sistema")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells("A1:F1")
ws2.row_dimensions[1].height = 28
ws2["A1"] = "USUARIOS Y ROLES OFICIALES DE ACCESO PWA - GZG MINERALES"
ws2["A1"].fill = fill_banner
ws2["A1"].font = font_banner
ws2["A1"].alignment = align_center

ws2.merge_cells("A2:F2")
ws2.row_dimensions[2].height = 18
ws2["A2"] = "GZG Minerales | Control de Acceso y Niveles de Jerarquía de Aprobación"
ws2["A2"].fill = fill_sub
ws2["A2"].font = font_sub
ws2["A2"].alignment = align_center

headers2 = ["Usuario", "Nombre Completo", "Rol", "Área", "Cargo", "Estado"]
ws2.row_dimensions[3].height = 26
ws2.append(headers2)

for cell in ws2[3]:
    cell.fill = fill_header
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border

for _, r in df_users.iterrows():
    u = str(r['Usuario']).strip()
    nc = str(r['Nombre Completo']).strip()
    rol = str(r['Rol']).strip()
    area = str(r['Área']).strip()
    cargo = str(r['Cargo']).strip()
    estado = str(r['Estado']).strip()

    row_data = [u, nc, rol, area, cargo, estado]
    ws2.append(row_data)
    curr_row = ws2.max_row
    ws2.row_dimensions[curr_row].height = 20

    for c_idx in range(1, 7):
        cell = ws2.cell(row=curr_row, column=c_idx)
        cell.font = font_data
        cell.border = thin_border
        if c_idx in (1, 3, 6):
            cell.alignment = align_center
        else:
            cell.alignment = align_left

widths2 = {1: 16, 2: 34, 3: 20, 4: 20, 5: 26, 6: 14}
for col_idx, w in widths2.items():
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

ws2.auto_filter.ref = f"A3:F{ws2.max_row}"

wb.save(padron_path)
print(f"[OK] Padron_Trabajadores_GZG.xlsx restaurado con 2 hojas completas y aprobadores oficiales!")
