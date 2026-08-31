import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

local_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_cruda\Transacciones_Acumuladas.xlsx"
gdrive_path = r"G:\.shortcut-targets-by-id\1YpKPT9uTbWzHguHqJrJMb8U5V3GwnEoU\AGOSTO\Transacciones_Acumuladas.xlsx"

def aplicar_formato_excel(file_path):
    if not os.path.exists(file_path):
        print(f"Archivo no encontrado: {file_path}")
        return
    
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    # Congelar panel para que los encabezados queden fijos
    ws.freeze_panes = "A2"
    
    # Estilos de encabezado
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Calibri", size=11, bold=False, color="000000")
    
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    thin_gray = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    
    # Formatear Fila 1 (Encabezados)
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_header
        cell.border = thin_border
    
    # Formatear Filas de datos
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = font_data
            cell.border = thin_border
            # Alinear según contenido
            if col_idx in (1, 8, 10): # ID/DNI, Fecha, Tiempo
                cell.alignment = align_center
                cell.number_format = '@'
            else:
                cell.alignment = align_left
    
    # Ajustar anchos de columna automáticamente
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    try:
        wb.save(file_path)
        print(f"[OK] Formato aplicado exitosamente a: {file_path}")
    except PermissionError:
        print(f"[Aviso] El archivo '{os.path.basename(file_path)}' esta abierto en Excel.")
        print("  Por favor cierralo en Excel para actualizar los estilos.")
        return False
    return True

print("=== APLICANDO FORMATO CORPORATIVO A TRANSACCIONES ACUMULADAS ===")

# 1. Aplicar a archivo local
success = aplicar_formato_excel(local_path)

# 2. Copiar/Sincronizar a Google Drive
if success and os.path.exists(local_path) and os.path.exists(os.path.dirname(gdrive_path)):
    try:
        with open(local_path, "rb") as src_f:
            data = src_f.read()
        with open(gdrive_path, "wb") as dst_f:
            dst_f.write(data)
        print(f"[OK] Sincronizado a Google Drive AGOSTO: {gdrive_path}")
    except Exception as e:
        print(f"Error sincronizando a Google Drive: {e}")
