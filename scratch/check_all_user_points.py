import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_104042.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

headers = [cell for cell in ws[4]]
print("=== VERIFICACIÓN DE ENCABEZADOS DE COLUMNA (FILA 4) ===")
print([h.value for h in headers[16:20]])

print("\n=== VERIFICACIÓN 1: LLAMYR BERMEO CRUZ (IMAGE 1) ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).strip() if row[1] is not None else ""
    if "BERMEO" in apellidos or "LLAMYR" in str(row[2]):
        vals = [cell for cell in row]
        print(f"Fecha: {vals[5]} | Turno: {vals[7]} | Ent: {vals[8]} {vals[9]} | Sal: {vals[10]} {vals[11]} | Tipo: {vals[20]} | Obs: {vals[21]}")

print("\n=== VERIFICACIÓN 2: JOSMELL HUAYAMA - MANTENIMIENTO (IMAGE 2) ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).strip() if row[1] is not None else ""
    if "HUAYAMA" in apellidos or "JOSMELL" in str(row[2]):
        vals = [cell for cell in row]
        print(f"Fecha: {vals[5]} | Cargo: {vals[4]} | Ent: {vals[8]} {vals[9]} | Sal: {vals[10]} {vals[11]} | Trab: {vals[16]} | Exc: {vals[18]} | HE Tot: {vals[19]} | Obs: {vals[21]}")

print("\n=== VERIFICACIÓN 3: SANGAMA GUERRA & CAMBIO DE GUARDIA (IMAGE 3) ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).strip() if row[1] is not None else ""
    if any(k in apellidos for k in ["SANGAMA", "INFANTES", "NEYRA", "RAMIREZ", "RINARACHI"]):
        vals = [cell for cell in row]
        print(f"Nombre: {vals[1]} {vals[2]} | Fecha: {vals[5]} | Ent: {vals[9]} | Sal: {vals[11]} | Trab: {vals[16]} | Exc: {vals[18]} | HE Tot: {vals[19]} | Tipo: {vals[20]}")
