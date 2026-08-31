import os
import sys
import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG"
sys.path.insert(0, PROJECT_ROOT)

from data.database import obtener_trabajadores_master, obtener_datos_db
from core.attendance_engine import procesar_asistencia_df
from data.exporter import exportar_asistencia_excel
from scripts.gdrive_uploader import subir_archivo_a_gdrive

# 1. Obtener padrón actualizado de trabajadores y marcaciones de SQLite
df_trab = obtener_trabajadores_master()
df_trab_db, df_marc, df_asis_db, df_he_db, df_inc_db = obtener_datos_db()

print(f"Trabajadores recuperados: {len(df_trab)}")
print(f"Marcaciones recuperadas: {len(df_marc)}")

# 2. Generar datos procesados de asistencia actualizados
df_asis, df_he, df_inc, kpis = procesar_asistencia_df(df_trab, df_marc)

# 3. Actualizar Padrón de Trabajadores Excel existente en la raíz (Padron_Trabajadores_GZG.xlsx)
padron_file = os.path.join(PROJECT_ROOT, "Padron_Trabajadores_GZG.xlsx")

wb_p = openpyxl.Workbook()
ws_p = wb_p.active
ws_p.title = "Padron_Trabajadores"

# Encabezados en azul oscuro
fill_h = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
font_h = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

headers_p = ["DNI / ID", "Apellidos", "Nombres", "Cargo / Posición", "Área / Departamento"]
ws_p.append(headers_p)

for cell in ws_p[1]:
    cell.fill = fill_h
    cell.font = font_h
    cell.alignment = Alignment(horizontal="center", vertical="center")

if not df_trab.empty:
    for row in df_trab.itertuples(index=False):
        # DNI, Apellidos, Nombres, Cargo, Área
        ws_p.append([
            str(getattr(row, 'dni', '')).strip(),
            str(getattr(row, 'apellidos', '')).strip(),
            str(getattr(row, 'nombres', '')).strip(),
            str(getattr(row, 'cargo', '')).strip(),
            str(getattr(row, 'area', '')).strip()
        ])

wb_p.save(padron_file)
print(f"[OK] Padron de Trabajadores actualizado en archivo existente: {padron_file}")

# 4. Actualizar Reporte Consolidado de Asistencia en la raíz (Sistema_Asistencia_GZG_v1.0.xlsx)
master_excel = os.path.join(PROJECT_ROOT, "Sistema_Asistencia_GZG_v1.0.xlsx")
excel_bytes = exportar_asistencia_excel(df_trab, df_marc, df_asis, df_he, df_inc, master_excel)

with open(master_excel, "wb") as f_out:
    f_out.write(excel_bytes)
print(f"[OK] Excel maestro de asistencia actualizado en archivo existente: {master_excel}")

# 5. Subir Padrón de Trabajadores y Excel Maestro directamente a Google Drive
subir_archivo_a_gdrive(padron_file)
subir_archivo_a_gdrive(master_excel)

print("\nPadron de Trabajadores y Excel actualizados y subidos exitosamente a Google Drive!")
