import os
import pandas as pd
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def quitar_tildes(texto: str) -> str:
    if not isinstance(texto, str) or not texto or str(texto).strip().lower() in ('nan', 'none', ''):
        return ""
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U',
        'á': 'A', 'é': 'E', 'í': 'I', 'ó': 'O', 'ú': 'U', 'ü': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'à': 'A', 'è': 'E', 'ì': 'I', 'ò': 'O', 'ù': 'U',
    }
    res = str(texto)
    for k, v in replacements.items():
        res = res.replace(k, v)
    return res.strip()

print("=== ACTUALIZANDO ARCHIVO LOCAL Padron_Trabajadores_GZG.xlsx ===")

db_path = "data/asistencia.db"
conn = sqlite3.connect(db_path)
df_db = pd.read_sql_query("SELECT dni, apellidos, nombres, cargo, area FROM trabajadores", conn)
conn.close()

df_db['dni'] = df_db['dni'].astype(str).str.strip().str.zfill(8)
df_db['apellidos'] = df_db['apellidos'].astype(str).apply(quitar_tildes)
df_db['nombres'] = df_db['nombres'].astype(str).apply(quitar_tildes)
df_db['cargo'] = df_db['cargo'].astype(str).apply(quitar_tildes)
df_db['area'] = df_db['area'].astype(str).apply(quitar_tildes)

# Forzar cargo de Moncada
mask_moncada = df_db['dni'] == '46181231'
if mask_moncada.any():
    df_db.loc[mask_moncada, 'cargo'] = 'Ingeniero Metalurgico'

padron_path = "Padron_Trabajadores_GZG.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Trabajadores"
ws.views.sheetView[0].showGridLines = True

# Estilos corporativos
fill_banner = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
font_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

font_data = Font(name="Calibri", size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Fila 1: Banner Título
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 28
ws["A1"] = "PADRÓN OFICIAL DE TRABAJADORES Y PERSONAL REGISTRADO - GZG MINERALES"
ws["A1"].fill = fill_banner
ws["A1"].font = font_banner
ws["A1"].alignment = align_center

# Fila 2: Subtítulo Fecha
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 18
fill_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")
ws["A2"] = "GZG Minerales | Estado Actualizado de Personal en Biométrico y Sistema"
ws["A2"].fill = fill_sub
ws["A2"].font = font_sub
ws["A2"].alignment = align_center

# Fila 3: Encabezados
headers = ["DNI", "Apellidos", "Nombres", "Departamento / Área", "Posición / Cargo", "Estado en Sistema"]
ws.row_dimensions[3].height = 25
ws.append(headers)

for cell in ws[3]:
    cell.fill = fill_header
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border

# Filas de datos (Fila 4 en adelante)
for idx, r in df_db.sort_values(by=['apellidos', 'nombres']).iterrows():
    dni = str(r['dni']).strip().zfill(8)
    ape = str(r['apellidos']).strip()
    nom = str(r['nombres']).strip()
    area = str(r['area']).strip()
    cargo = str(r['cargo']).strip()
    estado = "Activo"

    ws.append([dni, ape, nom, area, cargo, estado])
    c_row = ws.max_row
    ws.row_dimensions[c_row].height = 20

    for c_i in range(1, 7):
        cell = ws.cell(row=c_row, column=c_i)
        cell.font = font_data
        cell.border = thin_border
        if c_i in (1, 6):
            cell.alignment = align_center
            if c_i == 1:
                cell.number_format = '@'
        else:
            cell.alignment = align_left

# Anchos de columna
widths = {1: 15, 2: 28, 3: 26, 4: 26, 5: 30, 6: 18}
for c_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(c_idx)].width = w

wb.save(padron_path)
print(f"Éxito: Archivo local {padron_path} actualizado correctamente con {len(df_db)} trabajadores.")
