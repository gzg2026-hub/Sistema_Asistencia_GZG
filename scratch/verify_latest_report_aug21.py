import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE REPORTE FINAL CON DATA AL 21/08/2026 ===")

print("\n--- 1. COMPROBACION IMAGEN 1: VERIFICAR QUE 'Cambio de guardia' NO SE REPITE EN COLUMNA V ---")
cambio_redundante = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    col_u = str(row[20])
    col_v = str(row[21])
    if col_u == "Cambio de guardia":
        if "Cambio de guardia" in col_v:
            print(f"  [REPETIDO] {row[1]} {row[2]} ({row[5]}): Col U='{col_u}' | Col V='{col_v}'")
            cambio_redundante += 1
        else:
            print(f"  [CORRECTO] {row[1]} {row[2]} ({row[5]}): Col U='{col_u}' | Col V='{col_v if col_v != 'None' else ''}'")

print(f"  Total filas redundantes encontradas: {cambio_redundante}")

print("\n--- 2. COMPROBACION IMAGEN 2: VERIFICAR SALIDAS EN TURNO NOCHE DEL 20/08/2026 (SALIDA EL 21/08/2026) ---")
target_workers = ["MAXIMILIANO", "VIGO", "CELEN", "MONCADA", "ALBITRES", "LADINES", "SANCHEZ", "BERMEO OJEDA", "SOLANO"]
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).upper()
    fecha_ent = str(row[5])
    turno = str(row[7])
    if fecha_ent == "20/08/2026" and turno == "NOCHE" and any(w in apellidos for w in target_workers):
        ent_str = f"{row[8]} {row[9]}"
        sal_str = f"{row[10]} {row[11]}"
        horas_trab = row[16]
        tipo_reg = row[20]
        obs = row[21]
        print(f"  * {row[1]} {row[2]}: Ent {ent_str} | Sal {sal_str} | Trab {horas_trab} | Tipo: {tipo_reg} | Obs: {obs}")
