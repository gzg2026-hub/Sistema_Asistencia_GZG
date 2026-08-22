import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-22.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE ORDEN DE COLUMNAS EN {report_path} ===")
for c_idx in range(1, ws.max_column + 1):
    letter = openpyxl.utils.get_column_letter(c_idx)
    header_val = ws.cell(row=4, column=c_idx).value
    fill_hex = ws.cell(row=4, column=c_idx).fill.start_color.rgb if ws.cell(row=4, column=c_idx).fill else "None"
    print(f"Col {letter:2s} ({c_idx:2d}): {header_val:30s} | Color Header: {fill_hex}")

print("\n--- MUESTRA PRIMERAS 5 FILAS DE DATOS (COLUMNAS Q a W) ---")
for r_idx in range(5, 10):
    ap = ws.cell(row=r_idx, column=2).value
    nom = ws.cell(row=r_idx, column=3).value
    ht = ws.cell(row=r_idx, column=17).value  # Col Q: Horas de Turno
    tard = ws.cell(row=r_idx, column=18).value # Col R: Tardanza
    exc = ws.cell(row=r_idx, column=19).value  # Col S: Exceso de Turno
    he = ws.cell(row=r_idx, column=20).value   # Col T: Horas Extras
    tot = ws.cell(row=r_idx, column=21).value  # Col U: Total Horas Adicionales
    tipo = ws.cell(row=r_idx, column=22).value # Col V: Tipo Registro
    obs = ws.cell(row=r_idx, column=23).value  # Col W: Observaciones
    print(f"Fila {r_idx}: {ap} {nom}")
    print(f"   Q(Horas Turno)={ht} | R(Tardanza)={tard} | S(Exceso Turno)={exc} | T(Horas Extras)={he} | U(Total Adic)={tot} | V='{tipo}' | W='{obs}'")
