import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21_141850.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACIÓN FINAL EN {report_path} ===")

print("\n--- 1. MUESTRA DE TIPO REGISTRO (COLUMNA U - PUNTO 2) ---")
tipo_counts = {}
for row in ws.iter_rows(min_row=5, values_only=True):
    tipo = str(row[20]).strip() if row[20] is not None else ""
    tipo_counts[tipo] = tipo_counts.get(tipo, 0) + 1

for t, count in tipo_counts.items():
    print(f"  * {t}: {count} filas")

print("\n--- 2. MUESTRA DE OBSERVACIONES CONCISAS Y DEDUPLICADAS (COLUMNA V - PUNTO 3) ---")
obs_samples = []
for row in ws.iter_rows(min_row=5, values_only=True):
    obs = str(row[21]).strip() if row[21] is not None else ""
    if obs and obs != "None":
        obs_samples.append(f"  * {row[1]} {row[2]} ({row[5]}): {obs}")

for sample in obs_samples[:20]:
    print(sample)

print("\n--- 3. VERIFICACIÓN CASOS DE LA IMAGEN 1 (PUNTO 4 - SIN FALSA ENTRADA DUPLICADA) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    obs = str(row[21]).strip() if row[21] is not None else ""
    if "Jornada Parcial" in obs:
        print(f"  * {row[1]} {row[2]} ({row[5]} {row[7]}): Ent {row[9]} | Sal {row[11]} | Trab {row[16]} | Obs: {obs}")
