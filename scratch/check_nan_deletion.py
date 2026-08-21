import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_v2.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

total_rows = ws.max_row
print(f"Total filas en reporte procesado: {total_rows}")

nan_i_and_k_count = 0
valid_rows_count = 0

for r_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
    col_i = str(row[8]).strip() if row[8] is not None else "" # Fecha Entrada
    col_k = str(row[10]).strip() if row[10] is not None else "" # Fecha Salida
    hora_ent = str(row[9]).strip() if row[9] is not None else ""
    hora_sal = str(row[11]).strip() if row[11] is not None else ""

    if (not col_i or col_i.lower() in ('nan', 'none', '', '-')) and (not col_k or col_k.lower() in ('nan', 'none', '', '-')) and not hora_ent and not hora_sal:
        nan_i_and_k_count += 1
    else:
        valid_rows_count += 1

print(f"Filas con datos de marcación válidos: {valid_rows_count}")
print(f"Filas con doble NaN en Col I y K restantes: {nan_i_and_k_count}")
