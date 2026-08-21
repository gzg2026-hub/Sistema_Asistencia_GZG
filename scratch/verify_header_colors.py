import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE COLORES DE ENCABEZADO EN FILA 4 EN {report_path} ===")

for col_idx in range(1, 24):
    cell = ws.cell(row=4, column=col_idx)
    header_name = cell.value
    fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else "None"
    print(f"  Col {col_idx:2d} ({openpyxl.utils.get_column_letter(col_idx)}): '{header_name}' -> RGB Fill: {fill_color}")
