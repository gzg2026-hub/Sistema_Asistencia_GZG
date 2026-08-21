import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE REGLAS 1 Y 2 EN EN REPORTE DE RAIZ ===")

print("\n--- 1. VERIFICACION DNI 46181231 (JOSE MONCADA REJAS - HORARIOS LIBRES) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    dni = str(row[0]).strip()
    if dni == "46181231":
        fecha = row[5]
        ent = row[9]
        sal = row[11]
        trab = row[16]
        tard = row[17]
        tipo_reg = row[20]
        obs = row[21]
        print(f"  * DNI 46181231 ({fecha}): Ent {ent} | Sal {sal} | Trab {trab} | Tard {tard} | Tipo: '{tipo_reg}' | Obs: '{obs}'")

print("\n--- 2. VERIFICACION TARDANZA > 30 MIN EN COLUMNA U (TIPO REGISTRO) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    tard_str = str(row[17]) if row[17] is not None else "00:00"
    tipo_reg = str(row[20])
    obs = str(row[21])
    
    # Convertir HH:MM a minutos
    if ":" in tard_str:
        h, m = tard_str.split(":")
        tard_min = int(h) * 60 + int(m)
    else:
        tard_min = 0
        
    if tard_min > 30:
        print(f"  * [TARDANZA > 30 MIN ({tard_str})] {row[1]} {row[2]} ({row[5]}): Col U (Tipo): '{tipo_reg}' | Col V (Obs): '{obs}'")
    elif 0 < tard_min <= 30:
        print(f"  * [TARDANZA <= 30 MIN ({tard_str})] {row[1]} {row[2]} ({row[5]}): Col U (Tipo): '{tipo_reg}' | Col V (Obs): '{obs}'")
