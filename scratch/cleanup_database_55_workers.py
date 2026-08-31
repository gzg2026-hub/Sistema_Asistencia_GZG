import os
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def quitar_tildes(texto: str) -> str:
    if not isinstance(texto, str) or not texto or str(texto).strip().lower() in ('nan', 'none', ''):
        return ""
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U',
        'á': 'A', 'é': 'E', 'í': 'I', 'ó': 'O', 'ú': 'U', 'ü': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'à': 'A', 'è': 'E', 'ì': 'I', 'ò': 'O', 'ù': 'U',
    }
    res = str(texto)
    for k, v in replacements.items():
        res = res.replace(k, v)
    return res.strip()

def normalizar_dni(dni_raw) -> str:
    if pd.isna(dni_raw) or not str(dni_raw).strip():
        return ""
    val = str(dni_raw).strip().split('.')[0]
    val_clean = val.lstrip('0')
    if not val_clean:
        return ""
    return val_clean.zfill(8)

print("=== DEPURACIÓN DE BASE DE DATOS Y NORMALIZACIÓN DE 55 TRABAJADORES ===")

db_path = "data/asistencia.db"
conn = sqlite3.connect(db_path)
c = conn.cursor()

# 1. Depurar y consolidar tabla trabajadores
df_trab = pd.read_sql_query("SELECT * FROM trabajadores", conn)
df_trab['dni'] = df_trab['dni'].apply(normalizar_dni)
df_trab['apellidos'] = df_trab['apellidos'].apply(quitar_tildes)
df_trab['nombres'] = df_trab['nombres'].apply(quitar_tildes)
df_trab['cargo'] = df_trab['cargo'].apply(quitar_tildes)
df_trab['area'] = df_trab['area'].apply(quitar_tildes)

# Forzar cargo de Moncada Rejas
df_trab.loc[df_trab['dni'] == '46181231', 'cargo'] = 'Ingeniero Metalurgico'

# Deduplicar conservando el último registro actualizado
df_trab_clean = df_trab.drop_duplicates(subset=['dni'], keep='last')

print(f"Trabajadores consolidados: {len(df_trab_clean)} (debe ser 55)")

# Reemplazar tabla trabajadores
c.execute("DROP TABLE IF EXISTS trabajadores")
c.execute("""
CREATE TABLE trabajadores (
    dni TEXT PRIMARY KEY,
    apellidos TEXT,
    nombres TEXT,
    cargo TEXT,
    area TEXT,
    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
)
""")

for _, r in df_trab_clean.iterrows():
    c.execute("""
        INSERT INTO trabajadores (dni, apellidos, nombres, cargo, area)
        VALUES (?, ?, ?, ?, ?)
    """, (r['dni'], r['apellidos'], r['nombres'], r['cargo'], r['area']))

# 2. Normalizar DNIs en marcaciones_raw
c.execute("SELECT id, dni FROM marcaciones_raw")
rows_marc = c.fetchall()
for row_id, raw_dni in rows_marc:
    clean_d = normalizar_dni(raw_dni)
    if clean_d != raw_dni:
        c.execute("UPDATE marcaciones_raw SET dni = ? WHERE id = ?", (clean_d, row_id))

# Actualizar cargos en marcaciones_raw según trabajadores
for _, r in df_trab_clean.iterrows():
    c.execute("UPDATE marcaciones_raw SET cargo = ? WHERE dni = ?", (r['cargo'], r['dni']))

conn.commit()
conn.close()
print("Éxito: Base de datos SQLite depurada y consolidada.")

# 3. Limpiar Transacciones_Acumuladas.xlsx
excel_raw = "downloads/data_cruda/Transacciones_Acumuladas.xlsx"
if os.path.exists(excel_raw):
    df_raw = pd.read_excel(excel_raw)
    dni_col = 'ID' if 'ID' in df_raw.columns else 'DNI'
    if dni_col in df_raw.columns:
        df_raw[dni_col] = df_raw[dni_col].apply(normalizar_dni)
        df_raw['Nombre'] = df_raw['Nombre'].apply(quitar_tildes)
        df_raw['Apellido'] = df_raw['Apellido'].apply(quitar_tildes)
        
        pos_col = 'Posición' if 'Posición' in df_raw.columns else ('Posicion' if 'Posicion' in df_raw.columns else 'Cargo')
        if pos_col in df_raw.columns:
            # Sincronizar cargos con tabla trabajadores
            for _, r in df_trab_clean.iterrows():
                mask = df_raw[dni_col] == r['dni']
                df_raw.loc[mask, pos_col] = r['cargo']
        
        df_raw.to_excel(excel_raw, index=False)
        print("Éxito: Transacciones_Acumuladas.xlsx normalizado.")

# 4. Regenerar Padron_Trabajadores_GZG.xlsx
padron_path = "Padron_Trabajadores_GZG.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Trabajadores"
ws.views.sheetView[0].showGridLines = True

fill_banner = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
font_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

font_data = Font(name="Calibri", size=10)
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
)

ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 28
ws["A1"] = "PADRÓN OFICIAL DE TRABAJADORES Y PERSONAL REGISTRADO - GZG MINERALES"
ws["A1"].fill = fill_banner
ws["A1"].font = font_banner
ws["A1"].alignment = align_center

ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 18
fill_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
font_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")
ws["A2"] = "GZG Minerales | Estado Actualizado de Personal en Biométrico y Sistema"
ws["A2"].fill = fill_sub
ws["A2"].font = font_sub
ws["A2"].alignment = align_center

headers = ["DNI", "Apellidos", "Nombres", "Departamento / Área", "Posición / Cargo", "Estado en Sistema"]
ws.row_dimensions[3].height = 25
ws.append(headers)

for cell in ws[3]:
    cell.fill = fill_header
    cell.font = font_header
    cell.alignment = align_center
    cell.border = thin_border

for idx, r in df_trab_clean.sort_values(by=['apellidos', 'nombres']).iterrows():
    dni = str(r['dni']).strip().zfill(8)
    ape = str(r['apellidos']).strip()
    nom = str(r['nombres']).strip()
    area = str(r['area']).strip()
    cargo = str(r['cargo']).strip()
    estado = "Activo"

    ws.append([dni, ape, nom, area, cargo, estado])
    c_row = ws.max_row
    ws.row_dimensions[c_row].height = 20

    for c_i in range(1, 7):
        cell = ws.cell(row=c_row, column=c_i)
        cell.font = font_data
        cell.border = thin_border
        if c_i in (1, 6):
            cell.alignment = align_center
            if c_i == 1:
                cell.number_format = '@'
        else:
            cell.alignment = align_left

widths = {1: 15, 2: 28, 3: 26, 4: 26, 5: 30, 6: 18}
for c_idx, w in widths.items():
    ws.column_dimensions[get_column_letter(c_idx)].width = w

wb.save(padron_path)
print(f"Éxito: Archivo Padron_Trabajadores_GZG.xlsx regenerado con exactamente {len(df_trab_clean)} trabajadores.")
