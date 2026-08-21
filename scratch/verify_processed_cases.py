import os
import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print("=== VERIFICACIÓN EN REPORTE EXCEL PROCESADO ===")
print(f"Fila Título: {ws['A1'].value}")
print(f"Subtítulo: {ws['A2'].value}\n")

headers = [cell.value for cell in ws[4]]
print("Encabezados:", headers)

print("\n--- CASO 1: IVAN ANTONIO VASQUEZ PUELLES (48455175) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    if row[0] == '48455175':
        print(f"Fecha Turno: {row[5]} | Turno: {row[7]} | Ent: {row[8]} {row[9]} | Sal: {row[10]} {row[11]} | Trab: {row[16]} | Tipo: {row[20]} | Obs: {row[21]}")

print("\n--- CASO 2: LUIS FERNANDO RAMIREZ GUERRERO (70088280) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    if row[0] == '70088280':
        print(f"Fecha Turno: {row[5]} | Turno: {row[7]} | Ent: {row[8]} {row[9]} | Sal: {row[10]} {row[11]} | Trab: {row[16]} | Tipo: {row[20]} | Obs: {row[21]}")
