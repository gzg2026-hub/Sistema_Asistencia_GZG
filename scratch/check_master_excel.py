import openpyxl

fpath = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Sistema_Asistencia_GZG_v1.0.xlsx"
wb = openpyxl.load_workbook(fpath, data_only=True)

print(f"File: {fpath}")
print(f"Sheets: {wb.sheetnames}")
ws = wb.active
for idx, row in enumerate(ws.iter_rows(values_only=True)):
    if idx < 10:
        print([str(c) for c in row if c is not None])
