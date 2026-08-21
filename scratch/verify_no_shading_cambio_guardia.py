import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE SOMBREADO EN {report_path} ===")

cg_count = 0
for row_idx in range(5, ws.max_row + 1):
    tipo_reg = str(ws.cell(row=row_idx, column=21).value or "")
    if "Cambio de guardia" in tipo_reg:
        cg_count += 1
        fill_obj = ws.cell(row=row_idx, column=1).fill
        fill_type = fill_obj.fill_type
        start_color = fill_obj.start_color.rgb if fill_obj and fill_obj.start_color else None
        print(f"  Fila {row_idx}: {ws.cell(row=row_idx, column=2).value} {ws.cell(row=row_idx, column=3).value} ({ws.cell(row=row_idx, column=6).value}) | Tipo: '{tipo_reg}' | Fill Type: {fill_type} | Fill Color: {start_color}")

print(f"\nTotal filas Cambio de guardia verificadas: {cg_count}")
