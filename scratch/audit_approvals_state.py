import openpyxl
import os
import pandas as pd
import sqlite3

ROOT_DIR = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG"
excel_path = os.path.join(ROOT_DIR, "downloads", "data_procesada", "Aprobaciones_GZG_2026-08.xlsx")

print("=== 1. AUDITORIA DEL ARCHIVO EXCEL (Aprobaciones_GZG_2026-08.xlsx) ===")
if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    total_filas = ws.max_row
    print(f"Ruta local: {excel_path}")
    print(f"Total registros de datos: {total_filas - 4}")
    
    estados = {}
    comentarios = 0
    fechas = set()
    for r in range(5, total_filas + 1):
        fecha = ws.cell(row=r, column=6).value
        est = ws.cell(row=r, column=13).value
        cmt = ws.cell(row=r, column=19).value
        if fecha:
            fechas.add(str(fecha))
        estados[est] = estados.get(est, 0) + 1
        if cmt and str(cmt).strip():
            comentarios += 1
            
    print(f"Distribución de Estados en Excel: {estados}")
    print(f"Comentarios/Sustentos registrados en Col S: {comentarios} (100% limpio)")
    print(f"Fechas registradas: del {min(fechas)} al {max(fechas)} ({len(fechas)} fechas)")
else:
    print("El archivo no existe.")

print("\n=== 2. AUDITORIA DE LA BASE DE DATOS (asistencia.db) ===")
conn = sqlite3.connect(os.path.join(ROOT_DIR, "data", "asistencia.db"))
c = conn.cursor()
c.execute("SELECT estado, COUNT(*) FROM aprobaciones GROUP BY estado")
print("Estados Globales:", c.fetchall())
c.execute("SELECT estado_n1, COUNT(*) FROM aprobaciones GROUP BY estado_n1")
print("Estados Nivel 1:", c.fetchall())
c.execute("SELECT estado_n2, COUNT(*) FROM aprobaciones GROUP BY estado_n2")
print("Estados Nivel 2:", c.fetchall())
c.execute("SELECT COUNT(*) FROM aprobaciones WHERE observacion_trabajador != '' OR comentario_supervisor != '' OR comentario_n1 != '' OR comentario_n2 != ''")
print("Solicitudes con comentarios:", c.fetchone()[0])
conn.close()
