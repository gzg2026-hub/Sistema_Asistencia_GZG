import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE DATA DESCARGADA Y PROCESADA EN {report_path} ===")
print(f"Total filas en reporte Excel: {ws.max_row} (Filas de datos: {ws.max_row - 4})")

dates = set()
for r_idx in range(5, ws.max_row + 1):
    f = ws.cell(row=r_idx, column=6).value
    if f:
        dates.add(str(f))

print(f"Fechas encontradas en el reporte: {sorted(list(dates))}")

print("\n--- ULTIMOS REGISTROS DEL 21/08/2026 ---")
count_21 = 0
for r_idx in range(5, ws.max_row + 1):
    f = str(ws.cell(row=r_idx, column=6).value or "")
    if "21/08/2026" in f or "2026-08-21" in f:
        count_21 += 1
        if count_21 <= 10:
            ap = ws.cell(row=r_idx, column=2).value
            nom = ws.cell(row=r_idx, column=3).value
            ent = ws.cell(row=r_idx, column=10).value
            sal = ws.cell(row=r_idx, column=12).value
            tipo = ws.cell(row=r_idx, column=22).value
            obs = ws.cell(row=r_idx, column=23).value
            print(f"  * {ap} {nom} | Ent: {ent} | Sal: {sal} | Tipo: '{tipo}' | Obs: '{obs}'")

print(f"Total marcaciones procesadas para el 21/08/2026: {count_21}")
