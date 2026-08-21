import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE FILAS 'Cambio de guardia' SIN SOMBREADO EN {report_path} ===")

cambio_count = 0
for r_idx in range(5, ws.max_row + 1):
    tipo_reg = str(ws.cell(row=r_idx, column=22).value or "").strip()
    if tipo_reg == "Cambio de guardia":
        cambio_count += 1
        # Verificar celdas de la fila
        fills = [ws.cell(row=r_idx, column=c).fill.fill_type for c in range(1, 24)]
        has_fill = any(f is not None for f in fills)
        obs = ws.cell(row=r_idx, column=23).value
        print(f"  Fila {r_idx:3d} | Tipo: '{tipo_reg}' | Obs: '{obs}' | Tiene Sombreado: {has_fill}")

print(f"\nTotal filas 'Cambio de guardia' verificadas: {cambio_count}")
