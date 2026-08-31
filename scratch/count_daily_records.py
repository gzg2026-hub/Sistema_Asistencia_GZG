import os, glob
import openpyxl
import pandas as pd

folder = r"downloads\data_procesada\diario"
files = sorted(glob.glob(os.path.join(folder, "Reporte_Asistencia_GZG_2026-08-*.xlsx")))

print(f"{'Archivo':<42} | {'Filas Datos':<12} | {'Asistencias':<12} | {'Horas Extras':<12} | {'Excesos':<10}")
print("-" * 95)

total_filas = 0
total_asist = 0
total_he = 0
total_exc = 0

for fpath in files:
    fname = os.path.basename(fpath)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb.active
    
    # Filas de datos (empiezan en fila 5, después del header que está en fila 4)
    # Header está en fila 4 (DNI, Apellidos, Nombres...)
    # Fila 1 es banner título, Fila 2 es subtítulo, Fila 3 es vacía/espacio, Fila 4 es headers
    data_rows = []
    for r in range(5, ws.max_row + 1):
        dni = ws.cell(row=r, column=1).value
        if dni is not None and str(dni).strip() != '':
            data_rows.append(r)
            
    # Leer con pandas para desglose
    df = pd.read_excel(fpath, header=3) # header en fila índice 3 (fila 4 de excel)
    
    # Limpiar columnas
    n_filas = len(data_rows)
    total_filas += n_filas
    
    # Contar tipos
    tipo_col = [c for c in df.columns if 'Tipo' in str(c) or 'Registro' in str(c)]
    he_col = [c for c in df.columns if 'Horas Extras' in str(c)]
    exc_col = [c for c in df.columns if 'Exceso' in str(c)]
    
    n_he = 0
    n_exc = 0
    if he_col:
        n_he = len(df[df[he_col[0]].notna() & (df[he_col[0]].astype(str).str.strip().isin(['00:00', '0.0', '0', 'nan']) == False)])
    if exc_col:
        n_exc = len(df[df[exc_col[0]].notna() & (df[exc_col[0]].astype(str).str.strip().isin(['00:00', '0.0', '0', 'nan']) == False)])
        
    total_he += n_he
    total_exc += n_exc
    
    print(f"{fname:<42} | {n_filas:<12} | {n_filas:<12} | {n_he:<12} | {n_exc:<10}")
    wb.close()

print("-" * 95)
print(f"{'TOTAL ACUMULADO EN LOS 8 ARCHIVOS':<42} | {total_filas:<12} | {total_filas:<12} | {total_he:<12} | {total_exc:<10}")
