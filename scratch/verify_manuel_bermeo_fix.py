import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21_121034.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print("=== VERIFICACIÓN MANUEL BERMEO OJEDA (IMAGEN 2) ===")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).strip() if row[1] is not None else ""
    if "BERMEO" in apellidos and "MANUEL" in str(row[2]):
        vals = [cell for cell in row]
        print(f"Fecha: {vals[5]} | Ent: {vals[8]} {vals[9]} | Sal: {vals[10]} {vals[11]} | Trab: {vals[16]} | Tard: {vals[17]} | Tipo: {vals[20]} | Obs: {vals[21]}")

print("\n=== VERIFICACIÓN DE SOMBREADO PARA SALIDA ANTICIPADA (PUNTO 1) ===")
for r_idx in range(5, ws.max_row + 1):
    tipo_reg = str(ws.cell(row=r_idx, column=21).value or '')
    obs = str(ws.cell(row=r_idx, column=22).value or '')
    if "salida anticipada" in (tipo_reg + " " + obs).lower():
        cell_fill = ws.cell(row=r_idx, column=1).fill.start_color.rgb
        apellidos = ws.cell(row=r_idx, column=2).value
        nombres = ws.cell(row=r_idx, column=3).value
        fecha = ws.cell(row=r_idx, column=6).value
        print(f"Trabajador: {apellidos} {nombres} ({fecha}) | Fill RGB: {cell_fill} | Obs: {obs}")
