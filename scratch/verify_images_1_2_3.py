import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE LAS 3 IMAGENES EN {report_path} ===")

print("\n--- 1. VERIFICACION IMAGEN 1: 'Exceso de Jornada' CON 'J' MAYUSCULA ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    col_u = str(row[20])
    col_v = str(row[21])
    if "exceso" in col_u.lower() or "exceso" in col_v.lower():
        print(f"  * {row[1]} {row[2]} ({row[5]}): Col U: '{col_u}' | Col V: '{col_v}'")

print("\n--- 2. VERIFICACION IMAGEN 2: CASO ALVA MEDINA JHON KENEDY (18/08/2026) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    apellidos = str(row[1]).upper()
    if "ALVA MEDINA" in apellidos:
        print(f"  * {row[1]} {row[2]} ({row[5]}): Col U (Tipo): '{row[20]}' | Col V (Obs): '{row[21]}'")

print("\n--- 3. VERIFICACION IMAGEN 3: OBSERVACION JORNADA PARCIAL COMO 'Cambio de guardia (Medio día) (hh:mm)' ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    tipo = str(row[20])
    obs = str(row[21])
    if tipo == "Jornada parcial":
        print(f"  * {row[1]} {row[2]} ({row[5]}): Tipo: '{tipo}' | Obs: '{obs}'")
