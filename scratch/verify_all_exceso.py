import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION GENERAL DE EXCESO DE TURNO EN {report_path} ===")

count = 0
for r_idx in range(5, ws.max_row + 1):
    exc = ws.cell(row=r_idx, column=18).value
    if exc and str(exc) != "00:00":
        count += 1
        dni = ws.cell(row=r_idx, column=1).value
        ap = ws.cell(row=r_idx, column=2).value
        nom = ws.cell(row=r_idx, column=3).value
        f = ws.cell(row=r_idx, column=6).value
        ent = ws.cell(row=r_idx, column=10).value
        sal = ws.cell(row=r_idx, column=12).value
        ht = ws.cell(row=r_idx, column=17).value
        tot = ws.cell(row=r_idx, column=20).value
        tipo = ws.cell(row=r_idx, column=22).value
        obs = ws.cell(row=r_idx, column=23).value
        print(f"Fila {r_idx:3d} | {ap} {nom} ({f}): Ent={ent} | Sal={sal} | HT={ht} | Exceso={exc} | Total={tot} | Tipo='{tipo}' | Obs='{obs}'")

print(f"\nTotal filas con Exceso de Turno detectadas: {count}")
