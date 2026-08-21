import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE ESTRUCTURA DE 23 COLUMNAS EN {report_path} ===")

# Encabezados Fila 4
headers = [ws.cell(row=4, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
print(f"\nTotal columnas detectadas en Fila 4: {len(headers)}")
for idx, h in enumerate(headers, 1):
    print(f"  * Columna {idx} ({openpyxl.utils.get_column_letter(idx)}): '{h}'")

print("\n--- MUESTRA DE DATOS PROCESADOS (FILAS 5 A 15) ---")
for r_idx in range(5, 16):
    vals = [ws.cell(row=r_idx, column=c).value for c in range(1, 24)]
    print(f"Fila {r_idx:2d} | {vals[1]} {vals[2]} ({vals[5]}):")
    print(f"   Q (Horas de Turno): {vals[16]} | R (Exceso de Turno): {vals[17]} | S (Horas Extras): {vals[18]} | T (Total Adicionales): {vals[19]} | U (Tardanza): {vals[20]}")
    print(f"   V (Tipo Registro): '{vals[21]}' | W (Observaciones): '{vals[22]}'\n")

print("\n--- VERIFICACION CASOS CON HORAS EXTRAS BIOMETRICAS (COLUMNA S) ---")
for r_idx in range(5, ws.max_row + 1):
    he_val = str(ws.cell(row=r_idx, column=19).value or "00:00")
    if he_val != "00:00":
        apellidos = ws.cell(row=r_idx, column=2).value
        nombres = ws.cell(row=r_idx, column=3).value
        fecha = ws.cell(row=r_idx, column=6).value
        exc_val = ws.cell(row=r_idx, column=18).value
        tot_val = ws.cell(row=r_idx, column=20).value
        tipo_reg = ws.cell(row=r_idx, column=22).value
        obs = ws.cell(row=r_idx, column=23).value
        print(f"  * {apellidos} {nombres} ({fecha}): Exceso (R)={exc_val} + HE (S)={he_val} -> Total Adic (T)={tot_val} | Tipo: '{tipo_reg}' | Obs: '{obs}'")
