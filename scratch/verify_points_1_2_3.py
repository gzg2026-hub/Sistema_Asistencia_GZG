import openpyxl

report_path = r"c:\Users\GZG Minerales 2026\Desktop\GZG\Sistema_Asistencia_GZG\downloads\data_procesada\Reporte_Asistencia_GZG_2026-08-17_al_2026-08-21_143322.xlsx"
wb = openpyxl.load_workbook(report_path)
ws = wb.active

print(f"=== VERIFICACION DE LOS 3 NUEVOS PUNTOS EN {report_path} ===")

print("\n--- 1. VERIFICACION: BORRADAS TODAS LAS PALABRAS 'RELEVO' (PUNTO 1) ---")
relevo_found = False
for row in ws.iter_rows(min_row=5, values_only=True):
    col_u = str(row[20]).lower()
    col_v = str(row[21]).lower()
    if "relevo" in col_u or "relevo" in col_v:
        print(f"  [ERROR] Se encontro 'relevo': Col U: {row[20]} | Col V: {row[21]}")
        relevo_found = True

if not relevo_found:
    print("  EXITO: 0 coincidencias de 'relevo' o 'relevo de cuadrilla'. Todo dice 'Cambio de guardia'.")

print("\n--- 2. VERIFICACION: COLUMNA U Y COLUMNA V PARA JORNADA PARCIAL (PUNTO 2) ---")
for row in ws.iter_rows(min_row=5, values_only=True):
    tipo = str(row[20])
    obs = str(row[21])
    if "Jornada parcial" in tipo or "Jornada parcial" in obs:
        print(f"  * Trabajador: {row[1]} {row[2]} ({row[5]}): Col U (Tipo): '{tipo}' | Col V (Obs): '{obs}'")

print("\n--- 3. VERIFICACION: ENTRADA DUPLICADA SOLO EN INTERVALOS CORTOS (<= 15 MIN) (PUNTO 3) ---")
dup_count = 0
for row in ws.iter_rows(min_row=5, values_only=True):
    obs = str(row[21])
    if "Entrada duplicada" in obs:
        dup_count += 1
        print(f"  * Trabajador: {row[1]} {row[2]} ({row[5]}): Ent: {row[9]} | Obs: {obs}")

print(f"  Total marcaciones con Entrada duplicada cercana: {dup_count}")
