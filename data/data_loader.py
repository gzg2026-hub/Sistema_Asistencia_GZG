import openpyxl
import pandas as pd
import re
from datetime import datetime, time
from typing import Tuple, Dict, List

def parse_hikvision_transaction_file(excel_path_or_file) -> pd.DataFrame:
    """
    Parsea archivos descargados de transacciones de Hikvision (ej. Transacciones_2026-08-11_2026-08-11.xlsx).
    Maneja filas de metadatos superiores, encabezados de fecha (Fecha:2026-08-11) y mapea las columnas.
    """
    try:
        wb = openpyxl.load_workbook(excel_path_or_file, data_only=True)
    except Exception as e:
        print(f"Error al abrir Excel con openpyxl: {e}")
        return pd.DataFrame()
        
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    
    current_date = None
    header_idx = None
    data_rows = []
    headers = []
    
    # Expresión regular para buscar fechas en filas de título/sección
    date_regex = re.compile(r'Fecha\s*:\s*(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4})', re.IGNORECASE)
    period_regex = re.compile(r'Periodo\s*:\s*(\d{4}-\d{2}-\d{2})\s*-\s*(\d{4}-\d{2}-\d{2})', re.IGNORECASE)
    
    for idx, row in enumerate(rows):
        row_vals = [str(cell).strip() if cell is not None else "" for cell in row]
        row_str = " ".join(row_vals)
        
        # Buscar fecha en la fila
        match_date = date_regex.search(row_str)
        if match_date:
            current_date = match_date.group(1)
        elif not current_date:
            match_period = period_regex.search(row_str)
            if match_period:
                current_date = match_period.group(1)
                
        # Identificar la fila de encabezados
        if 'ID' in row_vals and ('Tiempo' in row_vals or 'Tipo de pase de tarjeta' in row_vals):
            header_idx = idx
            headers = [v for v in row_vals if v]
            continue
            
        # Si ya pasamos los encabezados y la fila contiene datos válidos
        if header_idx is not None and idx > header_idx:
            first_val = row_vals[0] if len(row_vals) > 0 else ""
            first_val_clean = first_val.lower().strip()
            
            # Omitir filas vacías, repetidas de encabezado o banners como "Fecha:2026-08-02 Semana:Domingo"
            if not first_val or first_val_clean == 'id' or first_val_clean == 'none' or 'fecha:' in first_val_clean or 'semana:' in first_val_clean or 'periodo:' in first_val_clean:
                continue
                
            row_dict = {}
            for col_idx, col_name in enumerate(headers):
                val = row_vals[col_idx] if col_idx < len(row_vals) else ""
                row_dict[col_name] = val
                    
            # Asignar la fecha extraída
            row_dict['Fecha'] = current_date if current_date else datetime.now().strftime("%Y-%m-%d")
            
            # Limpiar departamento (solo texto después del >)
            dept_val = str(row_dict.get('Departamento', '')).strip()
            if '>' in dept_val:
                row_dict['Departamento'] = dept_val.split('>')[-1].strip()
            else:
                row_dict['Departamento'] = dept_val
                
            data_rows.append(row_dict)
                
    if not data_rows:
        return pd.DataFrame()
        
    df = pd.DataFrame(data_rows)
    official_cols = [
        'ID', 'Fecha', 'Nombre', 'Apellido', 'Cargo', 'Departamento', 'Grupo de asistencia',
        'Tiempo', 'Tipo de pase de tarjeta', 'Método de verificación',
        'Punto de control de asistencia'
    ]
    # Reordenar según columnas oficiales si están presentes
    present_cols = [c for c in official_cols if c in df.columns]
    other_cols = [c for c in df.columns if c not in official_cols]
    df = df[present_cols + other_cols]
    return df

def cargar_datos_excel(excel_source) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Lee las pestañas 01_TRABAJADORES, 02_MARCACIONES, 04_HORAS_EXTRA del Excel base
    o parsea un archivo crudo de transacciones de Hikvision.
    """
    try:
        wb = openpyxl.load_workbook(excel_source, data_only=True)
        sheetnames = wb.sheetnames
    except Exception:
        sheetnames = []
        
    df_trabajadores = pd.DataFrame()
    df_marcaciones = pd.DataFrame()
    df_horas_extra = pd.DataFrame()

    # Si es el Excel base estructurado (Sistema_Asistencia_GZG_v1.0.xlsm)
    if '01_TRABAJADORES' in sheetnames:
        ws_trab = wb['01_TRABAJADORES']
        data = list(ws_trab.iter_rows(values_only=True))
        if data and len(data) > 1:
            cols = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(data[0])]
            df_trabajadores = pd.DataFrame(data[1:], columns=cols).dropna(how='all')
            # Eliminar columna auxiliar DNI_STR si existe
            if 'DNI_STR' in df_trabajadores.columns:
                df_trabajadores = df_trabajadores.drop(columns=['DNI_STR'])

    if '02_MARCACIONES' in sheetnames:
        ws_mar = wb['02_MARCACIONES']
        data = list(ws_mar.iter_rows(values_only=True))
        if data and len(data) > 1:
            cols = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(data[0])]
            df_marcaciones = pd.DataFrame(data[1:], columns=cols).dropna(how='all')

    if '04_HORAS_EXTRA' in sheetnames or '05_HORAS_EXTRA' in sheetnames:
        sheet_name = '04_HORAS_EXTRA' if '04_HORAS_EXTRA' in sheetnames else '05_HORAS_EXTRA'
        ws_he = wb[sheet_name]
        data = list(ws_he.iter_rows(values_only=True))
        if data and len(data) > 1:
            cols = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(data[0])]
            df_horas_extra = pd.DataFrame(data[1:], columns=cols).dropna(how='all')

    # Si df_marcaciones está vacío pero se cargó un Excel de transacciones de Hikvision
    if df_marcaciones.empty:
        df_marcaciones = parse_hikvision_transaction_file(excel_source)

    return df_trabajadores, df_marcaciones, df_horas_extra
