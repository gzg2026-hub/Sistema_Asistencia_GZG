"""
exporter.py
===========
Genera el reporte consolidado de asistencia en un ÚNICO Excel de una sola pestaña
("Asistencia y Horas Extras") con el formato ejecutivo oficial limpio de 23 columnas
(sin ninguna columna de punto de control).
"""

import io
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import streamlit as st
from datetime import datetime
from typing import Optional, Dict

BASE_EXCEL_TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Sistema_Asistencia_GZG_v1.0.xlsx")

DIAS_SEMANA = {
    0: "Lunes", 1: "Martes", 2: "Miércoles",
    3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"
}


def quitar_tildes(texto: str) -> str:
    if not isinstance(texto, str) or not texto or str(texto).strip().lower() in ('nan', 'none', ''):
        return ""
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U',
        'á': 'A', 'é': 'E', 'í': 'I', 'ó': 'O', 'ú': 'U', 'ü': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'à': 'A', 'è': 'E', 'ì': 'I', 'ò': 'O', 'ù': 'U',
    }
    res = str(texto)
    for k, v in replacements.items():
        res = res.replace(k, v)
    return res.strip()


def format_date_ddmmyyyy(date_val) -> str:
    """Convierte cualquier representación de fecha a formato DD/MM/YYYY (ej. 18/08/2026)."""
    if pd.isna(date_val) or date_val is None or date_val == "" or str(date_val).strip() == "-":
        return "-" if str(date_val).strip() == "-" else ""
    val_str = str(date_val).strip().split(' ')[0]
    try:
        if '-' in val_str:
            parts = val_str.split('-')
            if len(parts[0]) == 4:  # YYYY-MM-DD -> DD/MM/YYYY
                return f"{int(parts[2]):02d}/{int(parts[1]):02d}/{parts[0]}"
            elif len(parts[2]) == 4:  # DD-MM-YYYY
                return f"{int(parts[0]):02d}/{int(parts[1]):02d}/{parts[2]}"
        elif '/' in val_str:
            parts = val_str.split('/')
            if len(parts[2]) == 4:  # DD/MM/YYYY
                return f"{int(parts[0]):02d}/{int(parts[1]):02d}/{parts[2]}"
    except Exception:
        pass
    return val_str


def format_hhmm_cell(val, is_hours_float=False) -> str:
    """Convierte minutos enteros u horas flotantes a string HH:MM (ej. 11:51, 00:15)."""
    if pd.isna(val) or val is None or val == "":
        return "00:00"
    val_str = str(val).strip()
    if ":" in val_str and len(val_str.split(":")) == 2:
        return val_str
    try:
        num = float(val)
        if num <= 0:
            return "00:00"
        total_min = int(round(num * 60.0)) if is_hours_float else int(round(num))
        h = total_min // 60
        m = total_min % 60
        return f"{h:02d}:{m:02d}"
    except Exception:
        return "00:00"

# ... (resto de funciones)



@st.cache_data(ttl=60, show_spinner=False)
def exportar_asistencia_excel(
    df_trabajadores: pd.DataFrame,
    df_marcaciones: pd.DataFrame,
    df_asistencia: pd.DataFrame,
    df_horas_extra: pd.DataFrame = None,
    df_incidencias: pd.DataFrame = None,
    template_path: str = BASE_EXCEL_TEMPLATE
) -> bytes:
    """
    Genera el archivo Excel procesado oficial de UNA SOLA HOJA ('Asistencia y Horas Extras')
    con 23 columnas limpias (sin ninguna columna de punto de control).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asistencia y Horas Extras"
    ws.views.sheetView[0].showGridLines = True

    # 1. Rango de Fechas para el banner
    f_min = "2026-08-17"
    f_max = "2026-08-18"
    if df_asistencia is not None and not df_asistencia.empty and 'FECHA' in df_asistencia.columns:
        fechas_val = df_asistencia['FECHA'].dropna().unique()
        if len(fechas_val) > 0:
            f_min = str(min(fechas_val))
            f_max = str(max(fechas_val))

    # Estilos profesionales openpyxl
    fill_banner_title = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    font_banner_title = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

    fill_banner_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_banner_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    # Estilo Teal/Cian para las 4 columnas de cálculo (Fila 4: Horas de Turno, Exceso de Turno, Horas Extras, Total de Horas Adicionales)
    fill_header_calc = PatternFill(start_color="317F96", end_color="317F96", fill_type="solid")       # Azul Teal / Cian Oceánico
    font_header_calc = Font(name="Calibri", size=10, bold=True, color="FFFFFF")                       # Texto Blanco Bold

    font_data = Font(name="Calibri", size=10)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Banner Título (Fila 1 - 23 columnas A a W)
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:W1")
    ws["A1"] = "REPORTE DE ASISTENCIA Y HORAS EXTRAS PROCESADO (TURNOS DÍA Y NOCHE)"
    ws["A1"].fill = fill_banner_title
    ws["A1"].font = font_banner_title
    ws["A1"].alignment = align_center

    # Banner Subtítulo (Fila 2)
    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:W2")
    ws["A2"] = f"GZG Minerales | Período: {f_min} a {f_max} | Incluye Entrada, Salida, Inicio H.E. y Fin H.E."
    ws["A2"].fill = fill_banner_sub
    ws["A2"].font = font_banner_sub
    ws["A2"].alignment = align_center

    ws.row_dimensions[3].height = 10
    ws.append([])  # Fila 3 vacía de separación

    # Encabezados de Columna (Fila 4 - 23 columnas: A a W)
    ws.row_dimensions[4].height = 28
    headers = [
        "DNI", "Apellidos", "Nombres", "Departamento", "Posición",
        "Fecha Turno", "Día", "Turno", "Fecha Entrada", "Hora Entrada",
        "Fecha Salida", "Hora Salida",
        "Fecha Inicio H.E.", "Inicio H.E.",
        "Fecha Fin H.E.", "Fin H.E.",
        "Horas de Turno", "Tardanza", "Exceso de Turno", "Horas Extras",
        "Total de Horas Adicionales", "Tipo Registro", "Observación / Incidencias"
    ]

    ws.append(headers)  # Fila 4

    for c_idx, cell in enumerate(ws[4], 1):
        if 17 <= c_idx <= 21:  # Columnas 17 a 21 (Q a U: Horas de Turno, Tardanza, Exceso de Turno, Horas Extras, Total de Horas Adicionales)
            cell.fill = fill_header_calc
            cell.font = font_header_calc
        else:
            cell.fill = fill_header
            cell.font = font_header
        cell.alignment = align_center

    # Identificar trabajadores con doble turno / doble entrada en la misma fecha
    doble_turno_keys = set()
    if df_asistencia is not None and not df_asistencia.empty and 'DNI' in df_asistencia.columns and 'FECHA' in df_asistencia.columns:
        dups = df_asistencia[df_asistencia.duplicated(subset=['DNI', 'FECHA'], keep=False)]
        for _, d_row in dups.iterrows():
            doble_turno_keys.add((str(d_row['DNI']).strip(), str(d_row['FECHA']).strip()))

    # Escribir filas procesadas
    if df_asistencia is not None and not df_asistencia.empty:
        for _, row in df_asistencia.iterrows():
            dni = str(row.get('DNI', '')).strip()
            apellidos = quitar_tildes(str(row.get('APELLIDOS', '')))
            nombres = quitar_tildes(str(row.get('NOMBRES', '')))
            dept = str(row.get('ÁREA', row.get('Departamento', ''))).strip()
            posicion = str(row.get('CARGO', row.get('Posición', ''))).strip()
            fecha_t = str(row.get('FECHA', '')).strip()

            # Día de la semana
            dia_str = ""
            try:
                dt_obj = datetime.strptime(fecha_t, "%Y-%m-%d")
                dia_str = DIAS_SEMANA.get(dt_obj.weekday(), "")
            except Exception:
                dia_str = ""

            turno = str(row.get('TURNO', 'DIA')).strip()
            h_ent = str(row.get('ENTRADA', '')).strip() if pd.notna(row.get('ENTRADA')) else ""
            h_sal = str(row.get('SALIDA', '')).strip() if pd.notna(row.get('SALIDA')) else ""

            f_ent = str(row.get('FECHA_ENTRADA', fecha_t if h_ent else '')).strip()
            f_sal = str(row.get('FECHA_SALIDA', fecha_t if h_sal else '')).strip()

            # Marcaciones H.E.
            f_ini_he = str(row.get('FECHA_INICIO_HE', '-')).strip()
            h_ini_he = str(row.get('HORA_INICIO_HE', '-')).strip()
            f_fin_he = str(row.get('FECHA_FIN_HE', '-')).strip()
            h_fin_he = str(row.get('HORA_FIN_HE', '-')).strip()

            # Regla Punto 9: Eliminar filas cuando la doble columna I (Fecha Entrada) y K (Fecha Salida) sean NAN o vacías, salvo que sea un bloque de Horas Extras
            has_no_entry_exit = (not f_ent or f_ent.lower() in ('nan', 'none', '', '-')) and (not f_sal or f_sal.lower() in ('nan', 'none', '', '-')) and not h_ent and not h_sal
            has_no_he = (not h_ini_he or h_ini_he.lower() in ('nan', 'none', '', '-')) and (not h_fin_he or h_fin_he.lower() in ('nan', 'none', '', '-'))
            if has_no_entry_exit and has_no_he:
                continue

            # Función flexible para extraer valores por nombres aproximados de columna
            def get_row_val(row_obj, *keys):
                for k in keys:
                    if k in row_obj and pd.notna(row_obj[k]):
                        return row_obj[k]
                return '00:00'

            val_h_turno = get_row_val(row, 'HORAS DE TURNO (HH:MM)', 'HORAS DE TURNO (hh:mm)', 'HORAS TRABAJADAS (HH:MM)', 'HORAS TRABAJADAS')
            val_exc_turno = get_row_val(row, 'EXCESO DE TURNO (HH:MM)', 'EXCESO DE TURNO (hh:mm)', 'EXCESO JORNADA (HH:MM)', 'EXCESO JORNADA')
            val_he_explicita = get_row_val(row, 'HORAS EXTRAS (HH:MM)', 'HORAS EXTRAS (hh:mm)', 'HORAS EXTRAS')
            val_tot_adic = get_row_val(row, 'TOTAL DE HORAS ADICIONALES (HH:MM)', 'TOTAL DE HORAS ADICIONALES (hh:mm)', 'TOTAL HORAS ADICIONALES', 'HORAS EXTRAS TOTALES (hh:mm)')
            val_tard = get_row_val(row, 'TARDANZA (HH:MM)', 'TARDANZA (hh:mm)', 'TARDANZA (MIN)')

            h_turno = format_hhmm_cell(val_h_turno, is_hours_float=True)
            exc_turno = format_hhmm_cell(val_exc_turno, is_hours_float=False)
            he_explicita = format_hhmm_cell(val_he_explicita, is_hours_float=False)
            tot_adic = format_hhmm_cell(val_tot_adic, is_hours_float=False)
            tard = format_hhmm_cell(val_tard, is_hours_float=False)

            fecha_t_formatted = format_date_ddmmyyyy(fecha_t)
            f_ent_formatted = format_date_ddmmyyyy(f_ent)
            f_sal_formatted = format_date_ddmmyyyy(f_sal)
            f_ini_he_formatted = format_date_ddmmyyyy(f_ini_he)
            f_fin_he_formatted = format_date_ddmmyyyy(f_fin_he)

            incid = str(row.get('INCIDENCIAS', '')).strip()
            tipo_reg = str(row.get('TIPO_REGISTRO', 'Normal')).strip()

            ws.append([
                dni, apellidos, nombres, dept, posicion,
                fecha_t_formatted, dia_str, turno,
                f_ent_formatted, h_ent,
                f_sal_formatted, h_sal,
                f_ini_he_formatted, h_ini_he,
                f_fin_he_formatted, h_fin_he,
                h_turno, tard, exc_turno, he_explicita, tot_adic,
                tipo_reg, incid
            ])

            # Colores pastel de sombreado
            fill_jornada_parcial = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo Pastel
            fill_cambio_turno = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")    # Azul Pastel
            fill_incidencia = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # Durazno Pastel

            # Aplicar bordes, fuente, alineaciones a la nueva fila (23 columnas: A a W)
            current_row = ws.max_row
            ws.row_dimensions[current_row].height = 20
            es_doble_turno = (dni, fecha_t) in doble_turno_keys
            
            for c_idx in range(1, 24):
                cell = ws.cell(row=current_row, column=c_idx)
                cell.font = font_data
                cell.border = thin_border
                cell.alignment = align_center if c_idx not in (2, 3, 4, 5, 23) else align_left
                
                # Resaltado Pastel:
                # - Durazno Pastel (#FCE4D6): Doble Turno, Pendientes, Faltas, Sin Registro, Salidas Anticipadas.
                # - Azul Pastel (#D9E1F2): Horas Extras, H.E., Exceso de Jornada / Horas Adicionales.
                comb_check = (tipo_reg + " " + incid).lower()
                has_he = ("horas extra" in comb_check or "h.e." in comb_check or "exceso" in comb_check or "adicional" in comb_check or he_explicita != '00:00' or tot_adic != '00:00' or exc_turno != '00:00')
                
                if es_doble_turno or "doble" in comb_check or "reingreso" in comb_check:
                    cell.fill = fill_incidencia
                elif "pendiente" in comb_check or "falta" in comb_check or "sin registro" in comb_check or "salida anticipada" in comb_check or "cambio de guardia" in comb_check or "jornada parcial" in comb_check:
                    cell.fill = fill_incidencia
                elif has_he:
                    cell.fill = fill_cambio_turno

                # Formato de celda DNI como Texto '@'
                if c_idx == 1:
                    cell.number_format = '@'

    # Anchos de columna holgados y proporcionales (23 columnas: A a W)
    PROPORTIONAL_WIDTHS = {
        1: 15,   # DNI
        2: 26,   # Apellidos
        3: 24,   # Nombres
        4: 28,   # Departamento
        5: 24,   # Posición
        6: 15,   # Fecha Turno
        7: 13,   # Día
        8: 12,   # Turno
        9: 15,   # Fecha Entrada
        10: 14,  # Hora Entrada
        11: 15,  # Fecha Salida
        12: 14,  # Hora Salida
        13: 16,  # Fecha Inicio H.E.
        14: 15,  # Inicio H.E.
        15: 16,  # Fecha Fin H.E.
        16: 15,  # Fin H.E.
        17: 20,  # Horas de Turno
        18: 20,  # Exceso de Turno
        19: 18,  # Horas Extras
        20: 25,  # Total de Horas Adicionales
        21: 16,  # Tardanza
        22: 22,  # Tipo Registro
        23: 50   # Observación / Incidencias
    }

    for col_idx, width in PROPORTIONAL_WIDTHS.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def guardar_excel_base(
    df_trabajadores: pd.DataFrame,
    df_marcaciones: pd.DataFrame,
    df_asistencia: pd.DataFrame,
    df_horas_extra: pd.DataFrame = None,
    df_incidencias: pd.DataFrame = None,
    target_path: str = BASE_EXCEL_TEMPLATE
) -> bool:
    """Guarda el reporte procesado en la raíz y en downloads/data_procesada."""
    excel_bytes = exportar_asistencia_excel(
        df_trabajadores, df_marcaciones, df_asistencia, df_horas_extra, df_incidencias, target_path
    )
    
    # 1. Guardar en carpeta raíz
    success = False
    try:
        with open(target_path, "wb") as f:
            f.write(excel_bytes)
        success = True
    except PermissionError:
        ts = datetime.now().strftime("%H%M%S")
        alt_path = target_path.replace(".xlsx", f"_{ts}.xlsx")
        try:
            with open(alt_path, "wb") as f:
                f.write(excel_bytes)
            print(f"[Warn] Archivo en uso. Guardado como: '{alt_path}'")
            success = True
        except Exception:
            pass
    except Exception as e:
        print(f"[Error] Error al guardar Excel base: {e}")

    return success


def guardar_transacciones_acumuladas_excel(df: pd.DataFrame, target_path: str) -> bool:
    """
    Guarda Transacciones_Acumuladas.xlsx aplicando el formato ejecutivo corporativo:
    - Encabezados Azul Corporativo (#1F4E78) con texto Blanco en negrita.
    - Anchos de columna auto-ajustados y cuadrículas activas.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transacciones"
        ws.views.sheetView[0].showGridLines = True

        fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

        font_data = Font(name="Calibri", size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        headers = list(df.columns)
        ws.row_dimensions[1].height = 28
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = align_header
            cell.border = thin_border

        center_cols = {'ID', 'Fecha', 'Semana', 'Tiempo', 'Tipo de pase de tarjeta', 'Método de verificación'}

        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            ws.row_dimensions[row_idx].height = 20
            for col_idx, val in enumerate(row_data, start=1):
                col_name = headers[col_idx - 1]
                val_str = "" if pd.isna(val) else str(val).strip()
                if col_name in ('ID', 'DNI') and val_str:
                    if val_str.endswith(".0"):
                        val_str = val_str[:-2]
                    if val_str.isdigit():
                        val_str = val_str.zfill(8)
                c = ws.cell(row=row_idx, column=col_idx, value=val_str)
                c.font = font_data
                c.border = thin_border
                if col_name in ('ID', 'DNI'):
                    c.number_format = '@'
                if col_name in center_cols:
                    c.alignment = align_center
                else:
                    c.alignment = align_left

        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 6, 16)

        ws.auto_filter.ref = ws.dimensions

        try:
            wb.save(target_path)
            return True
        except PermissionError:
            print(f"[Aviso] No se pudo sobrescribir '{target_path}' porque está abierto en Excel. Ciérralo para guardar.")
            return False
    except Exception as e:
        print(f"[Error] Error al guardar Transacciones_Acumuladas.xlsx formateado: {e}")
        return False


def exportar_aprobaciones_excel(df_aprobaciones: pd.DataFrame, target_path: str) -> bool:
    """
    Genera el Excel oficial de Aprobaciones de HE y Excesos de Jornada con formato
    corporativo Azul Oscuro (#1F4E78) en downloads/data_procesada/Aprobaciones_GZG_YYYY-MM.xlsx.
    Se actualiza automáticamente cada vez que se aprueba/rechaza una solicitud en el app.
    REGLA: Usa OBLIGATORIAMENTE esta función, nunca pandas.to_excel directo.
    """
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aprobaciones HE"
        ws.views.sheetView[0].showGridLines = True

        fill_header_dark = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_header_calc = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        fill_banner = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        fill_approved = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_rejected = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_pending = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        font_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        font_banner_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")
        font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=10)
        font_estado_ok = Font(name="Calibri", size=10, bold=True, color="375623")
        font_estado_no = Font(name="Calibri", size=10, bold=True, color="843C0C")
        font_estado_pend = Font(name="Calibri", size=10, bold=True, color="7F6000")

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="D9D9D9")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Fila 1: Banner
        ws.merge_cells("A1:S1")
        ws.row_dimensions[1].height = 30
        mes_str = ""
        if df_aprobaciones is not None and not df_aprobaciones.empty and 'fecha' in df_aprobaciones.columns:
            try:
                mes_str = f" — {str(df_aprobaciones['fecha'].iloc[0])[:7]}"
            except Exception:
                pass
        ws["A1"] = f"REGISTRO DE APROBACIONES DE HORAS EXTRAS Y EXCESOS DE JORNADA - GZG MINERALES{mes_str}"
        ws["A1"].fill = fill_banner
        ws["A1"].font = font_banner
        ws["A1"].alignment = align_center

        # Fila 2: Sub-banner
        ws.merge_cells("A2:S2")
        ws.row_dimensions[2].height = 18
        fill_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws["A2"] = "GZG Minerales | Sistema Integrado de Control de Asistencia y Aprobaciones v1.0 | Generado automaticamente"
        ws["A2"].fill = fill_sub
        ws["A2"].font = font_banner_sub
        ws["A2"].alignment = align_center

        # Fila 3: separación
        ws.append([])
        ws.row_dimensions[3].height = 8

        # Fila 4: Encabezados (19 columnas A..S)
        ws.row_dimensions[4].height = 30
        headers = [
            "DNI", "Apellidos", "Nombres", "Cargo", "Area",
            "Fecha Turno", "Turno", "Hora Entrada", "Hora Salida",
            "Horas Trabajadas", "Horas Extras", "Exceso Jornada",
            "Estado Final", "Aprobador N1", "Estado N1",
            "Aprobador N2", "Estado N2",
            "Fecha Aprobacion", "Comentario Supervisor"
        ]
        ws.append(headers)
        for c_idx, cell in enumerate(ws[4], 1):
            cell.fill = fill_header_calc if c_idx in (11, 12) else fill_header_dark
            cell.font = font_header
            cell.alignment = align_center
            cell.border = thin_border

        # Filas de datos
        if df_aprobaciones is not None and not df_aprobaciones.empty:
            for _, row in df_aprobaciones.iterrows():
                dni_val = str(row.get('dni', '')).strip()
                if dni_val.endswith('.0'):
                    dni_val = dni_val[:-2]
                if dni_val.isdigit():
                    dni_val = dni_val.zfill(8)

                estado = str(row.get('estado', 'PENDIENTE')).strip().upper()
                if estado in ('NAN', 'NONE', ''):
                    estado = 'PENDIENTE'

                aprob_n1 = str(row.get('aprobador_n1', '') or '').strip()
                if aprob_n1.lower() in ('nan', 'none', ''):
                    aprob_n1 = '-'

                est_n1 = str(row.get('estado_n1', 'PENDIENTE') or 'PENDIENTE').strip().upper()
                if est_n1 in ('NAN', 'NONE', ''):
                    est_n1 = 'PENDIENTE'

                aprob_n2 = str(row.get('aprobador_n2', '') or '').strip()
                if aprob_n2.lower() in ('nan', 'none', ''):
                    aprob_n2 = '-'

                est_n2 = str(row.get('estado_n2', '-') or '-').strip().upper()
                if est_n2 in ('NAN', 'NONE', '') or aprob_n2 == '-':
                    est_n2 = '-'

                # Formatear Fecha Turno a DD/MM/YYYY
                raw_fecha = row.get('fecha', '')
                fecha_fmt = ""
                if raw_fecha and str(raw_fecha).strip().lower() not in ('', 'none', 'nan'):
                    f_str = str(raw_fecha).strip()
                    if len(f_str) >= 10 and f_str[4] == '-' and f_str[7] == '-':
                        pts = f_str[:10].split('-')
                        fecha_fmt = f"{pts[2]}/{pts[1]}/{pts[0]}"
                    elif len(f_str) >= 10 and f_str[2] == '/' and f_str[5] == '/':
                        fecha_fmt = f_str[:10]
                    else:
                        try:
                            dt_obj = pd.to_datetime(f_str, errors='coerce')
                            fecha_fmt = dt_obj.strftime('%d/%m/%Y') if pd.notna(dt_obj) else f_str
                        except Exception:
                            fecha_fmt = f_str

                # Formatear Fecha Aprobacion a DD/MM/YYYY HH:MM
                fecha_aprob = ""
                raw_fa = row.get('fecha_aprobacion')
                if raw_fa and str(raw_fa).strip().lower() not in ('', 'none', 'nan'):
                    fa_str = str(raw_fa).strip()
                    if len(fa_str) >= 10 and fa_str[4] == '-' and fa_str[7] == '-':
                        pts = fa_str[:10].split('-')
                        h_part = fa_str[11:16] if len(fa_str) >= 16 else ""
                        fecha_aprob = f"{pts[2]}/{pts[1]}/{pts[0]} {h_part}".strip()
                    else:
                        try:
                            dt_fa = pd.to_datetime(fa_str, errors='coerce')
                            fecha_aprob = dt_fa.strftime('%d/%m/%Y %H:%M') if pd.notna(dt_fa) else fa_str
                        except Exception:
                            fecha_aprob = fa_str

                # Combinar justificación del trabajador y comentarios de Nivel 1 y Nivel 2 en la misma celda
                obs_trab = str(row.get('observacion_trabajador', '') or '').strip()
                c_n1 = str(row.get('comentario_n1', '') or '').strip()
                c_n2 = str(row.get('comentario_n2', '') or '').strip()
                c_sup = str(row.get('comentario_supervisor', '') or '').strip()

                if obs_trab.lower() in ('nan', 'none', ''): obs_trab = ""
                if c_n1.lower() in ('nan', 'none', ''): c_n1 = ""
                if c_n2.lower() in ('nan', 'none', ''): c_n2 = ""
                if c_sup.lower() in ('nan', 'none', ''): c_sup = ""

                comentarios_list = []
                if obs_trab:
                    # Identificar de qué usuario/trabajador proviene el sustento personal
                    dni_val = str(row.get('dni', '') or '').strip().lstrip('0').zfill(8)
                    mapa_dni_usuario = {
                        '47783594': 'jagreda',
                        '47034929': 'jalva',
                        '72559194': 'jdelariva',
                        '46671923': 'jhuayama',
                        '26696602': 'msanchez',
                        '75227437': 'lpretel',
                        '44955960': 'respinoza',
                        '70782038': 'jsanchez',
                    }
                    u_autor = mapa_dni_usuario.get(dni_val)
                    if not u_autor:
                        nom_raw = str(row.get('nombres', '') or '').strip()
                        ape_raw = str(row.get('apellidos', '') or '').strip()
                        if nom_raw:
                            u_autor = nom_raw.split()[0].capitalize()
                        elif ape_raw:
                            u_autor = ape_raw.split()[0].capitalize()
                        else:
                            u_autor = "Personal"

                    if obs_trab.lower().startswith(f"{u_autor.lower()}:"):
                        comentarios_list.append(obs_trab)
                    else:
                        comentarios_list.append(f"{u_autor}: {obs_trab}")

                ap_n1_name = str(row.get('aprobado_por_n1', '') or '').strip()
                est_n1_val = str(row.get('estado_n1', '') or '').strip().upper()
                if c_n1:
                    pref_1 = f"N1 ({ap_n1_name}): " if ap_n1_name and ap_n1_name != '-' else "N1: "
                    comentarios_list.append(f"{pref_1}{c_n1}")
                elif ap_n1_name.lower() == 'admin' and est_n1_val in ('APROBADO', 'RECHAZADO'):
                    comentarios_list.append(f"N1 (admin): {est_n1_val.capitalize()}")

                ap_n2_name = str(row.get('aprobado_por_n2', '') or '').strip()
                est_n2_val = str(row.get('estado_n2', '') or '').strip().upper()
                if c_n2:
                    pref_2 = f"N2 ({ap_n2_name}): " if ap_n2_name and ap_n2_name != '-' else "N2: "
                    comentarios_list.append(f"{pref_2}{c_n2}")
                elif ap_n2_name.lower() == 'admin' and est_n2_val in ('APROBADO', 'RECHAZADO'):
                    comentarios_list.append(f"N2 (admin): {est_n2_val.capitalize()}")

                if not comentarios_list and c_sup:
                    comentarios_list.append(c_sup)

                coment = "\n".join(comentarios_list)

                def _clean_hhmm(val):
                    if val is None or pd.isna(val):
                        return "00:00"
                    v_str = str(val).strip()
                    if v_str.lower() in ('nan', 'none', '', '-'):
                        return "00:00"
                    import re
                    m = re.match(r'^(\d+)\s*h\s*(\d+)\s*m?$', v_str, re.IGNORECASE)
                    if m:
                        return f"{int(m.group(1)):02d}:{int(m.group(2)):02d}"
                    m2 = re.match(r'^(\d+):(\d+)$', v_str)
                    if m2:
                        return f"{int(m2.group(1)):02d}:{int(m2.group(2)):02d}"
                    return v_str

                jornada_fmt = _clean_hhmm(row.get('jornada_trabajada_hhmm', ''))
                he_fmt = _clean_hhmm(row.get('horas_extras_hhmm', '00:00'))
                exceso_fmt = _clean_hhmm(row.get('exceso_jornada_hhmm', '00:00'))

                row_data = [
                    dni_val,
                    quitar_tildes(str(row.get('apellidos', '') or '')),
                    quitar_tildes(str(row.get('nombres', '') or '')),
                    str(row.get('cargo', '') or ''),
                    str(row.get('area', '') or ''),
                    fecha_fmt,
                    str(row.get('turno', '') or ''),
                    str(row.get('entrada', '') or ''),
                    str(row.get('salida', '') or ''),
                    jornada_fmt,
                    he_fmt,
                    exceso_fmt,
                    estado,
                    aprob_n1,
                    est_n1,
                    aprob_n2,
                    est_n2,
                    fecha_aprob,
                    coment,
                ]

                ws.append(row_data)
                r_idx = ws.max_row
                ws.row_dimensions[r_idx].height = 32 if "\n" in coment else 20

                # Estilos por columna de estado:
                def _get_status_style(st_val):
                    s = str(st_val).strip().upper()
                    if s == 'APROBADO':
                        return fill_approved, font_estado_ok
                    elif s == 'RECHAZADO':
                        return fill_rejected, font_estado_no
                    elif s == 'PENDIENTE':
                        return fill_pending, font_estado_pend
                    return PatternFill(), font_data

                fill_13, font_13 = _get_status_style(estado)
                fill_15, font_15 = _get_status_style(est_n1)
                fill_17, font_17 = _get_status_style(est_n2) if est_n2 != '-' else (PatternFill(), font_data)

                for c_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    if c_idx == 13:
                        cell.font = font_13
                        cell.fill = fill_13
                    elif c_idx == 15:
                        cell.font = font_15
                        cell.fill = fill_15
                    elif c_idx == 17:
                        cell.font = font_17
                        cell.fill = fill_17
                    else:
                        cell.font = font_data
                        cell.fill = PatternFill()

                    cell.border = thin_border
                    # Columnas centradas: DNI (1), Fecha Turno (6), Turno (7), Entrada (8), Salida (9),
                    # Horas Trabajadas (10), Horas Extras (11), Exceso Jornada (12), Estado Final (13),
                    # Aprobador N1 (14), Estado N1 (15), Aprobador N2 (16), Estado N2 (17), Fecha Aprobacion (18)
                    if c_idx in (1, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18):
                        cell.alignment = align_center
                    elif c_idx == 19:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = align_left

                    if c_idx == 1:
                        cell.number_format = '@'

        # Anchos de columna
        for c_idx, w in {
            1: 14, 2: 28, 3: 26, 4: 24, 5: 16,
            6: 16, 7: 10, 8: 14, 9: 14,
            10: 18, 11: 20, 12: 22,
            13: 16, 14: 16, 15: 14,
            16: 16, 17: 14, 18: 20, 19: 36
        }.items():
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        ws.auto_filter.ref = ws.dimensions
        os.makedirs(os.path.dirname(os.path.abspath(target_path)), exist_ok=True)
        try:
            wb.save(target_path)
            return True
        except PermissionError:
            print(f"[Aviso] No se pudo sobrescribir '{target_path}' porque esta abierto en Excel. Cierralo para guardar.")
            return False
    except Exception as e:
        print(f"[Error] Error al exportar aprobaciones Excel: {e}")
        import traceback
        traceback.print_exc()
        return False
