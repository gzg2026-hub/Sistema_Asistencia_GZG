import sqlite3
import pandas as pd
import os
import streamlit as st
from typing import Tuple, Dict, Any, Optional

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(ROOT_DIR, "data", "asistencia.db")

def get_connection(db_path: str = DB_PATH):
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    return conn

def init_db(db_path: str = DB_PATH):
    """Inicializa la base de datos con las tablas estructuradas si no existen."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    # 1. Tabla TRABAJADORES
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS trabajadores (
        dni TEXT PRIMARY KEY,
        apellidos TEXT,
        nombres TEXT,
        cargo TEXT,
        area TEXT,
        aprobador_n1 TEXT,
        aprobador_n2 TEXT,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    
    # 2. Tabla MARCACIONES_RAW
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS marcaciones_raw (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        dni TEXT,
        nombre TEXT,
        apellido TEXT,
        cargo TEXT,
        departamento TEXT,
        grupo TEXT,
        fecha TEXT,
        semana TEXT,
        tiempo TEXT,
        tipo_pase TEXT,
        metodo_verificacion TEXT,
        punto_control TEXT,
        archivo_origen TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(dni, fecha, tiempo, tipo_pase) ON CONFLICT REPLACE
    )
    """)
    
    # 3. Tabla ASISTENCIA (Hoja 03_ASISTENCIA)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS asistencia (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT,
        dni TEXT,
        apellidos TEXT,
        nombres TEXT,
        cargo TEXT,
        area TEXT,
        turno TEXT,
        entrada TEXT,
        salida TEXT,
        horas_trabajadas REAL,
        tardanza_min INTEGER,
        salida_anticipada_min INTEGER,
        exceso_jornada_min INTEGER,
        total_horas_adicionales_min INTEGER,
        incidencias TEXT,
        estado_asistencia TEXT,
        observaciones TEXT,
        UNIQUE(fecha, dni) ON CONFLICT REPLACE
    )
    """)
    
    # 4. Tabla HORAS_EXTRA (Hoja 04_HORAS_EXTRA)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS horas_extra (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT,
        dni TEXT,
        apellidos TEXT,
        nombres TEXT,
        cargo TEXT,
        area TEXT,
        turno TEXT,
        inicio_he TEXT,
        fin_he TEXT,
        duracion_min INTEGER,
        observacion TEXT,
        UNIQUE(fecha, dni, inicio_he, fin_he) ON CONFLICT REPLACE
    )
    """)
    
    # 5. Tabla INCIDENCIAS (Hoja 05_INCIDENCIAS)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS incidencias (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT,
        dni TEXT,
        apellidos TEXT,
        nombres TEXT,
        cargo TEXT,
        area TEXT,
        tipo TEXT,
        hora TEXT,
        descripcion TEXT,
        severidad TEXT,
        observacion TEXT
    )
    """)

    # 6. Tabla USUARIOS para RBAC y Autenticación
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        nombre_completo TEXT NOT NULL,
        rol TEXT NOT NULL,
        area_asignada TEXT DEFAULT 'TODAS',
        cargo TEXT DEFAULT '',
        activo INTEGER DEFAULT 1,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    # 7. Tabla APROBACIONES para MVP Móvil (Horas Extras, Excesos de Jornada e Incidencias)
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS aprobaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha TEXT NOT NULL,
        dni TEXT NOT NULL,
        apellidos TEXT,
        nombres TEXT,
        cargo TEXT,
        area TEXT,
        entrada TEXT,
        salida TEXT,
        horas_trabajadas REAL,
        jornada_trabajada_hhmm TEXT,
        horas_extras_min INTEGER DEFAULT 0,
        exceso_jornada_min INTEGER DEFAULT 0,
        horas_extras_hhmm TEXT,
        exceso_jornada_hhmm TEXT,
        motivo TEXT DEFAULT 'Trabajo operativo adicional en turno',
        observacion_trabajador TEXT,
        estado TEXT NOT NULL DEFAULT 'PENDIENTE',
        aprobado_por TEXT,
        fecha_aprobacion TIMESTAMP,
        comentario_supervisor TEXT,
        adjuntos TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(fecha, dni) ON CONFLICT IGNORE
    )
    """)

    # Migración segura de columnas aprobador_n1 y aprobador_n2 en trabajadores
    cols_trab = [row[1] for row in cursor.execute("PRAGMA table_info(trabajadores)").fetchall()]
    if 'aprobador_n1' not in cols_trab:
        cursor.execute("ALTER TABLE trabajadores ADD COLUMN aprobador_n1 TEXT")
    if 'aprobador_n2' not in cols_trab:
        cursor.execute("ALTER TABLE trabajadores ADD COLUMN aprobador_n2 TEXT")

    # Migración segura de columnas de aprobación en 2 niveles para aprobaciones
    cols_aprob = [row[1] for row in cursor.execute("PRAGMA table_info(aprobaciones)").fetchall()]
    if 'aprobador_n1' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN aprobador_n1 TEXT")
    if 'aprobador_n2' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN aprobador_n2 TEXT")
    if 'estado_n1' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN estado_n1 TEXT DEFAULT 'PENDIENTE'")
    if 'aprobado_por_n1' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN aprobado_por_n1 TEXT")
    if 'fecha_n1' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN fecha_n1 TIMESTAMP")
    if 'comentario_n1' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN comentario_n1 TEXT")
    if 'estado_n2' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN estado_n2 TEXT DEFAULT 'PENDIENTE'")
    if 'aprobado_por_n2' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN aprobado_por_n2 TEXT")
    if 'fecha_n2' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN fecha_n2 TIMESTAMP")
    if 'comentario_n2' not in cols_aprob:
        cursor.execute("ALTER TABLE aprobaciones ADD COLUMN comentario_n2 TEXT")

    # Migración segura de columnas de validación en horas_extra e incidencias
    cols_he = [row[1] for row in cursor.execute("PRAGMA table_info(horas_extra)").fetchall()]
    if 'estado_validacion' not in cols_he:
        cursor.execute("ALTER TABLE horas_extra ADD COLUMN estado_validacion TEXT DEFAULT 'PENDIENTE'")
        cursor.execute("ALTER TABLE horas_extra ADD COLUMN validado_por TEXT")
        cursor.execute("ALTER TABLE horas_extra ADD COLUMN fecha_validacion TEXT")
        cursor.execute("ALTER TABLE horas_extra ADD COLUMN observacion_validacion TEXT")

    cols_inc = [row[1] for row in cursor.execute("PRAGMA table_info(incidencias)").fetchall()]
    if 'estado_validacion' not in cols_inc:
        cursor.execute("ALTER TABLE incidencias ADD COLUMN estado_validacion TEXT DEFAULT 'PENDIENTE'")
        cursor.execute("ALTER TABLE incidencias ADD COLUMN validado_por TEXT")
        cursor.execute("ALTER TABLE incidencias ADD COLUMN fecha_validacion TEXT")
        cursor.execute("ALTER TABLE incidencias ADD COLUMN observacion_validacion TEXT")

    conn.commit()
    conn.close()

    # init_db completo sin relectura de Excel en cada ejecución

def clean_dni(val) -> str:
    """Normaliza cualquier DNI respetando estrictamente la Regla Invariante de 8 dígitos (zfill(8))."""
    if pd.isna(val) or val is None or str(val).strip() == '':
        return ''
    digits = ''.join(c for c in str(val) if c.isdigit())
    digits_clean = digits.lstrip('0')
    if not digits_clean:
        return ''
    return digits_clean.zfill(8)

def quitar_tildes(texto: str) -> str:
    if not isinstance(texto, str) or not texto or str(texto).strip().lower() in ('nan', 'none', ''):
        return ""
    replacements = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U',
        'á': 'A', 'é': 'E', 'í': 'I', 'ó': 'O', 'ú': 'U', 'ü': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'à': 'A', 'è': 'E', 'ì': 'I', 'ò': 'O', 'ù': 'U',
    }
    res = str(texto)
    for k, v in replacements.items():
        res = res.replace(k, v)
    return res.strip()

def guardar_trabajadores(df_trabajadores: pd.DataFrame, db_path: str = DB_PATH):
    """Guarda o actualiza la lista de trabajadores maestros normalizando DNIs."""
    if df_trabajadores.empty:
        return
    init_db(db_path)
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    for _, row in df_trabajadores.iterrows():
        dni = clean_dni(row.get('DNI', ''))
        if not dni:
            continue
        n1_val = str(row.get('APROBADOR_N1', row.get('Nivel de Aprobacion 1', ''))).strip() if pd.notna(row.get('APROBADOR_N1', row.get('Nivel de Aprobacion 1', ''))) else ''
        n2_val = str(row.get('APROBADOR_N2', row.get('Nivel de Aprobacion 2', ''))).strip() if pd.notna(row.get('APROBADOR_N2', row.get('Nivel de Aprobacion 2', ''))) else ''
        if n1_val.lower() in ('nan', 'none', ''): n1_val = None
        if n2_val.lower() in ('nan', 'none', ''): n2_val = None

        cursor.execute("""
        INSERT INTO trabajadores (dni, apellidos, nombres, cargo, area, aprobador_n1, aprobador_n2, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(dni) DO UPDATE SET
            apellidos=excluded.apellidos,
            nombres=excluded.nombres,
            cargo=excluded.cargo,
            area=excluded.area,
            aprobador_n1=excluded.aprobador_n1,
            aprobador_n2=excluded.aprobador_n2,
            updated_at=CURRENT_TIMESTAMP
        """, (
            dni,
            quitar_tildes(str(row.get('APELLIDOS', row.get('Apellidos', '')))),
            quitar_tildes(str(row.get('NOMBRES', row.get('Nombres', '')))),
            str(row.get('CARGO', row.get('Posición / Cargo', ''))).strip(),
            str(row.get('AREA', row.get('ÁREA', row.get('Departamento / Área', '')))).strip(),
            n1_val,
            n2_val
        ))
    conn.commit()
    conn.close()

    # Actualizar automáticamente el archivo local Padron_Trabajadores_GZG.xlsx
    try:
        actualizar_excel_padron_trabajadores(db_path)
    except Exception as e_p:
        print(f"Aviso al actualizar Padron_Trabajadores_GZG.xlsx: {e_p}")


def sincronizar_padron_desde_excel(db_path: str = DB_PATH):
    """Lee Padron_Trabajadores_GZG.xlsx e inserta/actualiza todos los trabajadores con sus aprobadores N1 y N2."""
    padron_path = os.path.join(ROOT_DIR, "Padron_Trabajadores_GZG.xlsx")
    if not os.path.exists(padron_path):
        return
    try:
        df_raw = pd.read_excel(padron_path, header=None, dtype=str)
        if len(df_raw) < 3:
            return
        df_data = df_raw.iloc[3:].copy()
        df_data.columns = ['dni', 'apellidos', 'nombres', 'area', 'cargo', 'estado', 'aprobador_n1', 'aprobador_n2']
        df_data = df_data.dropna(subset=['dni'])
        df_data['dni'] = df_data['dni'].astype(str).str.strip().str.lstrip('0').str.zfill(8)
        df_data = df_data[df_data['dni'].str.len() == 8]
        df_data = df_data[df_data['dni'] != '00000DNI']

        conn = get_connection(db_path)
        cursor = conn.cursor()

        # Asegurar columnas
        cols_trab = [row[1] for row in cursor.execute("PRAGMA table_info(trabajadores)").fetchall()]
        if 'aprobador_n1' not in cols_trab:
            cursor.execute("ALTER TABLE trabajadores ADD COLUMN aprobador_n1 TEXT")
        if 'aprobador_n2' not in cols_trab:
            cursor.execute("ALTER TABLE trabajadores ADD COLUMN aprobador_n2 TEXT")

        for _, r in df_data.iterrows():
            dni = r['dni']
            ape = quitar_tildes(str(r.get('apellidos', ''))).strip()
            nom = quitar_tildes(str(r.get('nombres', ''))).strip()
            cargo = str(r.get('cargo', '')).strip()
            area = str(r.get('area', '')).strip()
            n1 = str(r.get('aprobador_n1', '')).strip().lower()
            n2 = str(r.get('aprobador_n2', '')).strip().lower()
            if n1 in ('nan', 'none', ''): n1 = None
            if n2 in ('nan', 'none', ''): n2 = None

            cursor.execute("""
                INSERT INTO trabajadores (dni, apellidos, nombres, cargo, area, aprobador_n1, aprobador_n2, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(dni) DO UPDATE SET
                    apellidos = excluded.apellidos,
                    nombres = excluded.nombres,
                    cargo = excluded.cargo,
                    area = excluded.area,
                    aprobador_n1 = excluded.aprobador_n1,
                    aprobador_n2 = excluded.aprobador_n2,
                    updated_at = CURRENT_TIMESTAMP
            """, (dni, ape, nom, cargo, area, n1, n2))

        conn.commit()

        # Actualizar tabla aprobaciones existente
        for _, r in df_data.iterrows():
            dni = r['dni']
            cargo = str(r.get('cargo', '')).strip()
            area = str(r.get('area', '')).strip()
            n1 = str(r.get('aprobador_n1', '')).strip().lower()
            n2 = str(r.get('aprobador_n2', '')).strip().lower()
            if n1 in ('nan', 'none', ''): n1 = None
            if n2 in ('nan', 'none', ''): n2 = None
            cursor.execute("""
                UPDATE aprobaciones
                SET aprobador_n1 = ?, aprobador_n2 = ?, cargo = ?, area = ?
                WHERE dni = ?
            """, (n1, n2, cargo, area, dni))

        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Aviso] Sincronización Padrón desde Excel: {e}")


def actualizar_excel_padron_trabajadores(db_path: str = DB_PATH):
    """Sincroniza el archivo local Padron_Trabajadores_GZG.xlsx con la base de datos (8 columnas completas)."""
    padron_path = os.path.join(ROOT_DIR, "Padron_Trabajadores_GZG.xlsx")
    conn = get_connection(db_path)
    df_db = pd.read_sql_query("SELECT dni, apellidos, nombres, cargo, area, aprobador_n1, aprobador_n2 FROM trabajadores", conn)
    conn.close()

    if df_db.empty:
        return

    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    df_db['dni'] = df_db['dni'].astype(str).str.strip().str.zfill(8)
    df_db['apellidos'] = df_db['apellidos'].astype(str).apply(quitar_tildes)
    df_db['nombres'] = df_db['nombres'].astype(str).apply(quitar_tildes)
    df_db['cargo'] = df_db['cargo'].astype(str).apply(quitar_tildes)
    df_db['area'] = df_db['area'].astype(str).apply(quitar_tildes)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trabajadores"
    ws.views.sheetView[0].showGridLines = True

    fill_banner = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    font_banner = Font(name="Calibri", size=13, bold=True, color="FFFFFF")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    font_data = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 28
    ws["A1"] = "PADRÓN OFICIAL DE TRABAJADORES Y PERSONAL REGISTRADO - GZG MINERALES"
    ws["A1"].fill = fill_banner
    ws["A1"].font = font_banner
    ws["A1"].alignment = align_center

    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 18
    fill_sub = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_sub = Font(name="Calibri", size=10, italic=True, bold=True, color="1F4E78")
    ws["A2"] = "GZG Minerales | Estado Actualizado de Personal en Biométrico y Sistema"
    ws["A2"].fill = fill_sub
    ws["A2"].font = font_sub
    ws["A2"].alignment = align_center

    headers = ["DNI", "Apellidos", "Nombres", "Departamento / Área", "Posición / Cargo", "Estado en Sistema", "Nivel de Aprobacion 1", "Nivel de Aprobacion 2"]
    ws.row_dimensions[3].height = 25
    ws.append(headers)

    for cell in ws[3]:
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = align_center
        cell.border = thin_border

    for idx, r in df_db.sort_values(by=['apellidos', 'nombres']).iterrows():
        dni = str(r['dni']).strip().zfill(8)
        ape = str(r['apellidos']).strip()
        nom = str(r['nombres']).strip()
        area = str(r['area']).strip()
        cargo = str(r['cargo']).strip()
        estado = "Activo"
        n1 = str(r['aprobador_n1']).strip() if pd.notna(r['aprobador_n1']) and str(r['aprobador_n1']).strip() != 'None' else ""
        n2 = str(r['aprobador_n2']).strip() if pd.notna(r['aprobador_n2']) and str(r['aprobador_n2']).strip() != 'None' else ""

        ws.append([dni, ape, nom, area, cargo, estado, n1, n2])
        c_row = ws.max_row
        ws.row_dimensions[c_row].height = 20

        for c_i in range(1, 9):
            cell = ws.cell(row=c_row, column=c_i)
            cell.font = font_data
            cell.border = thin_border
            if c_i in (1, 6, 7, 8):
                cell.alignment = align_center
                if c_i == 1:
                    cell.number_format = '@'
            else:
                cell.alignment = align_left


    from openpyxl.utils import get_column_letter

    widths = {1: 15, 2: 28, 3: 26, 4: 26, 5: 30, 6: 18, 7: 24, 8: 24}
    for c_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    wb.save(padron_path)


def guardar_marcaciones_raw(df_marcaciones: pd.DataFrame, archivo_origen: str = "", db_path: str = DB_PATH):
    """Inserta transacciones de marcación raw en la base de datos."""
    if df_marcaciones.empty:
        return
    init_db(db_path)
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    # Obtener mapa DNI -> CARGO de la tabla trabajadores si existe
    cargo_map = {}
    try:
        df_trab_map = pd.read_sql_query("SELECT dni, cargo FROM trabajadores", conn)
        for _, tr in df_trab_map.iterrows():
            cargo_map[str(tr['dni']).strip()] = str(tr['cargo']).strip()
    except Exception:
        pass
        
    for _, row in df_marcaciones.iterrows():
        dni = str(row.get('ID', row.get('DNI', ''))).strip()
        fecha = str(row.get('Fecha', '')).strip()
        tiempo = str(row.get('Tiempo', '')).strip()
        tipo_pase = str(row.get('Tipo de pase de tarjeta', '')).strip()
        
        if not dni or not fecha:
            continue
            
        # Limpieza de departamento (extraer texto tras el >)
        dept_val = str(row.get('Departamento', '')).strip()
        if '>' in dept_val:
            dept_val = dept_val.split('>')[-1].strip()
            
        # Obtener cargo del row o del mapa de trabajadores por DNI
        cargo_val = str(row.get('Cargo', row.get('CARGO', ''))).strip()
        if not cargo_val and dni in cargo_map:
            cargo_val = cargo_map[dni]
            
        cursor.execute("""
        INSERT INTO marcaciones_raw (
            dni, nombre, apellido, cargo, departamento, grupo, fecha, semana, tiempo,
            tipo_pase, metodo_verificacion, punto_control, archivo_origen
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(dni, fecha, tiempo, tipo_pase) DO NOTHING
        """, (
            dni,
            quitar_tildes(str(row.get('Nombre', ''))),
            quitar_tildes(str(row.get('Apellido', ''))),
            cargo_val,
            dept_val,
            str(row.get('Grupo de asistencia', '')),
            fecha,
            str(row.get('Semana', '')),
            tiempo,
            tipo_pase,
            str(row.get('Método de verificación', '')),
            str(row.get('Punto de control de asistencia', '')),
            archivo_origen
        ))
    conn.commit()
    conn.close()

def guardar_asistencia_y_reportes(df_asistencia: pd.DataFrame, df_horas_extra: pd.DataFrame, df_incidencias: pd.DataFrame, db_path: str = DB_PATH):
    """Guarda los resultados procesados de asistencia, H.E. e incidencias."""
    init_db(db_path)
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    # 1. Asistencia
    if not df_asistencia.empty:
        for _, r in df_asistencia.iterrows():
            cursor.execute("""
            INSERT INTO asistencia (
                fecha, dni, apellidos, nombres, cargo, area, turno,
                entrada, salida, horas_trabajadas, tardanza_min, salida_anticipada_min,
                exceso_jornada_min, total_horas_adicionales_min, incidencias, estado_asistencia, observaciones
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(fecha, dni) DO UPDATE SET
                apellidos=excluded.apellidos,
                nombres=excluded.nombres,
                cargo=excluded.cargo,
                area=excluded.area,
                turno=excluded.turno,
                entrada=excluded.entrada,
                salida=excluded.salida,
                horas_trabajadas=excluded.horas_trabajadas,
                tardanza_min=excluded.tardanza_min,
                salida_anticipada_min=excluded.salida_anticipada_min,
                exceso_jornada_min=excluded.exceso_jornada_min,
                total_horas_adicionales_min=excluded.total_horas_adicionales_min,
                incidencias=excluded.incidencias,
                estado_asistencia=excluded.estado_asistencia,
                observaciones=excluded.observaciones
            """, (
                str(r.get('FECHA', '')),
                str(r.get('DNI', '')),
                str(r.get('APELLIDOS', '')),
                str(r.get('NOMBRES', '')),
                str(r.get('CARGO', '')),
                str(r.get('ÁREA', r.get('AREA', ''))),
                str(r.get('TURNO', '')),
                r.get('ENTRADA', None),
                r.get('SALIDA', None),
                float(r.get('HORAS TRABAJADAS', r.get('HORAS_TRABAJADAS', 0.0)) or 0.0),
                int(r.get('TARDANZA (MIN)', r.get('TARDANZA_MIN', 0)) or 0),
                int(r.get('SALIDA ANTICIPADA (MIN)', r.get('SALIDA_ANTICIPADA_MIN', 0)) or 0),
                int(r.get('EXCESO JORNADA', r.get('EXCESO_JORNADA_MIN', 0)) or 0),
                int(r.get('TOTAL HORAS ADICIONALES', r.get('TOTAL_HORAS_ADICIONALES_MIN', 0)) or 0),
                str(r.get('INCIDENCIAS', '')),
                str(r.get('ESTADO ASISTENCIA', r.get('ESTADO', ''))),
                str(r.get('OBSERVACIONES', ''))
            ))
            
    # 2. Horas Extra
    if not df_horas_extra.empty:
        for _, r in df_horas_extra.iterrows():
            cursor.execute("""
            INSERT INTO horas_extra (
                fecha, dni, apellidos, nombres, cargo, area, turno, inicio_he, fin_he, duracion_min, observacion
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(fecha, dni, inicio_he, fin_he) DO UPDATE SET
                cargo=excluded.cargo,
                area=excluded.area,
                duracion_min=excluded.duracion_min,
                observacion=excluded.observacion
            """, (
                str(r.get('FECHA', '')),
                str(r.get('DNI', '')),
                str(r.get('APELLIDOS', '')),
                str(r.get('NOMBRES', '')),
                str(r.get('CARGO', '')),
                str(r.get('ÁREA', r.get('AREA', ''))),
                str(r.get('TURNO', '')),
                str(r.get('INICIO H.E.', r.get('INICIO_HE', ''))),
                str(r.get('FIN H.E.', r.get('FIN_HE', ''))),
                int(r.get('DURACIÓN', r.get('DURACION_MIN', 0)) or 0),
                str(r.get('OBSERVACIÓN', r.get('OBSERVACION', '')))
            ))

    # 3. Incidencias
    if not df_incidencias.empty:
        for _, r in df_incidencias.iterrows():
            cursor.execute("""
            INSERT INTO incidencias (
                fecha, dni, apellidos, nombres, cargo, area, tipo, hora, descripcion, severidad, observacion
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(r.get('FECHA', '')),
                str(r.get('DNI', '')),
                str(r.get('APELLIDOS', '')),
                str(r.get('NOMBRES', '')),
                str(r.get('CARGO', '')),
                str(r.get('ÁREA', r.get('AREA', ''))),
                str(r.get('TIPO', '')),
                str(r.get('HORA', '')),
                str(r.get('DESCRIPCIÓN', r.get('DESCRIPCION', ''))),
                str(r.get('SEVERIDAD', '')),
                str(r.get('OBSERVACIÓN', r.get('OBSERVACION', '')))
            ))

    conn.commit()
    conn.close()

def format_hhmm_cell(val, is_hours_float=False) -> str:
    """Convierte minutos enteros u horas flotantes a string HH:MM (ej. 11:51, 00:15)."""
    if pd.isna(val) or val is None or val == "":
        return "00:00"
    val_str = str(val).strip()
    if ":" in val_str and len(val_str.split(":")) == 2:
        parts = val_str.split(":")
        try:
            h = int(parts[0])
            m = int(parts[1])
            return f"{h:02d}:{m:02d}"
        except Exception:
            return val_str
    try:
        num = float(val)
        if num <= 0:
            return "00:00"
        total_min = int(round(num * 60.0)) if is_hours_float else int(round(num))
        h = total_min // 60
        m = total_min % 60
        return f"{h:02d}:{m:02d}"
    except Exception:
        return "00:00"

def format_hhmm_series(series: pd.Series, is_hours_float: bool = False) -> pd.Series:
    """Convierte una Serie a formato HH:MM de forma vectorizada en milisegundos."""
    if series.empty:
        return series
    num_s = pd.to_numeric(series, errors='coerce').fillna(0)
    if is_hours_float:
        tot_min = (num_s * 60.0).round().astype(int)
    else:
        tot_min = num_s.round().astype(int)
    tot_min = tot_min.clip(lower=0)
    hours = (tot_min // 60).astype(str).str.zfill(2)
    mins = (tot_min % 60).astype(str).str.zfill(2)
    return hours + ":" + mins

def sincronizar_excel_madre_a_db(excel_path: str = None, db_path: str = DB_PATH):
    """Sincroniza automáticamente la lista de trabajadores desde la hoja 01_TRABAJADORES del Excel madre a SQLite."""
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    target_files = [excel_path] if excel_path else [
        os.path.join(base_dir, "Sistema_Asistencia_GZG_v1.0.xlsx"),
        os.path.join(base_dir, "Sistema_Asistencia_GZG_v1.0.xlsm")
    ]
    
    for fpath in target_files:
        if fpath and os.path.exists(fpath):
            try:
                df_trab = pd.read_excel(fpath, sheet_name="01_TRABAJADORES")
                if not df_trab.empty and 'DNI' in df_trab.columns:
                    guardar_trabajadores(df_trab, db_path)
                    break
            except Exception:
                pass

@st.cache_data(ttl=300, show_spinner=False)
def obtener_trabajadores_master(db_path: str = DB_PATH) -> pd.DataFrame:
    """Obtiene la lista master de trabajadores rápidamente para selectores de cargos y personal (deduplicada)."""
    init_db(db_path)
    sincronizar_excel_madre_a_db(db_path=db_path)
    conn = get_connection(db_path)
    df = pd.read_sql_query("SELECT dni as DNI, apellidos as APELLIDOS, nombres as NOMBRES, cargo as CARGO, area as AREA FROM trabajadores ORDER BY apellidos, nombres", conn)
    conn.close()
    if not df.empty:
        df['DNI'] = df['DNI'].apply(clean_dni)
        df = df.drop_duplicates(subset=['DNI'], keep='first')
        df = df.sort_values(by=['APELLIDOS', 'NOMBRES'])
    return df

@st.cache_data(ttl=60, show_spinner=False)
def obtener_datos_db(fecha_inicio: Optional[str] = None, fecha_fin: Optional[str] = None, db_path: str = DB_PATH) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Obtiene DataFrames acumulados desde la base de datos:
    (df_trabajadores, df_marcaciones, df_asistencia, df_horas_extra, df_incidencias)
    """
    init_db(db_path)
    sincronizar_desde_hcweb_downloadcenter(db_path)
    conn = get_connection(db_path)
    
    # 1. Trabajadores
    df_trabajadores = pd.read_sql_query("SELECT dni as DNI, apellidos as APELLIDOS, nombres as NOMBRES, cargo as CARGO, area as AREA FROM trabajadores ORDER BY apellidos, nombres", conn)
    
    where_clause = ""
    params = []
    if fecha_inicio and fecha_fin:
        where_clause = " WHERE fecha >= ? AND fecha <= ? "
        params = [fecha_inicio, fecha_fin]
    elif fecha_inicio:
        where_clause = " WHERE fecha >= ? "
        params = [fecha_inicio]
    elif fecha_fin:
        where_clause = " WHERE fecha <= ? "
        params = [fecha_fin]

    # 2. Marcaciones Raw (Orden oficial estricto: ID, Fecha, Nombre, Apellido, Cargo, Departamento...)
    query_mar = f"""
    SELECT COALESCE(t.dni, m.dni) as ID, m.fecha as Fecha, m.nombre as Nombre, m.apellido as Apellido,
           COALESCE(NULLIF(m.cargo, ''), t.cargo, '') as Cargo,
           m.departamento as Departamento, m.grupo as 'Grupo de asistencia',
           m.tiempo as Tiempo, m.tipo_pase as 'Tipo de pase de tarjeta',
           m.metodo_verificacion as 'Método de verificación',
           m.punto_control as 'Punto de control de asistencia'
    FROM marcaciones_raw m
    LEFT JOIN trabajadores t ON LTRIM(m.dni, '0') = LTRIM(t.dni, '0')
    {where_clause.replace('fecha', 'm.fecha') if where_clause else ''}
    ORDER BY m.fecha, m.tiempo
    """
    df_marcaciones = pd.read_sql_query(query_mar, conn, params=params)
    
    # 3. Asistencia
    query_asis = f"""
    SELECT fecha as FECHA, dni as DNI, apellidos as APELLIDOS, nombres as NOMBRES,
           cargo as CARGO, area as ÁREA, turno as TURNO,
           entrada as ENTRADA, salida as SALIDA,
           horas_trabajadas as 'HORAS TRABAJADAS',
           tardanza_min as 'TARDANZA (MIN)',
           salida_anticipada_min as 'SALIDA ANTICIPADA (MIN)',
           exceso_jornada_min as 'EXCESO JORNADA',
           total_horas_adicionales_min as 'TOTAL HORAS ADICIONALES',
           incidencias as INCIDENCIAS, estado_asistencia as 'ESTADO ASISTENCIA', observaciones as OBSERVACIONES
    FROM asistencia {where_clause} ORDER BY fecha, apellidos, nombres
    """
    df_asistencia = pd.read_sql_query(query_asis, conn, params=params)

    if not df_asistencia.empty:
        df_asistencia['HORAS TRABAJADAS (HH:MM)'] = format_hhmm_series(df_asistencia['HORAS TRABAJADAS'], is_hours_float=True)
        df_asistencia['TARDANZA (HH:MM)'] = format_hhmm_series(df_asistencia['TARDANZA (MIN)'], is_hours_float=False)
        df_asistencia['SALIDA ANTICIPADA (HH:MM)'] = format_hhmm_series(df_asistencia['SALIDA ANTICIPADA (MIN)'], is_hours_float=False)
        df_asistencia['EXCESO JORNADA (HH:MM)'] = format_hhmm_series(df_asistencia['EXCESO JORNADA'], is_hours_float=False)
        df_asistencia['TOTAL HORAS ADICIONALES (HH:MM)'] = format_hhmm_series(df_asistencia['TOTAL HORAS ADICIONALES'], is_hours_float=False)
    
    # 4. Horas Extra
    query_he = f"""
    SELECT h.id as ID_REGISTRO, h.fecha as FECHA, h.dni as DNI, h.apellidos as APELLIDOS, h.nombres as NOMBRES,
           COALESCE(NULLIF(h.cargo, ''), t.cargo, '') as CARGO,
           COALESCE(NULLIF(h.area, ''), t.area, '') as ÁREA,
           h.turno as TURNO, h.inicio_he as 'INICIO H.E.', h.fin_he as 'FIN H.E.',
           h.duracion_min as 'DURACIÓN',
           h.observacion as OBSERVACIÓN,
           COALESCE(h.estado_validacion, 'PENDIENTE') as 'ESTADO VALIDADOR',
           COALESCE(h.validado_por, '-') as 'VALIDADO POR',
           COALESCE(h.fecha_validacion, '-') as 'FECHA VALIDACIÓN',
           COALESCE(h.observacion_validacion, '') as 'OBSERVACIÓN VALIDADOR'
    FROM horas_extra h
    LEFT JOIN trabajadores t ON h.dni = t.dni
    {where_clause.replace('fecha', 'h.fecha') if where_clause else ''}
    ORDER BY h.fecha, h.apellidos, h.nombres
    """
    df_horas_extra = pd.read_sql_query(query_he, conn, params=params)

    if not df_horas_extra.empty:
        df_horas_extra['DURACIÓN (HH:MM)'] = format_hhmm_series(df_horas_extra['DURACIÓN'], is_hours_float=False)
    
    # 5. Incidencias
    query_inc = f"""
    SELECT i.id as ID_REGISTRO, i.fecha as FECHA, i.dni as DNI, i.apellidos as APELLIDOS, i.nombres as NOMBRES,
           COALESCE(NULLIF(i.cargo, ''), t.cargo, '') as CARGO,
           COALESCE(NULLIF(i.area, ''), t.area, '') as ÁREA,
           i.tipo as TIPO, i.hora as HORA, i.descripcion as DESCRIPCIÓN,
           i.severidad as SEVERIDAD, i.observacion as OBSERVACIÓN,
           COALESCE(i.estado_validacion, 'PENDIENTE') as 'ESTADO VALIDADOR',
           COALESCE(i.validado_por, '-') as 'VALIDADO POR',
           COALESCE(i.fecha_validacion, '-') as 'FECHA VALIDACIÓN',
           COALESCE(i.observacion_validacion, '') as 'OBSERVACIÓN VALIDADOR'
    FROM incidencias i
    LEFT JOIN trabajadores t ON i.dni = t.dni
    {where_clause.replace('fecha', 'i.fecha') if where_clause else ''}
    ORDER BY i.fecha, i.apellidos, i.nombres
    """
    df_incidencias = pd.read_sql_query(query_inc, conn, params=params)
    
    conn.close()
    return df_trabajadores, df_marcaciones, df_asistencia, df_horas_extra, df_incidencias

# ==============================================================================
# FUNCIONES DE GESTIÓN DE USUARIOS Y AUTENTICACIÓN (RBAC)
# ==============================================================================

def seed_default_users(hash_fn, db_path: str = DB_PATH):
    """Crea los usuarios iniciales autorizados del sistema y elimina usuarios legacy."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    # Eliminar usuarios de muestra no autorizados
    cursor.execute("""
    DELETE FROM usuarios 
    WHERE username IN ('raul.espinoza', 'jhon.alva', 'carlos.mendoza', 'manuel.benitez', 'javier.delariva')
    """)

    users = [
        ('admin', hash_fn('gzg2026*'), 'Administración (Control Total)', 'ADMINISTRADOR', 'TODAS', 'Administrador de Sistema'),
        ('jagreda', hash_fn('jagreda2026*'), 'Jhon Robert Ágreda Aspajo', 'JEFE', 'OPER&MTTO', 'Jefe'),
        ('jalva', hash_fn('jalva2026*'), 'Jhon Kenedy Alva Medina', 'JEFE', 'JEFATURA', 'Jefe'),
        ('jdelariva', hash_fn('jdelariva2026*'), 'Javier Adrián De La Riva Aguilar', 'JEFE', 'JEFATURA', 'Jefe'),
        ('jhuayama', hash_fn('jhuayama2026*'), 'Josmell Waldir Huayama Adriano', 'JEFE', 'OPER&MTTO', 'Jefe'),
        ('msanchez', hash_fn('msanchez2026*'), 'Manuel Ysidoro Sánchez Montoya', 'SUPERINTENDENTE', 'JEFATURA', 'Superintendente')
    ]
    
    for username, pass_hash, nombre, rol, area, cargo in users:
        # Verificar si ya existe
        cursor.execute("SELECT id FROM usuarios WHERE LOWER(username) = LOWER(?)", (username,))
        row = cursor.fetchone()
        if row:
            cursor.execute("""
            UPDATE usuarios SET rol = ?, area_asignada = ?, cargo = ? WHERE id = ?
            """, (rol, area, cargo, row[0]))
        else:
            cursor.execute("""
            INSERT INTO usuarios (username, password_hash, nombre_completo, rol, area_asignada, cargo)
            VALUES (?, ?, ?, ?, ?, ?)
            """, (username, pass_hash, nombre, rol, area, cargo))
        
    conn.commit()
    conn.close()

def cambiar_password_usuario(username: str, new_password_hash: str, db_path: str = DB_PATH) -> bool:
    """Actualiza el password_hash de un usuario en la base de datos."""
    try:
        conn = get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("UPDATE usuarios SET password_hash = ? WHERE LOWER(username) = LOWER(?)", (new_password_hash, username.strip()))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al cambiar contraseña de {username}: {e}")
        return False

def obtener_usuario_by_username(username: str, db_path: str = DB_PATH) -> Optional[Dict[str, Any]]:
    """Obtiene la información de un usuario según su nombre de usuario (búsqueda case-insensitive)."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT id, username, password_hash, nombre_completo, rol, area_asignada, cargo, activo FROM usuarios WHERE LOWER(username) = LOWER(?)", (username.strip(),))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {
            'id': row[0],
            'username': row[1],
            'password_hash': row[2],
            'nombre_completo': row[3],
            'rol': row[4],
            'area_asignada': row[5],
            'cargo': row[6],
            'activo': row[7]
        }
    return None

def obtener_todos_usuarios(db_path: str = DB_PATH) -> pd.DataFrame:
    """Retorna la lista de todos los usuarios registrados."""
    conn = get_connection(db_path)
    df = pd.read_sql_query("SELECT id, username, nombre_completo, rol, area_asignada, cargo, activo, created_at FROM usuarios ORDER BY nombre_completo", conn)
    conn.close()
    return df

def crear_token_sesion(username: str, db_path: str = DB_PATH) -> str:
    """Genera y almacena un token de sesión permanente para la función Recordarme."""
    import secrets
    token = secrets.token_hex(24)
    conn = get_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS user_tokens (
        token TEXT PRIMARY KEY,
        username TEXT NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)
    cursor.execute("INSERT OR REPLACE INTO user_tokens (token, username) VALUES (?, ?)", (token, username.lower().strip()))
    conn.commit()
    conn.close()
    return token

def _normalize_token_str(tok) -> str:
    if isinstance(tok, (list, tuple)):
        tok = tok[0] if len(tok) > 0 else ""
    return str(tok or "").strip()

def validar_token_sesion(token: Any, db_path: str = DB_PATH) -> Optional[Dict[str, Any]]:
    """Valida un token de sesión y devuelve el usuario si es válido (soporta strings y listas de query_params)."""
    t_clean = _normalize_token_str(token)
    if not t_clean:
        return None
    try:
        conn = get_connection(db_path)
        cursor = conn.cursor()
        cursor.execute("""
        CREATE TABLE IF NOT EXISTS user_tokens (
            token TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        """)
        cursor.execute("SELECT username FROM user_tokens WHERE token = ?", (t_clean,))
        row = cursor.fetchone()
        conn.close()
        if row:
            return obtener_usuario_by_username(row[0], db_path)
    except Exception as e:
        print(f"Error al validar token de sesion: {e}")
    return None

def eliminar_token_sesion(token: Any = None, username: str = None, db_path: str = DB_PATH):
    """Elimina tokens de sesión al cerrar sesión manualmente para evitar re-ingresos accidentales."""
    t_clean = _normalize_token_str(token)
    u_clean = str(username or "").strip()
    if not t_clean and not u_clean:
        return
    try:
        conn = get_connection(db_path)
        cursor = conn.cursor()
        if t_clean and u_clean:
            cursor.execute("DELETE FROM user_tokens WHERE token = ? OR LOWER(username) = LOWER(?)", (t_clean, u_clean))
        elif t_clean:
            cursor.execute("DELETE FROM user_tokens WHERE token = ?", (t_clean,))
        elif u_clean:
            cursor.execute("DELETE FROM user_tokens WHERE LOWER(username) = LOWER(?)", (u_clean,))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"Error al eliminar token de sesion: {e}")

def crear_usuario(username: str, password_hash: str, nombre_completo: str, rol: str, area_asignada: str = 'TODAS', cargo: str = '', db_path: str = DB_PATH) -> bool:
    """Crea un nuevo usuario en la base de datos."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    try:
        cursor.execute("""
        INSERT INTO usuarios (username, password_hash, nombre_completo, rol, area_asignada, cargo)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (username.strip(), password_hash, nombre_completo.strip(), rol, area_asignada, cargo.strip()))
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Error al crear usuario: {e}")
        conn.close()
        return False

def eliminar_usuario(user_id: int, db_path: str = DB_PATH) -> bool:
    """Desactiva o elimina un usuario."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM usuarios WHERE id = ?", (user_id,))
    conn.commit()
    conn.close()
    return True

def actualizar_estado_he(he_id: int, nuevo_estado: str, usuario_validador: str, observacion: str = "", db_path: str = DB_PATH) -> bool:
    """Actualiza el estado de validación de un registro de Horas Extra."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    fecha_val = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    cursor.execute("""
    UPDATE horas_extra
    SET estado_validacion = ?, validado_por = ?, fecha_validacion = ?, observacion_validacion = ?
    WHERE id = ?
    """, (nuevo_estado, usuario_validador, fecha_val, observacion, he_id))
    conn.commit()
    conn.close()
    return True

def actualizar_estado_incidencia(inc_id: int, nuevo_estado: str, usuario_validador: str, observacion: str = "", db_path: str = DB_PATH) -> bool:
    """Actualiza el estado de validación de un registro de Incidencias."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    fecha_val = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    cursor.execute("""
    UPDATE incidencias
    SET estado_validacion = ?, validado_por = ?, fecha_validacion = ?, observacion_validacion = ?
    WHERE id = ?
    """, (nuevo_estado, usuario_validador, fecha_val, observacion, inc_id))
    conn.commit()
    conn.close()
    return True


def sincronizar_desde_hcweb_downloadcenter(db_path: str = DB_PATH):
    """Busca automáticamente archivos de transacciones descargados por HikCentral en HCWebControlService\\Downloadcenter e inserta sus marcaciones."""
    import glob
    hcweb_dir = r"C:\Users\GZG Minerales 2026\HCWebControlService\Downloadcenter"
    if not os.path.exists(hcweb_dir):
        return
    pattern = os.path.join(hcweb_dir, "**", "Transacciones_*.xlsx")
    found_files = glob.glob(pattern, recursive=True)
    if not found_files:
        return
    
    found_files.sort(key=os.path.getmtime)
    
    for raw_excel in found_files:
        try:
            df_raw = pd.read_excel(raw_excel)
            if 'Unnamed: 0' in df_raw.columns:
                df_test = pd.read_excel(raw_excel, header=None)
                header_idx = None
                for i in range(min(10, len(df_test))):
                    row_vals = [str(v).strip() for v in df_test.iloc[i].values]
                    if 'ID' in row_vals or 'Nombre' in row_vals or 'Fecha' in row_vals:
                        header_idx = i
                        break
                if header_idx is not None:
                    df_raw = pd.read_excel(raw_excel, header=header_idx)
            
            col_rename = {}
            for c in df_raw.columns:
                c_str = str(c).strip()
                if c_str.lower() in ['id', 'dni', 'nro persona', 'id persona']:
                    col_rename[c] = 'ID'
                elif c_str.lower() in ['nombre', 'nombres']:
                    col_rename[c] = 'Nombre'
                elif c_str.lower() in ['apellido', 'apellidos']:
                    col_rename[c] = 'Apellido'
                elif c_str.lower() in ['fecha']:
                    col_rename[c] = 'Fecha'
                elif c_str.lower() in ['tiempo', 'hora', 'hora marcacion']:
                    col_rename[c] = 'Tiempo'
                elif 'tipo de pase' in c_str.lower():
                    col_rename[c] = 'Tipo de pase de tarjeta'
                elif 'metodo' in c_str.lower() or 'método' in c_str.lower():
                    col_rename[c] = 'Método de verificación'
                elif 'punto de control' in c_str.lower():
                    col_rename[c] = 'Punto de control de asistencia'
                    
            df_raw = df_raw.rename(columns=col_rename)
            guardar_marcaciones_raw(df_raw, archivo_origen=raw_excel, db_path=db_path)
        except Exception as e:
            pass


def sincronizar_aprobaciones_desde_asistencia(db_path: str = DB_PATH):
    """Sincronizar marcaciones con incidencias/HE hacia la tabla aprobaciones asociando N1 y N2."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    
    # 1. Asegurar tabla creada y sincronizar cargos oficiales del Padrón
    init_db(db_path)
    cursor.execute("""
        UPDATE asistencia
        SET cargo = (SELECT t.cargo FROM trabajadores t WHERE t.dni = asistencia.dni),
            area = (SELECT t.area FROM trabajadores t WHERE t.dni = asistencia.dni)
        WHERE EXISTS (SELECT 1 FROM trabajadores t WHERE t.dni = asistencia.dni AND t.cargo IS NOT NULL AND t.cargo != '');
    """)
    cursor.execute("""
        UPDATE aprobaciones
        SET cargo = (SELECT t.cargo FROM trabajadores t WHERE t.dni = aprobaciones.dni),
            area = (SELECT t.area FROM trabajadores t WHERE t.dni = aprobaciones.dni)
        WHERE EXISTS (SELECT 1 FROM trabajadores t WHERE t.dni = aprobaciones.dni AND t.cargo IS NOT NULL AND t.cargo != '');
    """)
    cursor.execute("""
        DELETE FROM aprobaciones 
        WHERE LOWER(cargo) LIKE '%administrativo%' 
           OR dni IN ('74546819', '77134790', '48455175')
           OR ((COALESCE(horas_extras_min, 0) <= 0) AND (COALESCE(exceso_jornada_min, 0) <= 0))
    """)
    
    # 2. Leer registros de asistencia usando el cargo oficial del Padrón (t.cargo)
    cursor.execute("""
        SELECT a.fecha, a.dni, a.apellidos, a.nombres, COALESCE(t.cargo, a.cargo) as cargo, COALESCE(t.area, a.area) as area, a.entrada, a.salida,
               a.horas_trabajadas, a.exceso_jornada_min, a.total_horas_adicionales_min,
               a.observaciones, t.aprobador_n1, t.aprobador_n2
        FROM asistencia a
        LEFT JOIN trabajadores t ON a.dni = t.dni
        WHERE (a.exceso_jornada_min > 0 OR a.total_horas_adicionales_min > 0)
          AND LOWER(COALESCE(t.cargo, a.cargo, '')) NOT LIKE '%administrativo%'
          AND a.dni NOT IN ('74546819', '77134790', '48455175')
    """)
    rows = cursor.fetchall()
    
    for r in rows:
        fecha, dni, apellidos, nombres, cargo, area, entrada, salida, h_trab, exceso_min, total_adic_min, obs, n1_app, n2_app = r
        
        # Formatear HH:MM
        exceso_min = exceso_min or 0
        total_adic_min = total_adic_min or 0
        he_min = total_adic_min - exceso_min if total_adic_min > exceso_min else 0
        
        def _to_hhmm(minutos):
            if not minutos or minutos <= 0:
                return "00:00"
            h = minutos // 60
            m = minutos % 60
            return f"{h:02d}:{m:02d}"
        
        def _htrab_to_hhmm(val):
            if not val:
                return "00:00"
            h = int(val)
            m = int(round((val - h) * 60))
            return f"{h:02d}:{m:02d}"
        
        jornada_str = _htrab_to_hhmm(h_trab)
        he_str = _to_hhmm(he_min)
        exceso_str = _to_hhmm(exceso_min)
        
        cursor.execute("""
            INSERT OR IGNORE INTO aprobaciones (
                fecha, dni, apellidos, nombres, cargo, area, entrada, salida,
                horas_trabajadas, jornada_trabajada_hhmm, horas_extras_min, exceso_jornada_min,
                horas_extras_hhmm, exceso_jornada_hhmm, observacion_trabajador,
                aprobador_n1, aprobador_n2, estado_n1, estado_n2
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PENDIENTE', 'PENDIENTE')
        """, (
            fecha, dni, apellidos, nombres, cargo, area, entrada, salida,
            h_trab, jornada_str, he_min, exceso_min, he_str, exceso_str, obs or '',
            n1_app, n2_app
        ))

        # Actualizar N1 y N2 si ya existía el registro
        cursor.execute("""
            UPDATE aprobaciones
            SET aprobador_n1 = ?, aprobador_n2 = ?
            WHERE fecha = ? AND dni = ?
        """, (n1_app, n2_app, fecha, dni))
        
    conn.commit()
    conn.close()


def sincronizar_aprobaciones_con_gdrive(db_path: str = DB_PATH):
    """
    Sincroniza el estado de aprobaciones desde Google Drive (o Excel local) hacia SQLite.
    Garantiza que cualquier aprobación realizada en la nube persista permanentemente en todos los dispositivos.
    """
    try:
        import os, datetime
        import pandas as pd
        from scripts.gdrive_uploader import descargar_archivo_de_gdrive
        
        root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        mes_str = datetime.date.today().strftime('%Y-%m')
        file_name = f"Aprobaciones_GZG_{mes_str}.xlsx"
        local_path = os.path.join(root_dir, 'downloads', 'data_procesada', file_name)
        
        # Intentar descargar desde Google Drive si existe
        descargar_archivo_de_gdrive(file_name, local_path)
        
        if not os.path.exists(local_path):
            return
            
        df_excel = pd.read_excel(local_path, header=3)
        if df_excel.empty or 'DNI' not in df_excel.columns:
            return
            
        conn = get_connection(db_path)
        cursor = conn.cursor()
        
        for _, r in df_excel.iterrows():
            dni_raw = str(r.get('DNI', '')).split('.')[0].strip()
            if not dni_raw or dni_raw.lower() in ('nan', 'none'):
                continue
            dni = dni_raw.zfill(8)
            fecha_raw = str(r.get('Fecha Turno', '')).strip()
            if not fecha_raw or fecha_raw.lower() in ('nan', 'none'):
                continue
            if '/' in fecha_raw:
                pts = fecha_raw.split('/')
                if len(pts) == 3:
                    fecha_iso = f"{pts[2]}-{pts[1].zfill(2)}-{pts[0].zfill(2)}"
                else:
                    fecha_iso = fecha_raw
            else:
                fecha_iso = fecha_raw[:10]
                
            est_final = str(r.get('Estado Final', '')).strip().upper()
            est_n1 = str(r.get('Estado N1', '')).strip().upper()
            est_n2 = str(r.get('Estado N2', '')).strip().upper()
            coment = str(r.get('Comentario Supervisor', '')).strip()
            
            if est_final in ('APROBADO', 'RECHAZADO', 'PENDIENTE'):
                cursor.execute("""
                    UPDATE aprobaciones
                    SET estado = ?,
                        estado_n1 = CASE WHEN ? IN ('APROBADO', 'RECHAZADO', 'PENDIENTE') THEN ? ELSE estado_n1 END,
                        estado_n2 = CASE WHEN ? IN ('APROBADO', 'RECHAZADO', 'PENDIENTE') THEN ? ELSE estado_n2 END,
                        comentario_supervisor = CASE WHEN ? != '' AND ? NOT IN ('NAN', 'NONE') THEN ? ELSE comentario_supervisor END
                    WHERE fecha = ? AND dni = ?
                """, (
                    est_final,
                    est_n1, est_n1,
                    est_n2, est_n2,
                    coment, coment, coment,
                    fecha_iso, dni
                ))
                
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[Aviso] Sincronizacion Aprobaciones desde Drive: {e}")


def obtener_solicitudes_aprobacion(estado_filter: str = None, db_path: str = DB_PATH) -> pd.DataFrame:
    """Obtener DataFrame de solicitudes de aprobacion directamente desde SQLite sin cache para respuesta instantanea."""
    conn = get_connection(db_path)
    
    query = "SELECT * FROM aprobaciones"
    params = []
    if estado_filter and estado_filter.upper() != 'TODAS':
        query += " WHERE estado = ?"
        params.append(estado_filter.upper())
        
    query += " ORDER BY fecha DESC, id DESC"
    df = pd.read_sql_query(query, conn, params=params)
    conn.close()
    return df


def regenerar_aprobaciones_excel(db_path: str = DB_PATH) -> bool:
    """
    Regenera el Excel oficial de aprobaciones desde SQLite y lo sube inmediatamente a Google Drive en segundo plano.
    Se ejecuta tras cada accion de aprobacion/rechazo en el app movil sin bloquear la UI (0ms latencia).
    Archivo autorizado: Aprobaciones_GZG_YYYY-MM.xlsx (3er archivo autorizado segun GEMINI.md)
    """
    try:
        import threading
        def _tarea_excel_background():
            try:
                import sys, os, datetime
                root_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
                from data.exporter import exportar_aprobaciones_excel
                
                conn = get_connection(db_path)
                df_aprob = pd.read_sql_query("SELECT * FROM aprobaciones ORDER BY fecha DESC, id DESC", conn)
                conn.close()
                
                mes_str = datetime.date.today().strftime('%Y-%m')
                out_path = os.path.join(root_dir, 'downloads', 'data_procesada', f'Aprobaciones_GZG_{mes_str}.xlsx')
                os.makedirs(os.path.dirname(out_path), exist_ok=True)
                
                ok_local = exportar_aprobaciones_excel(df_aprob, out_path)
                
                # Subida inmediata a Google Drive en segundo plano si el guardado local fue exitoso
                if ok_local and out_path and os.path.exists(out_path):
                    try:
                        from scripts.gdrive_uploader import subir_archivo_a_gdrive
                        subir_archivo_a_gdrive(out_path)
                    except Exception as e_drive:
                        print(f"[Aviso] Subida Drive Aprobaciones: {e_drive}")
            except Exception as e_bg:
                print(f"[Aviso] Error en background regenerar Excel: {e_bg}")

        threading.Thread(target=_tarea_excel_background, daemon=True).start()
        return True
    except Exception as e:
        print(f"[Aviso] No se pudo lanzar hilo de Excel de aprobaciones: {e}")
        return False



def actualizar_estado_aprobacion(
    id_solicitud: int,
    nuevo_estado: str,
    aprobado_por: str,
    comentario: str = "",
    adjunto_path: str = None,
    db_path: str = DB_PATH
) -> bool:
    """
    Actualiza inteligentemente la aprobacion detectando si el usuario es Nivel 1 o Nivel 2
    para esta solicitud especifica, y regenera el Excel en segundo plano.
    """
    conn = get_connection(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT aprobador_n1, aprobador_n2 FROM aprobaciones WHERE id = ?", (id_solicitud,))
    row = cursor.fetchone()
    conn.close()
    
    nivel = 1
    if row:
        n1_usr = (row[0] or '').strip().lower()
        n2_usr = (row[1] or '').strip().lower()
        curr_usr = (aprobado_por or '').strip().lower()
        
        if curr_usr == n2_usr:
            nivel = 2
        elif curr_usr == n1_usr:
            nivel = 1
        else:
            nivel = 1

    resultado = actualizar_estado_aprobacion_nivel(
        id_solicitud=id_solicitud,
        nivel=nivel,
        nuevo_estado=nuevo_estado,
        aprobado_por=aprobado_por,
        comentario=comentario,
        adjunto_path=adjunto_path,
        db_path=db_path
    )
    try:
        regenerar_aprobaciones_excel(db_path)
    except Exception:
        pass
    return resultado


def actualizar_estado_aprobacion_nivel(
    id_solicitud: int,
    nivel: int,
    nuevo_estado: str,
    aprobado_por: str,
    comentario: str = "",
    adjunto_path: str = None,
    db_path: str = DB_PATH
) -> bool:
    """Actualiza la aprobación especificando Nivel 1 o Nivel 2, guarda adjuntos y evalúa el estado global."""
    conn = get_connection(db_path)
    cursor = conn.cursor()
    try:
        st_upper = nuevo_estado.upper()
        if nivel == 1:
            cursor.execute("""
                UPDATE aprobaciones
                SET estado_n1 = ?,
                    aprobado_por_n1 = ?,
                    comentario_n1 = ?,
                    adjuntos = COALESCE(?, adjuntos),
                    fecha_n1 = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (st_upper, aprobado_por, comentario, adjunto_path, id_solicitud))
        elif nivel == 2:
            cursor.execute("""
                UPDATE aprobaciones
                SET estado_n2 = ?,
                    aprobado_por_n2 = ?,
                    comentario_n2 = ?,
                    adjuntos = COALESCE(?, adjuntos),
                    fecha_n2 = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (st_upper, aprobado_por, comentario, adjunto_path, id_solicitud))

        # Evaluar estado final global de la aprobación:
        cursor.execute("SELECT estado_n1, estado_n2, aprobador_n2 FROM aprobaciones WHERE id = ?", (id_solicitud,))
        row = cursor.fetchone()
        if row:
            e1, e2, app_n2 = row[0], row[1], row[2]
            # Si cualquiera rechaza -> RECHAZADO
            if e1 == 'RECHAZADO' or e2 == 'RECHAZADO':
                final_state = 'RECHAZADO'
            # Si no hay aprobador_n2 asignado -> el estado global depende de N1
            elif not app_n2 or str(app_n2).strip().lower() in ('nan', 'none', ''):
                final_state = e1 or 'PENDIENTE'
            # Si hay aprobador_n2 -> requiere que ambos N1 y N2 estén APROBADOS
            elif e1 == 'APROBADO' and e2 == 'APROBADO':
                final_state = 'APROBADO'
            else:
                final_state = 'PENDIENTE'

            cursor.execute("""
                UPDATE aprobaciones
                SET estado = ?,
                    aprobado_por = ?,
                    comentario_supervisor = ?,
                    fecha_aprobacion = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (final_state, aprobado_por, comentario, id_solicitud))

        conn.commit()
        conn.close()
        # Invalidar caché de solicitudes y regenerar Excel de aprobaciones automáticamente
        try:
            obtener_solicitudes_aprobacion.clear()
        except Exception:
            pass
        try:
            regenerar_aprobaciones_excel(db_path)
        except Exception as e_excel:
            print(f"[Aviso] Excel de aprobaciones no regenerado: {e_excel}")
        return True
    except Exception as e:
        conn.close()
        print(f"Error actualizando aprobación nivel {nivel}: {e}")
        return False
