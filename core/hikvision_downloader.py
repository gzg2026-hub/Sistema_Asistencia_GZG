"""
hikvision_downloader.py
=======================
Módulo de descarga de transacciones desde HikCentral Access Control / biométricos Hikvision.
Extrae directamente las marcaciones de HikCentral y genera el Excel 1:1 idéntico al reporte web de 11 columnas.
"""

import os
import json
import time
import datetime
import urllib3
from typing import Optional
from openpyxl import Workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_ASISTENCIA_DIR = os.path.join(PROJECT_ROOT, "downloads", "data_cruda")

DIAS_SEMANA = {
    0: "Lunes",
    1: "Martes",
    2: "Miércoles",
    3: "Jueves",
    4: "Viernes",
    5: "Sábado",
    6: "Domingo"
}

# Mapeo verificado 1:1 con la base de datos de HikCentral Access Control
VERIFY_TYPE_MAP = {
    3: "Huella dactilar",
    2: "Imagen de cara",
    1: "Tarjeta",
    4: "Iris",
    5: "Tarjeta + Huella",
    6: "Tarjeta + Rostro"
}

SWIPE_TYPE_MAP = {
    1: "Registro de entrada",
    2: "Registrar salida",
    3: "Inicio de descansos",
    4: "Fin de descansos",
    5: "Inicio de horas extra",
    6: "Fin de horas extra",
    0: "Indefinido"
}


FECHA_INICIO_OFICIAL = "2026-08-17"


def cargar_config_hikvision():
    """Lee la configuración de IP y credenciales desde config_hikvision.json."""
    config_file = os.path.join(PROJECT_ROOT, "config_hikvision.json")
    defaults = {
        "host": "127.0.0.1",
        "port": 443,
        "scheme": "https",
        "username": "admin",
        "password": ""
    }
    if os.path.exists(config_file):
        try:
            with open(config_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                defaults.update(data)
        except Exception:
            pass
    if not defaults.get("password"):
        print("[HikCentral] ADVERTENCIA: no hay password en config_hikvision.json")
    return defaults


def descargar_transacciones_hikvision(
    carpeta_destino: str = DEFAULT_ASISTENCIA_DIR,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    host: Optional[str] = None,
    port: Optional[int] = None,
    username: Optional[str] = None,
    password: Optional[str] = None
) -> str:
    """
    Extrae automáticamente las marcaciones de transacciones desde HikCentral Access Control V2.4
    utilizando Playwright en segundo plano y genera el Excel 1:1 idéntico al reporte web de 11 columnas.
    """
    cfg = cargar_config_hikvision()
    host = host or cfg.get("host", "127.0.0.1")
    port = port or cfg.get("port", 443)
    scheme = cfg.get("scheme", "https")
    username = username or cfg.get("username", "admin")
    password = password or cfg.get("password", "GzG@ACCESO2026")

    os.makedirs(carpeta_destino, exist_ok=True)

    ayer = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")

    if not fecha_inicio:
        fecha_inicio = ayer
    if not fecha_fin:
        fecha_fin = fecha_inicio

    if fecha_inicio < FECHA_INICIO_OFICIAL:
        print(f"[HikCentral] Ajustando fecha_inicio de {fecha_inicio} al piso oficial: {FECHA_INICIO_OFICIAL}")
        fecha_inicio = FECHA_INICIO_OFICIAL
    if fecha_fin < FECHA_INICIO_OFICIAL:
        fecha_fin = FECHA_INICIO_OFICIAL

    filename = f"Transacciones_{fecha_inicio}_{fecha_fin}.xlsx"
    target_path = os.path.join(carpeta_destino, filename)

    base_url = f"{scheme}://{host}:{port}" if port not in (80, 443) else f"{scheme}://{host}"

    print(f"[HikCentral] Extrayendo transacciones del {fecha_inicio} al {fecha_fin}...")
    print(f"[HikCentral] Servidor: {base_url} | Usuario: {username}")

    records = []

    try:
        from playwright.sync_api import sync_playwright

        for intento_global in range(1, 4):
            try:
                with sync_playwright() as p:
                    browser = p.chromium.launch(
                        headless=True,
                        args=["--ignore-certificate-errors", "--no-sandbox"]
                    )
                    context = browser.new_context(ignore_https_errors=True)
                    page = context.new_page()

                    page.goto(f"{base_url}/#/", wait_until="domcontentloaded")
                    time.sleep(3)

                    # Llenar credenciales en componentes Vue (Element UI) forzando eventos input/change/blur
                    page.evaluate(f"""
                        () => {{
                            const u = document.querySelector('#username') || document.querySelector("input[placeholder='Nombre de usuario']");
                            const p = document.querySelector('#password') || document.querySelector("input[type='password']");
                            if (u) {{
                                u.removeAttribute('readonly');
                                u.focus();
                                u.value = '{username}';
                                u.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                u.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                u.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                            }}
                            if (p) {{
                                p.removeAttribute('readonly');
                                p.focus();
                                p.value = '{password}';
                                p.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                p.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                p.dispatchEvent(new Event('blur', {{ bubbles: true }}));
                            }}
                        }}
                    """)

                    time.sleep(2)
                    pass_input = page.locator("#password, input[type='password']").first
                    if pass_input.count() > 0:
                        pass_input.press("Enter")
                    time.sleep(1)

                    login_btn = page.locator(".login-btn, button:has-text('Iniciar'), button:has-text('Log In')").first
                    if login_btn.count() > 0:
                        try:
                            login_btn.click(timeout=3000)
                        except Exception:
                            pass

                    # ── Verificar que la sesión quedó establecida ─────────────
                    # HikCentral V2.4 puede devolver networkidle antes de que la
                    # cookie de sesión esté lista. Si la URL sigue apuntando al
                    # login, la sesión NO está activa y el ISAPI devolverá [].
                    sesion_activa = False
                    for _espera in range(6):          # máx 12 s
                        time.sleep(2)
                        url_actual = page.url.lower()
                        if "login" not in url_actual and "/#/" != url_actual.split("?")[0].rstrip("/")[-3:]:
                            sesion_activa = True
                            break
                        # Intentar login de nuevo si sigue en pantalla de login
                        if _espera == 2:
                            try:
                                login_btn2 = page.locator(".login-btn, button:has-text('Iniciar'), button:has-text('Log In')").first
                                if login_btn2.count() > 0:
                                    login_btn2.click(timeout=2000)
                            except Exception:
                                pass

                    if not sesion_activa:
                        print(f"[HikCentral] Intento {intento_global}/3: sesión no establecida (URL={page.url}). Reintentando ciclo completo...")
                        browser.close()
                        time.sleep(4)
                        continue   # siguiente intento_global

                    try:
                        page.wait_for_load_state("networkidle", timeout=6000)
                    except Exception:
                        time.sleep(2)

                    payload_js = json.dumps({
                        "RecordRequest": {
                            "PageIndex": 1,
                            "PageSize": 5000,
                            "QueryInfo": {
                                "SortInfo": { "SortField": 1, "SortType": 1 },
                                "BeginTime": f"{fecha_inicio}T00:00:00-05:00",
                                "EndTime": f"{fecha_fin}T23:59:59-05:00",
                                "PersonID": [],
                                "PersonCustomFiledID": [],
                                "RecordType": 1
                            }
                        }
                    })

                    for intento in range(1, 4):
                        try:
                            res = page.evaluate(f"""
                                async () => {{
                                    const resp = await fetch('/ISAPI/Bumblebee/AttendancePlugin/V1/Record?MT=GET', {{
                                        method: 'POST',
                                        headers: {{
                                            'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'
                                        }},
                                        body: JSON.stringify({payload_js})
                                    }});
                                    return await resp.json();
                                }}
                            """)
                            records = res.get("ResponseStatus", {}).get("Data", {}).get("OriginalRecord", [])
                            if records and len(records) > 0:
                                break
                            time.sleep(3)
                        except Exception:
                            time.sleep(3)

                    browser.close()
                    if records and len(records) > 0:
                        print(f"[HikCentral] Autenticación exitosa. {len(records)} marcaciones extraídas (intento {intento_global}/3).")
                        break
                    else:
                        print(f"[HikCentral] Intento {intento_global}/3 devolvió 0 marcaciones. Reintentando login...")
                        time.sleep(3)
            except Exception as e_ciclo:
                print(f"[HikCentral] Error en intento {intento_global}/3: {e_ciclo}")
                time.sleep(3)

    except Exception as e:
        print(f"[HikCentral] Error general durante autodescarga: {e}")

    if not records or len(records) == 0:
        print(f"[HikCentral] ADVERTENCIA: No se obtuvieron registros de transacciones para {fecha_inicio} al {fecha_fin}.")
        return ""

    # Generar Excel EXACTO 1:1 idéntico al reporte de 11 columnas de HikCentral Web Client
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    # Estilos elegantes para los encabezados de columna (Azul marino #1F4E78 + Texto blanco #FFFFFF)
    from openpyxl.styles import PatternFill, Font, Alignment

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "ID", "Apellido", "Nombre", "Departamento", "Posición", "Fecha",
        "Semana", "Tiempo", "Tipo de pase de tarjeta",
        "Método de verificación", "Punto de control de asistencia"
    ]
    ws.append(headers)

    # Aplicar formato pastel azul a la fila 1 de encabezados
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for r in records:
        swipe_time = r.get("SwipeTime", "")  # "2026-08-18T07:06:30-05:00"
        fecha_ev = swipe_time[:10] if len(swipe_time) >= 10 else fecha_inicio
        if fecha_ev < FECHA_INICIO_OFICIAL:
            continue
        hora_ev = swipe_time[11:19] if len(swipe_time) >= 19 else "00:00:00"

        info = r.get("PersonInfo", {})
        dni = str(info.get("EmployeeID", ""))
        nombre = info.get("GivenName", "")
        apellido = info.get("FamilyName", "")
        dept = info.get("DepartmentName", "")
        posicion = info.get("Post", "")

        # Día de la semana en español
        try:
            dt_obj = datetime.datetime.strptime(fecha_ev, "%Y-%m-%d")
            semana = DIAS_SEMANA.get(dt_obj.weekday(), "")
        except Exception:
            semana = ""

        # Mapeo exacto 1:1 de Tipo de pase de tarjeta (SwipeType)
        swipe_type_code = r.get("SwipeType", 0)
        tipo_pase = SWIPE_TYPE_MAP.get(swipe_type_code, "Indefinido")

        # Mapeo exacto 1:1 de Método de verificación (VerifyType)
        verify_type_code = r.get("VerifyType", 0)
        metodo = VERIFY_TYPE_MAP.get(verify_type_code, "--")

        door = r.get("DoorName", "Planta_biometrico_wifi-1")

        ws.append([
            dni, apellido, nombre, dept, posicion, fecha_ev,
            semana, hora_ev, tipo_pase, metodo, door
        ])

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    try:
        wb.save(target_path)
    except PermissionError:
        ts = datetime.datetime.now().strftime("%H%M%S")
        target_path = os.path.join(carpeta_destino, f"Transacciones_{fecha_inicio}_{fecha_fin}_{ts}.xlsx")
        wb.save(target_path)
        print(f"[HikCentral] Archivo en uso, guardado como: {target_path}")

    print(f"[HikCentral] Archivo guardado con éxito en: {target_path}")
    return target_path
